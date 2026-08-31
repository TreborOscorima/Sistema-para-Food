"""Helpers para exportaciones a Excel autoexplicativas (openpyxl).

Compartidos por Reportes y Carta. El objetivo: que cualquiera que abra el .xlsx
entienda qué reporte es, de qué empresa y de qué período, sin contexto extra.
"""
from __future__ import annotations


def escribir_encabezado(ws, titulo: str, descripcion: str,
                        meta: list[tuple[str, str]], n_cols: int) -> int:
    """Escribe un encabezado (título + descripción + metadatos) al inicio de la
    hoja y devuelve el número de fila donde debe ir la fila de encabezados de la
    tabla. Las ``ws.append`` posteriores caen debajo.
    """
    from openpyxl.styles import Font

    ncol = max(1, n_cols)
    ws.append([titulo])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ncol)
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14, color="EA580C")
    if descripcion:
        ws.append([descripcion])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ncol)
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=10, color="555555")
    for label, val in meta:
        ws.append([f"{label}:", val])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
    # Fila en blanco separadora. Se escribe "" (no lista vacía) para que
    # ``max_row`` la cuente y el número de fila devuelto sea el correcto.
    ws.append([""])
    return ws.max_row + 1


def autofit_columnas(ws, start_row: int = 1, max_width: int = 48) -> None:
    """Ajusta el ancho de columnas mirando solo las filas de datos (``>= start_row``),
    para que el título/descripción fusionados de arriba no inflen el ancho."""
    from openpyxl.utils import get_column_letter

    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        lengths = [len(str(c.value)) for c in col_cells
                   if c.value is not None and c.row >= start_row]
        ancho = (max(lengths) + 2) if lengths else 10
        ws.column_dimensions[letter].width = min(ancho, max_width)


def formato_moneda(ws, hdr_row: int, cols, simbolo: str = "S/") -> None:
    """Aplica formato de moneda a las columnas ``cols`` (índices 1-based) desde la
    primera fila de datos (``hdr_row + 1``) hasta el final, para que los importes
    se vean como dinero y sean sumables en Excel."""
    fmt = f'"{simbolo}" #,##0.00'
    for col in cols:
        for row in ws.iter_rows(min_row=hdr_row + 1, min_col=col, max_col=col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt


def finalizar_hoja(ws, hdr_row: int, money_cols=None, simbolo: str = "S/",
                   max_width: int = 48) -> None:
    """Da el acabado final a una hoja de datos ya escrita, para que cualquier
    persona (dueño, cajero, contador) la pueda leer y explotar sin retoques:

    - congela la fila de encabezados (``freeze_panes``) para que no se pierda al
      hacer scroll;
    - activa el autofiltro sobre la tabla, para ordenar/filtrar por columna;
    - aplica formato de moneda a ``money_cols`` (índices 1-based);
    - autoajusta el ancho de las columnas.
    """
    from openpyxl.utils import get_column_letter

    if money_cols:
        formato_moneda(ws, hdr_row, money_cols, simbolo)
    # Autofiltro + congelado sobre el rango de la tabla (encabezados + datos).
    if ws.max_row > hdr_row and ws.max_column >= 1:
        ultima_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{hdr_row}:{ultima_col}{ws.max_row}"
    ws.freeze_panes = f"A{hdr_row + 1}"
    autofit_columnas(ws, start_row=hdr_row, max_width=max_width)
