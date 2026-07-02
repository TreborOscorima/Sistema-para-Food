"""Estado global de TUWAYKIFOOD — mozos, caja, cocina, mostrador, carta."""

from __future__ import annotations

import asyncio
import base64
import io
import os
import pathlib
import re
import uuid
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation

import bcrypt as _bcrypt

import reflex as rx
from pydantic import BaseModel
from sqlalchemy import or_
from sqlmodel import select

from app.models.company import Company
from app.models.food import (
    Categoria,
    Cliente,
    ConfigImpresora,
    CuentaCorriente,
    DetallePedido,
    EstadoMesa,
    EstadoPedido,
    EstadoProduccion,
    Insumo,
    Mesa,
    MovimientoCuenta,
    Pedido,
    Producto,
    Promocion,
    RecetaItem,
    RolUsuario,
    TipoPedido,
    TipoPromocion,
    UsuarioFood,
)
from app.services.receipt_service import (
    TicketLine,
    build_print_script,
    generate_cashier_ticket_html,
    generate_kitchen_ticket_html,
)
from app.models.food import PagoPedido
from app.services.pago_service import (
    metodo_pago_resumen,
    registrar_pagos_pedido,
    validar_pagos,
)
from app.states.caja_turno_mixin import CajaTurnoMixin, get_turno_abierto
from app.utils.db import get_session
from app.utils.tenant import set_tenant_context, tenant_bypass

# ─── Constantes de negocio ───────────────────────────────────────────────────
CURRENCY_SYMBOL = "S/"

OPEN_ORDER_STATES = (
    EstadoPedido.BORRADOR.value,
    EstadoPedido.ENVIADO.value,
    EstadoPedido.EN_PREPARACION.value,
    EstadoPedido.LISTO.value,
)

KITCHEN_VISIBLE_STATES = (
    EstadoProduccion.PENDIENTE.value,
    EstadoProduccion.EN_PREPARACION.value,
)

MESA_LABELS = {
    EstadoMesa.LIBRE.value: "Libre",
    EstadoMesa.OCUPADA.value: "Ocupada",
    EstadoMesa.ESPERANDO_CUENTA.value: "Esperando cuenta",
}
MESA_BADGE_BACKGROUNDS = {
    EstadoMesa.LIBRE.value: "rgba(51,65,85,0.5)",
    EstadoMesa.OCUPADA.value: "rgba(234,88,12,0.18)",
    EstadoMesa.ESPERANDO_CUENTA.value: "rgba(245,158,11,0.18)",
}
MESA_BADGE_TEXTS = {
    EstadoMesa.LIBRE.value: "#94A3B8",
    EstadoMesa.OCUPADA.value: "#FDBA74",
    EstadoMesa.ESPERANDO_CUENTA.value: "#FCD34D",
}
MESA_CARD_BACKGROUNDS = {
    EstadoMesa.LIBRE.value: "#1E293B",
    EstadoMesa.OCUPADA.value: "#1E293B",
    EstadoMesa.ESPERANDO_CUENTA.value: "#1E293B",
}
MESA_CARD_BORDERS = {
    EstadoMesa.LIBRE.value: "2px solid #334155",
    EstadoMesa.OCUPADA.value: "2px solid #EA580C",
    EstadoMesa.ESPERANDO_CUENTA.value: "2px solid #F59E0B",
}
READY_ALERT_BORDER = "3px solid #F59E0B"

PRODUCTION_LABELS = {
    EstadoProduccion.PENDIENTE.value: "Pendiente",
    EstadoProduccion.EN_PREPARACION.value: "En preparacion",
    EstadoProduccion.LISTO_PARA_ENTREGAR.value: "Listo para entregar",
    EstadoProduccion.ENTREGADO_AL_CLIENTE.value: "Entregado al cliente",
}
PRODUCTION_BADGE_BACKGROUNDS = {
    EstadoProduccion.PENDIENTE.value: "#F59E0B",
    EstadoProduccion.EN_PREPARACION.value: "#EA580C",
    EstadoProduccion.LISTO_PARA_ENTREGAR.value: "#16A34A",
    EstadoProduccion.ENTREGADO_AL_CLIENTE.value: "#3B82F6",
}
PRODUCTION_BADGE_TEXTS = {
    EstadoProduccion.PENDIENTE.value: "#FEF3C7",
    EstadoProduccion.EN_PREPARACION.value: "#FEF3C7",
    EstadoProduccion.LISTO_PARA_ENTREGAR.value: "#DCFCE7",
    EstadoProduccion.ENTREGADO_AL_CLIENTE.value: "#DBEAFE",
}
KITCHEN_CARD_BACKGROUNDS = {
    EstadoProduccion.PENDIENTE.value: "#0F172A",
    EstadoProduccion.EN_PREPARACION.value: "#0F172A",
}
KITCHEN_CARD_BORDERS = {
    EstadoProduccion.PENDIENTE.value: "#F59E0B",
    EstadoProduccion.EN_PREPARACION.value: "#EA580C",
}
KITCHEN_DEMORADO_MINUTOS = 15
KITCHEN_DEMORADO_COLOR = "#DC2626"
CLIENTE_VIP_VISITAS_MIN = 15

ROLE_HOME_ROUTES: dict[str, str] = {
    RolUsuario.MOZO.value: "/mozos",
    RolUsuario.CAJA.value: "/caja",
    RolUsuario.COCINA.value: "/cocina",
    RolUsuario.ADMIN.value: "/carta",
}
ROLE_ALLOWED_ROUTES: dict[str, set[str]] = {
    "mozos": {RolUsuario.MOZO.value, RolUsuario.ADMIN.value},
    "caja": {RolUsuario.CAJA.value, RolUsuario.ADMIN.value},
    "mostrador": {RolUsuario.CAJA.value, RolUsuario.ADMIN.value},
    "cocina": {RolUsuario.COCINA.value, RolUsuario.ADMIN.value},
    "carta": {RolUsuario.ADMIN.value},
    "reportes": {RolUsuario.ADMIN.value},
    "usuarios": {RolUsuario.ADMIN.value},
    "configuracion": {RolUsuario.ADMIN.value},
}

_ROL_LABELS: dict[str, str] = {
    RolUsuario.ADMIN.value: "Admin",
    RolUsuario.MOZO.value: "Mozo",
    RolUsuario.CAJA.value: "Caja",
    RolUsuario.COCINA.value: "Cocina",
}
_ROL_BADGE_BG: dict[str, str] = {
    RolUsuario.ADMIN.value: "#FFEDD5",
    RolUsuario.MOZO.value: "#DBEAFE",
    RolUsuario.CAJA.value: "#DCFCE7",
    RolUsuario.COCINA.value: "#FEF3C7",
}
_ROL_BADGE_TEXT: dict[str, str] = {
    RolUsuario.ADMIN.value: "#9A3412",
    RolUsuario.MOZO.value: "#1D4ED8",
    RolUsuario.CAJA.value: "#15803D",
    RolUsuario.COCINA.value: "#B45309",
}

# Base URLs leídas del entorno en tiempo de importación. El company_id ya no es
# fijo: FoodState._company_id() lo resuelve por sesión (ver clase FoodState).
_FOOD_BASE_URL: str = os.getenv("FOOD_BASE_URL", "http://localhost:3003").rstrip("/")
_FOOD_API_URL: str = os.getenv("PUBLIC_API_URL", "http://localhost:3004").rstrip("/")


def _utcnow() -> datetime:
    """Datetime UTC naive compatible con columnas MySQL sin TZ."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _hash_pin(plain: str) -> str:
    """Hashea un PIN con bcrypt. Retorna el hash como str."""
    return _bcrypt.hashpw(plain.encode(), _bcrypt.gensalt()).decode()


def _verify_pin(plain: str, hashed: str) -> bool:
    """Verifica un PIN contra su hash bcrypt."""
    try:
        return _bcrypt.checkpw(plain.encode(), hashed.encode())
    except Exception:
        return False


# ─── Helpers puros ───────────────────────────────────────────────────────────

def _slugify(texto: str) -> str:
    """Convierte texto a slug URL-safe (minúsculas, solo alfanumérico y guión)."""
    texto = texto.lower().strip()
    texto = re.sub(r"[áàä]", "a", texto)
    texto = re.sub(r"[éèë]", "e", texto)
    texto = re.sub(r"[íìï]", "i", texto)
    texto = re.sub(r"[óòö]", "o", texto)
    texto = re.sub(r"[úùü]", "u", texto)
    texto = re.sub(r"[ñ]", "n", texto)
    texto = re.sub(r"[^a-z0-9\s-]", "", texto)
    texto = re.sub(r"[\s]+", "-", texto)
    texto = re.sub(r"-+", "-", texto)
    return texto[:80].strip("-") or "mi-restaurante"


def _generar_qr_base64(url: str) -> str:
    try:
        import qrcode
        qr = qrcode.QRCode(version=None, box_size=6, border=3)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#0F172A", back_color="#FFFFFF")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
    except Exception:
        return ""


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    return Decimal(str(value)).quantize(Decimal("0.01"))


def _money_text(value) -> str:
    return f"{CURRENCY_SYMBOL} {_to_decimal(value):.2f}"


_PRODUCTO_EMOJI_KEYWORDS: list[tuple[str, str]] = [
    ("pizza", "🍕"),
    ("hamburgues", "🍔"),
    ("sandwich", "🥪"),
    ("sanguche", "🥪"),
    ("pan ", "🥪"),
    ("pescado", "🐟"),
    ("ceviche", "🐟"),
    ("gallina", "🍗"),
    ("pollo", "🍗"),
    ("brasa", "🍗"),
    ("carne", "🥩"),
    ("lomo", "🥩"),
    ("bife", "🥩"),
    ("milanesa", "🥩"),
    ("chancho", "🥩"),
    ("cerdo", "🥩"),
    ("pasta", "🍝"),
    ("tallarin", "🍝"),
    ("spaghetti", "🍝"),
    ("fideos", "🍝"),
    ("ensalada", "🥗"),
    ("sopa", "🍲"),
    ("caldo", "🍲"),
    ("chupe", "🍲"),
    ("causa", "🥔"),
    ("papa", "🍟"),
    ("arroz", "🍚"),
    ("huevo", "🥚"),
    ("torta", "🍰"),
    ("mazamorra", "🍰"),
    ("helado", "🍨"),
    ("flan", "🍮"),
    ("postre", "🍰"),
    ("cafe", "☕"),
    ("café", "☕"),
    ("cerveza", "🍺"),
    ("vino", "🍷"),
    ("chicha", "🥤"),
    ("agua", "💧"),
    ("gaseosa", "🥤"),
    ("cola", "🥤"),
    ("kola", "🥤"),
    ("jugo", "🥤"),
    ("bebida", "🥤"),
    ("empanada", "🥟"),
    ("wrap", "🫔"),
]

_CATEGORIA_EMOJI_KEYWORDS: list[tuple[str, str]] = [
    ("bebida", "🥤"),
    ("postre", "🍰"),
    ("entrada", "🥗"),
    ("ensalada", "🥗"),
    ("hamburgues", "🍔"),
    ("pizza", "🍕"),
    ("principal", "🍖"),
    ("fondo", "🍖"),
    ("carta", "🍽️"),
]


def _emoji_para_producto(nombre: str) -> str:
    n = nombre.lower()
    for kw, emoji in _PRODUCTO_EMOJI_KEYWORDS:
        if kw in n:
            return emoji
    return "🍽️"


def _emoji_para_categoria(nombre: str) -> str:
    n = nombre.lower()
    for kw, emoji in _CATEGORIA_EMOJI_KEYWORDS:
        if kw in n:
            return emoji
    return "🍽️"


def _parse_positive_price(raw: str) -> Decimal | None:
    try:
        value = Decimal(raw.replace(",", ".").strip())
    except (InvalidOperation, AttributeError):
        return None
    return value.quantize(Decimal("0.01")) if value > 0 else None


def _normalize_pin(raw: str) -> str:
    return "".join(c for c in str(raw) if c.isdigit())[:6]


def _role_home_route(role: str) -> str:
    return ROLE_HOME_ROUTES.get(role, "/login")


def _actor_name(value: str | None) -> str:
    return (value or "").strip()


def _pedido_table_label(pedido: Pedido, mesas: dict) -> str:
    if pedido.mesa_id is None:
        return "Mesa no asignada"
    mesa = mesas.get(pedido.mesa_id)
    if mesa is None:
        return f"Mesa {pedido.mesa_id}"
    return mesa.nombre or f"Mesa {mesa.numero}"


def _pedido_kitchen_label(pedido: Pedido, mesas: dict) -> str:
    if pedido.tipo_pedido == TipoPedido.MOSTRADOR.value:
        return f"Para Llevar - Cliente: {_actor_name(pedido.nombre_cliente) or 'Sin nombre'}"
    return _pedido_table_label(pedido, mesas)


def _pedido_sales_label(pedido: Pedido, mesas: dict) -> str:
    if pedido.tipo_pedido == TipoPedido.MOSTRADOR.value:
        return f"Mostrador ({_actor_name(pedido.nombre_cliente) or 'Sin nombre'})"
    return _pedido_table_label(pedido, mesas)


def _get_open_order(session, mesa_id: int, company_id: int) -> Pedido | None:
    return session.exec(
        select(Pedido).where(
            Pedido.company_id == company_id,
            Pedido.mesa_id == mesa_id,
            Pedido.estado.in_(OPEN_ORDER_STATES),
        ).order_by(Pedido.id.desc())
    ).first()


def _get_unsent_details(session, pedido_id: int) -> list:
    return session.exec(
        select(DetallePedido).where(
            DetallePedido.pedido_id == pedido_id,
            DetallePedido.impreso_cocina.is_(False),
        ).order_by(DetallePedido.id)
    ).all()


def _get_ready_details(session, pedido_id: int) -> list:
    return session.exec(
        select(DetallePedido).where(
            DetallePedido.pedido_id == pedido_id,
            DetallePedido.impreso_cocina.is_(True),
            DetallePedido.estado_produccion == EstadoProduccion.LISTO_PARA_ENTREGAR.value,
        ).order_by(DetallePedido.id)
    ).all()


def _get_not_delivered_details(session, pedido_id: int) -> list:
    return session.exec(
        select(DetallePedido).where(
            DetallePedido.pedido_id == pedido_id,
            DetallePedido.impreso_cocina.is_(True),
            DetallePedido.estado_produccion != EstadoProduccion.ENTREGADO_AL_CLIENTE.value,
        ).order_by(DetallePedido.id)
    ).all()


def _recalculate_order_total(session, pedido: Pedido) -> Decimal:
    detalles = session.exec(
        select(DetallePedido).where(DetallePedido.pedido_id == pedido.id)
    ).all()
    total = sum((_to_decimal(d.subtotal) for d in detalles), Decimal("0.00"))
    pedido.total = total
    pedido.updated_at = _utcnow()
    session.add(pedido)
    return total


def _sync_order_status(session, pedido: Pedido) -> None:
    sent_details = session.exec(
        select(DetallePedido).where(
            DetallePedido.pedido_id == pedido.id,
            DetallePedido.impreso_cocina.is_(True),
        )
    ).all()
    if not sent_details:
        pedido.estado = EstadoPedido.BORRADOR.value
    elif any(d.estado_produccion == EstadoProduccion.LISTO_PARA_ENTREGAR.value for d in sent_details):
        pedido.estado = EstadoPedido.LISTO.value
    elif any(d.estado_produccion == EstadoProduccion.EN_PREPARACION.value for d in sent_details):
        pedido.estado = EstadoPedido.EN_PREPARACION.value
    elif any(d.estado_produccion == EstadoProduccion.PENDIENTE.value for d in sent_details):
        pedido.estado = EstadoPedido.ENVIADO.value
    elif pedido.pagado and all(d.estado_produccion == EstadoProduccion.ENTREGADO_AL_CLIENTE.value for d in sent_details):
        pedido.estado = EstadoPedido.COBRADO.value
    else:
        pedido.estado = EstadoPedido.ENVIADO.value
    pedido.updated_at = _utcnow()
    session.add(pedido)


def _ensure_open_order(session, mesa: Mesa, company_id: int, mozo_id: int | None = None) -> Pedido:
    pedido = _get_open_order(session, mesa.id or 0, company_id)
    if pedido is not None:
        if mozo_id is not None and pedido.mozo_id is None:
            pedido.mozo_id = mozo_id
            pedido.updated_at = _utcnow()
            session.add(pedido)
        return pedido
    pedido = Pedido(
        company_id=company_id,
        mesa_id=mesa.id or 0,
        mozo_id=mozo_id,
        estado=EstadoPedido.BORRADOR.value,
        total=Decimal("0.00"),
    )
    session.add(pedido)
    session.commit()
    session.refresh(pedido)
    return pedido


# ─── ViewModels (Pydantic, serializables a JSON) ─────────────────────────────

class MesaView(BaseModel):
    id: int
    numero: int
    label: str
    nombre: str
    estado: str
    estado_label: str
    badge_bg: str
    badge_text: str
    capacidad: int
    total_abierto: float
    total_abierto_texto: str
    card_bg: str
    card_border: str
    tiene_items_listos: bool
    items_listos_count: int
    items_total_count: int = 0
    tiempo_abierto_texto: str = ""


class PagoStagedView(BaseModel):
    """Pago agregado a la lista del cobro dividido/mixto (aún no persistido)."""

    metodo: str
    metodo_label: str
    monto: float
    monto_texto: str


class CajaItemView(BaseModel):
    producto_nombre: str = ""
    cantidad: int = 0
    precio_unitario_texto: str = ""
    subtotal_texto: str = ""
    notas: str = ""


class MesaAdminView(BaseModel):
    id: int
    numero: int
    nombre: str
    capacidad: int
    activa: bool
    estado: str


class InsumoView(BaseModel):
    id: int = 0
    nombre: str = ""
    unidad: str = ""
    stock_actual: float = 0.0
    stock_minimo: float = 0.0
    activo: bool = True
    bajo_stock: bool = False
    stock_texto: str = ""
    stock_minimo_texto: str = ""


class RecetaItemView(BaseModel):
    id: int = 0
    producto_id: int = 0
    insumo_id: int = 0
    insumo_nombre: str = ""
    insumo_unidad: str = ""
    cantidad: float = 0.0
    cantidad_texto: str = ""


class ClienteView(BaseModel):
    id: int = 0
    nombre: str = ""
    telefono: str = ""
    email: str = ""
    fecha_nac_iso: str = ""
    fecha_nac_texto: str = ""
    notas: str = ""
    puntos: int = 0
    activo: bool = True
    cumple_hoy: bool = False
    cumple_pronto: bool = False
    dias_para_cumple: int = 999
    visitas_count: int = 0
    gastado_texto: str = "S/ 0.00"
    ultima_visita_texto: str = "Sin visitas"
    es_vip: bool = False


class CuentaView(BaseModel):
    id: int = 0
    cliente_id: int = 0
    cliente_nombre: str = ""
    cliente_telefono: str = ""
    saldo_deuda: float = 0.0
    saldo_texto: str = ""
    limite_credito: float = 0.0


class MovimientoView(BaseModel):
    id: int = 0
    tipo: str = ""
    tipo_label: str = ""
    monto: float = 0.0
    monto_texto: str = ""
    descripcion: str = ""
    fecha_texto: str = ""


class PromocionView(BaseModel):
    id: int = 0
    nombre: str = ""
    tipo: str = ""
    tipo_label: str = ""
    valor: float = 0.0
    descripcion: str = ""
    hora_inicio: str = ""
    hora_fin: str = ""
    activa: bool = True
    aplica_ahora: bool = False
    descuento_texto: str = ""
    horario_texto: str = ""


class UsuarioSesion(BaseModel):
    id: int
    nombre: str
    rol: str
    company_id: int


class CompanyOptionView(BaseModel):
    id: int
    name: str
    slug: str
    logo_url: str = ""


class CategoriaView(BaseModel):
    id: int
    nombre: str
    descripcion: str
    orden: int
    activa: bool
    productos_count: int = 0
    emoji: str = "🍽️"


class ProductoView(BaseModel):
    id: int
    categoria_id: int
    categoria_nombre: str
    nombre: str
    descripcion: str
    precio: float
    precio_texto: str
    disponible: bool
    imagen_url: str
    emoji: str = "🍽️"


class CarritoItem(BaseModel):
    producto_id: int
    nombre: str
    cantidad: int
    precio_unitario: float
    subtotal: float
    subtotal_texto: str
    nota: str = ""


class HistorialItem(BaseModel):
    detalle_id: int
    nombre: str
    cantidad: int
    precio_unitario_texto: str
    subtotal_texto: str
    nota: str
    enviado_en_texto: str
    estado_clave: str
    estado_label: str
    estado_bg: str
    estado_color: str
    preparado_por_nombre: str
    puede_entregar: bool
    puede_cancelar: bool


class CocinaTicketView(BaseModel):
    pedido_id: int
    mesa_label: str
    hora_texto: str
    estado_produccion: str
    estado_label: str
    estado_bg: str
    estado_color: str
    mozo_nombre: str
    action_label: str
    accent_bg: str
    accent_border: str
    detalle_ids_csv: str
    items_lines: list[str]
    minutos_texto: str = ""
    demorado: bool = False


class VentaHistorialView(BaseModel):
    pedido_id: int
    mesa_label: str
    total: float
    total_texto: str
    propina: float
    propina_texto: str
    total_con_propina: float
    total_con_propina_texto: str
    metodo_pago: str
    mozo_nombre: str
    cajero_nombre: str


class VentaDetalleItemView(BaseModel):
    nombre: str
    cantidad: int
    precio_unitario_texto: str
    subtotal_texto: str
    notas: str = ""


class TopPlatoView(BaseModel):
    nombre: str
    cantidad: int
    total_generado: float
    total_texto: str


class MostradorEntregaView(BaseModel):
    pedido_id: int
    cliente_nombre: str
    hora_texto: str
    items_lines: list[str]
    items_count: int


class MostradorEntregadoView(BaseModel):
    pedido_id: int
    cliente_nombre: str
    hora_texto: str
    items_resumen: str
    total_texto: str


class UsuarioAdminView(BaseModel):
    id: int
    nombre: str
    rol: str
    rol_label: str
    pin_masked: str
    activo: bool
    badge_bg: str
    badge_text: str
    es_yo: bool


# ─── Helpers de inventario ───────────────────────────────────────────────────

def _validar_stock_para_items(session, items: list[tuple[int, int]], company_id: int) -> list[str]:
    """Devuelve mensajes de error si el stock es insuficiente. Lista vacía = OK.
    items: lista de (producto_id, cantidad)."""
    if not items:
        return []
    producto_ids = list({pid for pid, _ in items})
    recetas = session.exec(
        select(RecetaItem).where(
            RecetaItem.company_id == company_id,
            RecetaItem.producto_id.in_(producto_ids),
        )
    ).all()
    if not recetas:
        return []
    receta_por_producto: dict[int, list] = {}
    for r in recetas:
        receta_por_producto.setdefault(r.producto_id, []).append(r)
    uso_total: dict[int, Decimal] = {}
    for pid, cantidad in items:
        for ri in receta_por_producto.get(pid, []):
            uso = Decimal(str(ri.cantidad)) * cantidad
            uso_total[ri.insumo_id] = uso_total.get(ri.insumo_id, Decimal("0")) + uso
    if not uso_total:
        return []
    insumos = {
        i.id: i
        for i in session.exec(
            select(Insumo).where(
                Insumo.company_id == company_id,
                Insumo.id.in_(list(uso_total.keys())),
            )
        ).all()
    }
    errores: list[str] = []
    for insumo_id, uso in uso_total.items():
        ins = insumos.get(insumo_id)
        if ins is None:
            continue
        stock = Decimal(str(ins.stock_actual))
        if stock < uso:
            errores.append(f"{ins.nombre}: necesario {uso:.2f}, disponible {stock:.2f} {ins.unidad}")
    return errores


def _descontar_stock_por_pedido(session, pedido_id: int, company_id: int) -> None:
    """Descuenta stock de insumos según las recetas de los ítems del pedido."""
    detalles = session.exec(
        select(DetallePedido).where(DetallePedido.pedido_id == pedido_id)
    ).all()
    if not detalles:
        return
    producto_ids = list({d.producto_id for d in detalles})
    recetas = session.exec(
        select(RecetaItem).where(
            RecetaItem.company_id == company_id,
            RecetaItem.producto_id.in_(producto_ids),
        )
    ).all()
    if not recetas:
        return
    receta_por_producto: dict[int, list] = {}
    for r in recetas:
        receta_por_producto.setdefault(r.producto_id, []).append(r)
    insumo_ids = list({r.insumo_id for r in recetas})
    insumos = {
        i.id: i
        for i in session.exec(
            select(Insumo).where(
                Insumo.company_id == company_id,
                Insumo.id.in_(insumo_ids),
            )
        ).all()
    }
    descuentos: dict[int, Decimal] = {}
    for d in detalles:
        for ri in receta_por_producto.get(d.producto_id, []):
            uso = Decimal(str(ri.cantidad)) * d.cantidad
            descuentos[ri.insumo_id] = descuentos.get(ri.insumo_id, Decimal("0")) + uso
    now = _utcnow()
    for insumo_id, uso_total in descuentos.items():
        ins = insumos.get(insumo_id)
        if ins:
            ins.stock_actual = max(Decimal(str(ins.stock_actual)) - uso_total, Decimal("0"))
            ins.updated_at = now
            session.add(ins)


# ─── Estado principal ─────────────────────────────────────────────────────────

class FoodState(CajaTurnoMixin, rx.State):
    """Estado global de la app TUWAYKIFOOD."""

    mesas: list[MesaView] = []
    categorias: list[CategoriaView] = []
    productos: list[ProductoView] = []
    carrito: list[CarritoItem] = []
    mostrador_carrito: list[CarritoItem] = []
    historial_pedido: list[HistorialItem] = []
    tickets_cocina: list[CocinaTicketView] = []
    historial_ventas: list[VentaHistorialView] = []
    pedidos_mostrador_listos: list[MostradorEntregaView] = []
    pedidos_mostrador_entregados: list[MostradorEntregadoView] = []

    mesa_seleccionada_id: int = 0
    mesa_atendida_por_nombre: str = ""
    categoria_activa_id: int = 0
    mostrador_categoria_activa_id: int = 0
    mostrador_cliente_nombre: str = ""
    mostrador_metodo_pago: str = "efectivo"
    ultimo_pedido_id: int = 0
    mensaje: str = ""
    login_error: str = ""
    usuario_actual: UsuarioSesion | None = None
    login_pin_input: str = ""
    login_rol_seleccionado: str = RolUsuario.MOZO.value
    sidebar_collapsed: bool = False

    login_step: str = "restaurant"
    companies_activas: list[CompanyOptionView] = []
    login_selected_company_id: int = 0
    login_selected_company_slug: str = ""

    @rx.var
    def login_selected_company_name(self) -> str:
        c = next((c for c in self.companies_activas if c.id == self.login_selected_company_id), None)
        return c.name if c else ""

    @rx.var
    def login_selected_company_logo(self) -> str:
        c = next((c for c in self.companies_activas if c.id == self.login_selected_company_id), None)
        return c.logo_url if c else ""

    def _company_id(self) -> int:
        """company_id del tenant activo — de la sesión logueada, o del restaurante
        elegido en el paso previo al login. Arma el contexto de aislamiento tenant."""
        company_id = (
            self.usuario_actual.company_id
            if self.usuario_actual is not None
            else self.login_selected_company_id
        )
        set_tenant_context(company_id, None)
        return company_id

    @contextmanager
    def _tenant_session(self):
        """Como get_session(), pero arma el contexto tenant y evita que el listener
        de tuwayki_core dispare un RuntimeError — el filtro real sigue siendo
        explícito en cada query vía self._company_id()."""
        self._company_id()
        with get_session() as session:
            session.info["tenant_bypass"] = True
            yield session

    categoria_form_id: int = 0
    categoria_form_nombre: str = ""
    categoria_form_descripcion: str = ""
    categoria_form_orden: str = "1"

    producto_form_id: int = 0
    producto_form_categoria_nombre: str = ""
    producto_form_nombre: str = ""
    producto_form_descripcion: str = ""
    producto_form_precio: str = ""
    producto_form_disponible: bool = True
    producto_form_imagen_url: str = ""
    producto_form_emoji: str = ""

    mozos_polling_enabled: bool = False
    cocina_polling_enabled: bool = False
    caja_polling_enabled: bool = False
    mostrador_polling_enabled: bool = False

    usuarios_admin: list[UsuarioAdminView] = []
    usuario_form_id: int = 0
    usuario_form_nombre: str = ""
    usuario_form_rol: str = RolUsuario.MOZO.value
    usuario_form_pin: str = ""
    usuario_form_pin_confirm: str = ""
    usuario_form_activo: bool = True
    usuario_form_visible: bool = False

    mozos_tab_activa: str = "salon"
    nota_producto_activo_id: int = 0
    nota_input_temporal: str = ""

    # Caja — flujo de cobro con método de pago
    caja_cobro_mesa_id: int = 0
    caja_cobro_metodo: str = "efectivo"
    caja_cobro_propina: str = ""
    caja_cobro_descuento: str = ""
    caja_cobro_efectivo_recibido: str = ""
    caja_cobro_error: str = ""
    caja_cobro_items: list[CajaItemView] = []

    # Caja — cobro dividido / pago mixto
    caja_cobro_dividido: bool = False
    caja_pago_staged_metodo: str = "efectivo"
    caja_pago_staged_monto: str = ""
    caja_pagos_staged: list[PagoStagedView] = []

    # Dashboard KPIs
    dashboard_ventas_hoy_texto: str = "S/ 0.00"
    dashboard_pedidos_hoy: int = 0
    dashboard_mesas_ocupadas: int = 0
    dashboard_propina_hoy_texto: str = "S/ 0.00"
    dashboard_ticket_promedio_texto: str = "S/ 0.00"
    dashboard_top_platos: list[TopPlatoView] = []
    dashboard_ventas_trend_pct: int = 0
    dashboard_pedidos_trend: int = 0
    dashboard_ticket_trend_pct: int = 0
    dashboard_propina_trend_pct: int = 0

    # Historial — filtros
    historial_filtro_fecha_desde: str = ""
    historial_filtro_fecha_hasta: str = ""
    historial_filtro_metodo: str = ""
    historial_filtro_rapido: str = "hoy"
    historial_pagina: int = 0
    historial_total: int = 0
    _HISTORIAL_PAGE_SIZE: int = 50

    # Historial — detalle de venta (modal)
    venta_detalle_visible: bool = False
    venta_detalle_pedido_id: int = 0
    venta_detalle_mesa_label: str = ""
    venta_detalle_metodo: str = ""
    venta_detalle_mozo: str = ""
    venta_detalle_cajero: str = ""
    venta_detalle_total_texto: str = ""
    venta_detalle_propina_texto: str = ""
    venta_detalle_items: list[VentaDetalleItemView] = []

    # Nota global del pedido de mesa activo
    nota_pedido_mesa: str = ""

    # Configuración impresoras + carta digital
    config_nombre_local: str = "Mi Restaurante"
    config_logo_url: str = ""
    config_ticket_paper_width_mm: str = "80"
    config_slug: str = "mi-restaurante"
    config_menu_qr_base64: str = ""
    config_menu_url: str = ""
    config_admin_email: str = ""
    config_admin_password_nueva: str = ""
    config_admin_password_confirm: str = ""
    config_admin_show_password: bool = False

    # CRUD mesas (admin)
    mesas_config: list[MesaAdminView] = []
    mesa_config_form_id: int = 0
    mesa_config_form_numero: str = ""
    mesa_config_form_nombre: str = ""
    mesa_config_form_capacidad: str = "4"

    # Inventario de insumos / recetas
    inv_insumos: list[InsumoView] = []
    inv_alertas_bajo_stock: list[str] = []
    inv_form_id: int = 0
    inv_form_nombre: str = ""
    inv_form_unidad: str = "unidad"
    inv_form_stock_actual: str = ""
    inv_form_stock_minimo: str = ""
    inv_form_editando: bool = False
    inv_form_visible: bool = False
    inv_search: str = ""
    inv_producto_sel_nombre: str = ""
    inv_producto_sel_id: int = 0
    inv_receta_items: list[RecetaItemView] = []
    inv_receta_insumo_sel_nombre: str = ""
    inv_receta_cantidad: str = ""

    # Clientes
    clientes_lista: list[ClienteView] = []
    cli_busqueda: str = ""
    cli_form_id: int = 0
    cli_form_nombre: str = ""
    cli_form_telefono: str = ""
    cli_form_email: str = ""
    cli_form_fecha_nac: str = ""
    cli_form_notas: str = ""
    cli_form_editando: bool = False
    cli_form_visible: bool = False
    caja_cobro_cliente_nombre: str = ""
    caja_cobro_cliente_id: int = 0

    # Cuentas corrientes
    cuentas_lista: list[CuentaView] = []
    cuenta_sel_id: int = 0
    cuenta_movimientos: list[MovimientoView] = []
    cc_pago_monto: str = ""
    cc_pago_descripcion: str = ""
    cc_cliente_sel_nombre: str = ""

    # Promociones
    promociones_lista: list[PromocionView] = []
    promo_form_id: int = 0
    promo_form_nombre: str = ""
    promo_form_tipo: str = TipoPromocion.PORCENTAJE.value
    promo_form_valor: str = ""
    promo_form_descripcion: str = ""
    promo_form_hora_inicio: str = ""
    promo_form_hora_fin: str = ""
    promo_form_editando: bool = False
    promo_form_visible: bool = False

    # ─── Computed vars ────────────────────────────────────────────────────────

    @rx.var
    def historial_ventas_recientes(self) -> list[VentaHistorialView]:
        return self.historial_ventas[:5]

    @rx.var
    def inv_insumos_filtrados(self) -> list[InsumoView]:
        q = self.inv_search.strip().lower()
        if not q:
            return self.inv_insumos
        return [i for i in self.inv_insumos if q in i.nombre.lower()]

    def set_inv_search(self, v: str) -> None:
        self.inv_search = v

    def set_inv_form_visible(self, v: bool) -> None:
        self.inv_form_visible = v

    def abrir_nuevo_insumo(self) -> None:
        self.inv_form_id = 0
        self.inv_form_nombre = ""
        self.inv_form_unidad = "unidad"
        self.inv_form_stock_actual = ""
        self.inv_form_stock_minimo = ""
        self.inv_form_editando = False
        self.inv_form_visible = True

    @rx.var
    def autenticado(self) -> bool:
        return self.usuario_actual is not None

    @rx.var
    def usuario_nombre(self) -> str:
        return self.usuario_actual.nombre if self.usuario_actual else ""

    @rx.var
    def usuario_rol(self) -> str:
        return self.usuario_actual.rol if self.usuario_actual else ""

    @rx.var
    def usuario_home_route(self) -> str:
        if self.usuario_actual is None:
            return "/login"
        return _role_home_route(self.usuario_actual.rol)

    @rx.var
    def puede_ver_mozos(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["mozos"]

    @rx.var
    def puede_ver_caja(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["caja"]

    @rx.var
    def puede_ver_mostrador(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["mostrador"]

    @rx.var
    def puede_ver_cocina(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["cocina"]

    @rx.var
    def puede_ver_carta(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["carta"]

    @rx.var
    def puede_ver_reportes(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["reportes"]

    @rx.var
    def puede_ver_configuracion(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["configuracion"]

    @rx.var
    def puede_ver_panel_admin(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol == RolUsuario.ADMIN.value

    @rx.var
    def es_pagina_standalone(self) -> bool:
        """True si esta página (Cuentas/Promociones/Clientes/Inventario) se
        accede como ruta independiente — False si está embebida como pestaña
        dentro de /admin, donde el link "Panel Administrativo" es redundante."""
        return self.router.page.path != "/admin"

    @rx.var
    def historial_filtro_activo(self) -> bool:
        return bool(
            self.historial_filtro_fecha_desde
            or self.historial_filtro_fecha_hasta
            or self.historial_filtro_metodo
        )

    @rx.var
    def historial_tiene_anterior(self) -> bool:
        return self.historial_pagina > 0

    @rx.var
    def historial_tiene_siguiente(self) -> bool:
        return (self.historial_pagina + 1) * self._HISTORIAL_PAGE_SIZE < self.historial_total

    @rx.var
    def historial_pagina_label(self) -> str:
        if self.historial_total == 0:
            return "Sin resultados"
        desde = self.historial_pagina * self._HISTORIAL_PAGE_SIZE + 1
        hasta = min((self.historial_pagina + 1) * self._HISTORIAL_PAGE_SIZE, self.historial_total)
        return f"{desde}–{hasta} de {self.historial_total}"

    @rx.var
    def puede_ver_usuarios(self) -> bool:
        if self.usuario_actual is None:
            return False
        return self.usuario_actual.rol in ROLE_ALLOWED_ROUTES["usuarios"]

    @rx.var
    def roles_disponibles(self) -> list[str]:
        return [r.value for r in RolUsuario]

    @rx.var
    def usuario_form_es_edicion(self) -> bool:
        return self.usuario_form_id > 0

    @rx.var
    def mesa_seleccionada_label(self) -> str:
        mesa = next((m for m in self.mesas if m.id == self.mesa_seleccionada_id), None)
        return mesa.nombre if mesa else "Sin mesa"

    @rx.var
    def mesa_seleccionada_total_texto(self) -> str:
        mesa = next((m for m in self.mesas if m.id == self.mesa_seleccionada_id), None)
        return mesa.total_abierto_texto if mesa else _money_text(0)

    @rx.var
    def cantidad_items_carrito(self) -> int:
        return sum(item.cantidad for item in self.carrito)

    @rx.var
    def total_carrito_texto(self) -> str:
        total = sum(_to_decimal(item.subtotal) for item in self.carrito)
        return _money_text(total)

    @rx.var
    def hay_historial_pedido(self) -> bool:
        return len(self.historial_pedido) > 0

    @rx.var
    def cantidad_mesas_abiertas(self) -> int:
        return sum(1 for m in self.mesas if m.estado != EstadoMesa.LIBRE.value)

    @rx.var
    def mesas_por_cobrar(self) -> list[MesaView]:
        return [m for m in self.mesas if m.estado != EstadoMesa.LIBRE.value and m.total_abierto > 0]

    @rx.var
    def tickets_nuevos(self) -> list[CocinaTicketView]:
        return [t for t in self.tickets_cocina if t.estado_produccion == EstadoProduccion.PENDIENTE.value]

    @rx.var
    def tickets_en_preparacion(self) -> list[CocinaTicketView]:
        return [t for t in self.tickets_cocina if t.estado_produccion == EstadoProduccion.EN_PREPARACION.value]

    @rx.var
    def cantidad_tickets_nuevos(self) -> int:
        return len(self.tickets_nuevos)

    @rx.var
    def cantidad_tickets_en_preparacion(self) -> int:
        return len(self.tickets_en_preparacion)

    @rx.var
    def mesas_con_alerta_entrega(self) -> int:
        return sum(1 for m in self.mesas if m.tiene_items_listos)

    @rx.var
    def categorias_activas(self) -> list[CategoriaView]:
        return [c for c in self.categorias if c.activa]

    @rx.var
    def categorias_activas_nombres(self) -> list[str]:
        return [c.nombre for c in self.categorias if c.activa]

    @rx.var
    def inv_insumos_activos_nombres(self) -> list[str]:
        return [i.nombre for i in self.inv_insumos if i.activo]

    @rx.var
    def inv_productos_nombres(self) -> list[str]:
        return [p.nombre for p in self.productos]

    @rx.var
    def inv_alertas_bajo_stock_texto(self) -> str:
        return ", ".join(self.inv_alertas_bajo_stock)

    @rx.var
    def clientes_filtrados(self) -> list[ClienteView]:
        q = self.cli_busqueda.lower().strip()
        if not q:
            return self.clientes_lista
        return [c for c in self.clientes_lista if q in c.nombre.lower() or q in c.telefono]

    @rx.var
    def clientes_activos_nombres(self) -> list[str]:
        return [c.nombre for c in self.clientes_lista if c.activo]

    @rx.var
    def clientes_cumpleanos_hoy(self) -> list[ClienteView]:
        return [c for c in self.clientes_lista if c.cumple_hoy]

    @rx.var
    def clientes_cumpleanos_pronto(self) -> list[ClienteView]:
        return [c for c in self.clientes_lista if c.cumple_pronto and not c.cumple_hoy]

    @rx.var
    def caja_cobro_es_fiado(self) -> bool:
        return self.caja_cobro_metodo == "fiado"

    @rx.var
    def cuentas_con_deuda(self) -> list[CuentaView]:
        return [c for c in self.cuentas_lista if c.saldo_deuda > 0]

    @rx.var
    def cuentas_total_deuda_texto(self) -> str:
        total = sum(Decimal(str(c.saldo_deuda)) for c in self.cuentas_lista)
        return _money_text(total)

    @rx.var
    def cuenta_sel_nombre(self) -> str:
        c = next((x for x in self.cuentas_lista if x.id == self.cuenta_sel_id), None)
        return c.cliente_nombre if c else ""

    @rx.var
    def cuenta_sel_saldo(self) -> str:
        c = next((x for x in self.cuentas_lista if x.id == self.cuenta_sel_id), None)
        return c.saldo_texto if c else "S/ 0.00"

    @rx.var
    def hay_promo_activa(self) -> bool:
        return any(p.aplica_ahora and p.activa for p in self.promociones_lista)

    @rx.var
    def promo_activa_nombre(self) -> str:
        for p in self.promociones_lista:
            if p.aplica_ahora and p.activa:
                return p.nombre
        return ""

    @rx.var
    def promo_activa_descuento_texto(self) -> str:
        for p in self.promociones_lista:
            if p.aplica_ahora and p.activa:
                return p.descuento_texto
        return ""

    @rx.var
    def promo_activa_descuento_sugerido(self) -> float:
        for p in self.promociones_lista:
            if p.aplica_ahora and p.activa:
                if p.tipo == TipoPromocion.PORCENTAJE.value or p.tipo == TipoPromocion.HAPPY_HOUR.value:
                    return round(self.caja_cobro_total_base * p.valor / 100, 2)
                elif p.tipo == TipoPromocion.MONTO_FIJO.value:
                    return p.valor
        return 0.0

    @rx.var
    def tipos_promo_disponibles(self) -> list[str]:
        return [t.value for t in TipoPromocion]

    @rx.var
    def productos_filtrados(self) -> list[ProductoView]:
        if self.categoria_activa_id == 0:
            return [p for p in self.productos if p.disponible]
        return [p for p in self.productos if p.disponible and p.categoria_id == self.categoria_activa_id]

    @rx.var
    def mostrador_productos_filtrados(self) -> list[ProductoView]:
        if self.mostrador_categoria_activa_id == 0:
            return [p for p in self.productos if p.disponible]
        return [p for p in self.productos if p.disponible and p.categoria_id == self.mostrador_categoria_activa_id]

    @rx.var
    def total_mostrador_texto(self) -> str:
        total = sum(_to_decimal(item.subtotal) for item in self.mostrador_carrito)
        return _money_text(total)

    @rx.var
    def caja_cobro_activo(self) -> bool:
        return self.caja_cobro_mesa_id > 0

    @rx.var
    def caja_cobro_mesa_nombre(self) -> str:
        mesa = next((m for m in self.mesas if m.id == self.caja_cobro_mesa_id), None)
        return mesa.nombre if mesa else ""

    @rx.var
    def caja_cobro_total_base(self) -> float:
        mesa = next((m for m in self.mesas if m.id == self.caja_cobro_mesa_id), None)
        return mesa.total_abierto if mesa else 0.0

    @rx.var
    def caja_cobro_total_base_texto(self) -> str:
        return _money_text(self.caja_cobro_total_base)

    @rx.var
    def caja_cobro_propina_decimal(self) -> float:
        try:
            v = float(self.caja_cobro_propina.replace(",", ".").strip())
            return round(v, 2) if v >= 0 else 0.0
        except (ValueError, AttributeError):
            return 0.0

    @rx.var
    def caja_cobro_descuento_decimal(self) -> float:
        try:
            v = float(self.caja_cobro_descuento.replace(",", ".").strip())
            return round(max(v, 0.0), 2)
        except (ValueError, AttributeError):
            return 0.0

    @rx.var
    def caja_cobro_total_final(self) -> float:
        total = self.caja_cobro_total_base - self.caja_cobro_descuento_decimal + self.caja_cobro_propina_decimal
        return round(max(total, 0.0), 2)

    @rx.var
    def caja_cobro_total_final_texto(self) -> str:
        return _money_text(self.caja_cobro_total_final)

    @rx.var
    def caja_cobro_vuelto(self) -> float:
        try:
            recibido = float(self.caja_cobro_efectivo_recibido.replace(",", ".").strip())
            vuelto = round(recibido - self.caja_cobro_total_final, 2)
            return vuelto if vuelto >= 0 else 0.0
        except (ValueError, AttributeError):
            return 0.0

    @rx.var
    def caja_cobro_vuelto_texto(self) -> str:
        return _money_text(self.caja_cobro_vuelto)

    @rx.var
    def caja_cobro_es_efectivo(self) -> bool:
        return self.caja_cobro_metodo == "efectivo"

    # ─── Inicialización ───────────────────────────────────────────────────────

    def cargar_datos_iniciales(self) -> None:
        self.cargar_mesas()
        self.cargar_menu()
        self.cargar_cocina()
        self._bootstrap_forms()
        if self.mesa_seleccionada_id:
            self._cargar_carrito_mesa(self.mesa_seleccionada_id)
            self._cargar_historial_mesa(self.mesa_seleccionada_id)

    def _bootstrap_forms(self) -> None:
        if not self.producto_form_categoria_nombre and self.categorias:
            self.producto_form_categoria_nombre = self.categorias[0].nombre
        if self.categoria_form_orden == "1" and self.categorias:
            self.categoria_form_orden = str(len(self.categorias) + 1)

    def refrescar(self) -> None:
        self.cargar_datos_iniciales()
        self.cargar_pedidos_mostrador_listos()
        self.cargar_pedidos_mostrador_entregados()
        self.mensaje = "Datos actualizados."

    def _clear_operational_context(self) -> None:
        self.mesas = []
        self.categorias = []
        self.productos = []
        self.carrito = []
        self.mostrador_carrito = []
        self.historial_pedido = []
        self.tickets_cocina = []
        self.historial_ventas = []
        self.pedidos_mostrador_listos = []
        self.pedidos_mostrador_entregados = []
        self.mesa_seleccionada_id = 0
        self.mesa_atendida_por_nombre = ""
        self.categoria_activa_id = 0
        self.mostrador_categoria_activa_id = 0
        self.mostrador_cliente_nombre = ""
        self.mostrador_metodo_pago = "efectivo"
        self.ultimo_pedido_id = 0
        self.mensaje = ""
        self.login_pin_input = ""
        self.sidebar_collapsed = False
        self.mozos_polling_enabled = False
        self.cocina_polling_enabled = False
        self.caja_polling_enabled = False
        self.mostrador_polling_enabled = False
        self.usuarios_admin = []
        self._limpiar_usuario_form()
        self.caja_cobro_mesa_id = 0
        self.caja_cobro_metodo = "efectivo"
        self.caja_cobro_propina = ""
        self.caja_cobro_efectivo_recibido = ""
        self.nota_pedido_mesa = ""

    # ─── Navegación / Shell ───────────────────────────────────────────────────

    def toggle_sidebar(self) -> None:
        self.sidebar_collapsed = not self.sidebar_collapsed

    def _route_access_result(self, route_key: str):
        if self.usuario_actual is None:
            return rx.redirect("/login", replace=True)
        if self.usuario_actual.rol not in ROLE_ALLOWED_ROUTES[route_key]:
            return [
                rx.window_alert("No tienes permisos para este modulo."),
                rx.redirect(self.usuario_home_route, replace=True),
            ]
        self.cargar_datos_iniciales()
        return None

    def on_load_root(self):
        if self.usuario_actual is not None:
            return rx.redirect(self.usuario_home_route, replace=True)
        # Usuario no autenticado: el componente index() ya hace window.location.href='/login'

    def on_load_login(self):
        if self.usuario_actual is not None:
            return rx.redirect(self.usuario_home_route, replace=True)
        self.login_pin_input = ""
        self.login_error = ""
        with tenant_bypass():
            with get_session() as session:
                empresas = session.exec(
                    select(Company)
                    .where(Company.is_active.is_(True))
                    .order_by(Company.name)
                ).all()
        self.companies_activas = [
            CompanyOptionView(id=c.id or 0, name=c.name, slug=c.slug,
                               logo_url=c.logo_url or "")
            for c in empresas
        ]
        # Si venimos de "Ingresar como Administrador" con el link "¿Sos
        # empleado?", el slug en la URL nos devuelve directo al paso del PIN
        # de esa misma empresa en vez de hacer elegir de nuevo.
        slug = self.router.page.params.get("empresa", "")
        empresa = next((c for c in self.companies_activas if c.slug == slug), None) if slug else None
        if empresa is not None:
            self.login_selected_company_id = empresa.id
            self.login_selected_company_slug = empresa.slug
            self.login_step = "pin"
        else:
            self.login_step = "restaurant"
            self.login_selected_company_id = 0
            self.login_selected_company_slug = ""
        return None

    def seleccionar_restaurante(self, company_id: int) -> None:
        empresa = next((c for c in self.companies_activas if c.id == company_id), None)
        if empresa is None:
            return
        self.login_selected_company_id = empresa.id
        self.login_selected_company_slug = empresa.slug
        self.login_step = "pin"
        self.login_error = ""

    def volver_a_seleccion_restaurante(self) -> None:
        self.login_step = "restaurant"
        self.login_pin_input = ""
        self.login_error = ""

    def on_load_mozos(self):
        self.stop_caja_polling()
        self.stop_cocina_polling()
        self.stop_mostrador_polling()
        return self._route_access_result("mozos")

    def on_load_caja(self):
        self.stop_mozos_polling()
        self.stop_cocina_polling()
        self.stop_mostrador_polling()
        result = self._route_access_result("caja")
        if result is not None:
            return result
        self.cargar_turno_caja()
        return None

    def on_load_mostrador(self):
        self.stop_mozos_polling()
        self.stop_caja_polling()
        self.stop_cocina_polling()
        result = self._route_access_result("mostrador")
        if result is not None:
            return result
        self.cargar_turno_caja()
        self.cargar_pedidos_mostrador_listos()
        self.cargar_pedidos_mostrador_entregados()
        return None

    def on_load_cocina(self):
        self.stop_mozos_polling()
        self.stop_caja_polling()
        self.stop_mostrador_polling()
        return self._route_access_result("cocina")

    def on_load_carta(self):
        return self._route_access_result("carta")

    def on_load_reportes(self):
        result = self._route_access_result("reportes")
        if result is not None:
            return result
        self.historial_filtro_fecha_desde = _utcnow().strftime("%Y-%m-%d")
        self.cargar_dashboard()
        self.cargar_historial_ventas()
        return None

    def on_load_usuarios(self):
        if self.usuario_actual is None:
            return rx.redirect("/login", replace=True)
        if self.usuario_actual.rol not in ROLE_ALLOWED_ROUTES["usuarios"]:
            return [
                rx.window_alert("No tienes permisos para este modulo."),
                rx.redirect(self.usuario_home_route, replace=True),
            ]
        self.cargar_usuarios_admin()
        self._limpiar_usuario_form()
        return None

    def on_load_configuracion(self):
        result = self._route_access_result("configuracion")
        if result is not None:
            return result
        self.cargar_config_impresora()
        self.cargar_mesas_config()
        return None

    def on_load_dono_page(self) -> None:
        self.cargar_config_impresora()
        self.cargar_mesas_config()
        self.historial_filtro_fecha_desde = _utcnow().strftime("%Y-%m-%d")
        self.historial_filtro_fecha_hasta = ""
        self.historial_filtro_metodo = ""
        self.cargar_dashboard()
        self.cargar_historial_ventas()
        self.cargar_inventario()
        self.cargar_clientes()
        self.cargar_promociones()
        self.cargar_cuentas()
        self.cargar_usuarios_admin()

    # ─── Autenticación (PIN + company_id) ────────────────────────────────────

    def set_login_pin(self, value: str) -> None:
        self.login_pin_input = _normalize_pin(value)

    def append_login_digit(self, digit: str) -> None:
        if not digit.isdigit() or len(self.login_pin_input) >= 6:
            return
        self.login_error = ""
        self.login_pin_input = f"{self.login_pin_input}{digit}"

    def backspace_login_pin(self) -> None:
        self.login_pin_input = self.login_pin_input[:-1]

    def clear_login_pin(self) -> None:
        self.login_pin_input = ""

    def seleccionar_login_rol(self, rol: str) -> None:
        self.login_rol_seleccionado = rol
        self.login_error = ""

    def _authenticate_with_pin(self, pin: str):
        normalized = _normalize_pin(pin)
        if len(normalized) < 4:
            self.login_pin_input = ""
            self.login_error = "Ingresa un PIN válido de 4 a 6 dígitos."
            return
        self.login_error = ""
        with self._tenant_session() as session:
            candidatos = session.exec(
                select(UsuarioFood).where(
                    UsuarioFood.company_id == self._company_id(),
                    UsuarioFood.activo.is_(True),
                )
            ).all()
            usuario = next((u for u in candidatos if _verify_pin(normalized, u.pin)), None)
        if usuario is None:
            self.login_pin_input = ""
            self.login_error = "PIN incorrecto. Intenta nuevamente."
            return
        if (
            self.login_rol_seleccionado
            and usuario.rol != RolUsuario.ADMIN.value
            and usuario.rol != self.login_rol_seleccionado
        ):
            self.login_pin_input = ""
            self.login_error = (
                f"Ese PIN pertenece al rol {usuario.rol}. "
                f"Seleccioná {usuario.rol} para ingresar."
            )
            return
        self.login_error = ""
        self.usuario_actual = UsuarioSesion(
            id=usuario.id or 0,
            nombre=usuario.nombre,
            rol=usuario.rol,
            company_id=usuario.company_id,
        )
        self.login_pin_input = ""
        self.mensaje = f"Sesion iniciada como {usuario.nombre}."
        return rx.redirect(_role_home_route(usuario.rol), replace=True)

    def login(self, pin: str):
        return self._authenticate_with_pin(pin)

    def submit_login_pin(self):
        return self._authenticate_with_pin(self.login_pin_input)

    def logout(self):
        self.usuario_actual = None
        self._clear_operational_context()
        return rx.redirect("/login", replace=True)

    # ─── Gestión de usuarios del local ───────────────────────────────────────

    def on_change_uf_nombre(self, value: str) -> None:
        self.usuario_form_nombre = value

    def on_change_uf_rol(self, value: str) -> None:
        self.usuario_form_rol = value

    def on_change_uf_pin(self, value: str) -> None:
        self.usuario_form_pin = value

    def on_change_uf_pin_confirm(self, value: str) -> None:
        self.usuario_form_pin_confirm = value

    def toggle_uf_activo(self) -> None:
        self.usuario_form_activo = not self.usuario_form_activo

    def _limpiar_usuario_form(self) -> None:
        self.usuario_form_id = 0
        self.usuario_form_nombre = ""
        self.usuario_form_rol = RolUsuario.MOZO.value
        self.usuario_form_pin = ""
        self.usuario_form_pin_confirm = ""
        self.usuario_form_activo = True
        self.usuario_form_visible = False

    def set_usuario_form_visible(self, v: bool) -> None:
        self.usuario_form_visible = v

    def cargar_usuarios_admin(self) -> None:
        mi_id = self.usuario_actual.id if self.usuario_actual else 0
        with self._tenant_session() as session:
            rows = session.exec(
                select(UsuarioFood)
                .where(UsuarioFood.company_id == self._company_id())
                .order_by(UsuarioFood.rol, UsuarioFood.nombre)
            ).all()
        self.usuarios_admin = [
            UsuarioAdminView(
                id=u.id or 0,
                nombre=u.nombre,
                rol=u.rol,
                rol_label=_ROL_LABELS.get(u.rol, u.rol),
                pin_masked="●●●●",
                activo=u.activo,
                badge_bg=_ROL_BADGE_BG.get(u.rol, "rgba(100,116,139,0.16)"),
                badge_text=_ROL_BADGE_TEXT.get(u.rol, "#94A3B8"),
                es_yo=u.id == mi_id,
            )
            for u in rows
        ]

    def nuevo_usuario_form(self) -> None:
        self._limpiar_usuario_form()
        self.usuario_form_visible = True
        self.mensaje = ""

    def editar_usuario(self, user_id: int) -> None:
        with self._tenant_session() as session:
            u = session.get(UsuarioFood, user_id)
        if u is None or u.company_id != self._company_id():
            self.mensaje = "Usuario no encontrado."
            return
        self.usuario_form_id = u.id or 0
        self.usuario_form_nombre = u.nombre
        self.usuario_form_rol = u.rol
        self.usuario_form_pin = ""
        self.usuario_form_pin_confirm = ""
        self.usuario_form_activo = u.activo
        self.usuario_form_visible = True
        self.mensaje = ""

    def guardar_usuario(self) -> None:
        nombre = self.usuario_form_nombre.strip()
        if not nombre:
            self.mensaje = "El nombre es obligatorio."
            return
        rol = self.usuario_form_rol
        if rol not in [r.value for r in RolUsuario]:
            self.mensaje = "Rol invalido."
            return

        nuevo_pin = _normalize_pin(self.usuario_form_pin)
        es_edicion = self.usuario_form_id > 0

        if not es_edicion:
            if len(nuevo_pin) < 4:
                self.mensaje = "El PIN debe tener al menos 4 digitos."
                return
        else:
            if self.usuario_form_pin and len(nuevo_pin) < 4:
                self.mensaje = "El nuevo PIN debe tener al menos 4 digitos."
                return

        if self.usuario_form_pin:
            pin_confirm = _normalize_pin(self.usuario_form_pin_confirm)
            if nuevo_pin != pin_confirm:
                self.mensaje = "Los PINs no coinciden."
                return

        with self._tenant_session() as session:
            otros = session.exec(
                select(UsuarioFood).where(
                    UsuarioFood.company_id == self._company_id(),
                    UsuarioFood.id != self.usuario_form_id,
                )
            ).all()
            if es_edicion:
                u = session.get(UsuarioFood, self.usuario_form_id)
                if u is None or u.company_id != self._company_id():
                    self.mensaje = "Usuario no encontrado."
                    return
                if nuevo_pin:
                    conflicto = next((o for o in otros if _verify_pin(nuevo_pin, o.pin)), None)
                    if conflicto:
                        self.mensaje = f"El PIN {nuevo_pin} ya lo usa {conflicto.nombre}."
                        return
                    u.pin = _hash_pin(nuevo_pin)
                u.nombre = nombre
                u.rol = rol
                u.activo = self.usuario_form_activo
                u.updated_at = _utcnow()
                session.add(u)
                session.commit()
                self.mensaje = f"Usuario '{nombre}' actualizado."
            else:
                conflicto = next((o for o in otros if _verify_pin(nuevo_pin, o.pin)), None)
                if conflicto:
                    self.mensaje = f"El PIN {nuevo_pin} ya lo usa {conflicto.nombre}."
                    return
                u = UsuarioFood(
                    company_id=self._company_id(),
                    nombre=nombre,
                    pin=_hash_pin(nuevo_pin),
                    rol=rol,
                    activo=True,
                )
                session.add(u)
                session.commit()
                self.mensaje = f"Usuario '{nombre}' creado."

        self._limpiar_usuario_form()
        self.cargar_usuarios_admin()

    def toggle_usuario_activo(self, user_id: int) -> None:
        mi_id = self.usuario_actual.id if self.usuario_actual else 0
        if user_id == mi_id:
            self.mensaje = "No puedes desactivarte a ti mismo."
            return
        with self._tenant_session() as session:
            u = session.get(UsuarioFood, user_id)
            if u is None or u.company_id != self._company_id():
                self.mensaje = "Usuario no encontrado."
                return
            if u.activo and u.rol == RolUsuario.ADMIN.value:
                admins_activos = session.exec(
                    select(UsuarioFood).where(
                        UsuarioFood.company_id == self._company_id(),
                        UsuarioFood.rol == RolUsuario.ADMIN.value,
                        UsuarioFood.activo.is_(True),
                    )
                ).all()
                if len(admins_activos) <= 1:
                    self.mensaje = "No puedes desactivar al ultimo administrador."
                    return
            u.activo = not u.activo
            u.updated_at = _utcnow()
            session.add(u)
            session.commit()
            accion = "activado" if u.activo else "desactivado"
            self.mensaje = f"Usuario '{u.nombre}' {accion}."
        self.cargar_usuarios_admin()

    def cancelar_usuario_form(self) -> None:
        self._limpiar_usuario_form()
        self.mensaje = ""

    # ─── Configuración impresoras ─────────────────────────────────────────────

    def cargar_config_impresora(self) -> None:
        with self._tenant_session() as session:
            cfg = session.exec(
                select(ConfigImpresora).where(ConfigImpresora.company_id == self._company_id())
            ).first()
            if cfg:
                self.config_nombre_local = cfg.nombre_local
                self.config_ticket_paper_width_mm = str(cfg.ticket_paper_width_mm)
                self.config_slug = cfg.slug or "mi-restaurante"
                self.config_admin_email = cfg.admin_email or ""
                url = f"{_FOOD_BASE_URL}/menu/{self.config_slug}"
                self.config_menu_url = url
                self.config_menu_qr_base64 = _generar_qr_base64(url)
            empresa = session.get(Company, self._company_id())
            self.config_logo_url = (empresa.logo_url or "") if empresa else ""

    def guardar_config_impresora(self) -> None:
        with self._tenant_session() as session:
            cfg = session.exec(
                select(ConfigImpresora).where(ConfigImpresora.company_id == self._company_id())
            ).first()
            if cfg is None:
                cfg = ConfigImpresora(company_id=self._company_id())
            cfg.nombre_local = self.config_nombre_local.strip() or "Mi Restaurante"
            try:
                ancho = int(self.config_ticket_paper_width_mm.strip())
                cfg.ticket_paper_width_mm = ancho if ancho in (58, 80) else 80
            except (ValueError, AttributeError):
                cfg.ticket_paper_width_mm = 80
            slug = _slugify(self.config_slug) if self.config_slug.strip() else _slugify(cfg.nombre_local)
            cfg.slug = slug
            cfg.updated_at = _utcnow()
            session.add(cfg)
            empresa = session.get(Company, self._company_id())
            if empresa is not None:
                empresa.logo_url = self.config_logo_url or None
                session.add(empresa)
            session.commit()
        self.config_slug = slug
        url = f"{_FOOD_BASE_URL}/menu/{slug}"
        self.config_menu_url = url
        self.config_menu_qr_base64 = _generar_qr_base64(url)
        self.mensaje = "Configuracion guardada."

    def set_config_nombre_local(self, v: str) -> None:
        self.config_nombre_local = v

    def set_config_ticket_paper_width_mm(self, v: str) -> None:
        self.config_ticket_paper_width_mm = v

    def set_config_slug(self, v: str) -> None:
        self.config_slug = v

    def set_config_admin_email(self, v: str) -> None:
        self.config_admin_email = v

    def set_config_admin_password_nueva(self, v: str) -> None:
        self.config_admin_password_nueva = v

    def set_config_admin_password_confirm(self, v: str) -> None:
        self.config_admin_password_confirm = v

    def toggle_config_admin_show_password(self) -> None:
        self.config_admin_show_password = not self.config_admin_show_password

    def guardar_admin_cuenta(self) -> None:
        import hashlib
        email = self.config_admin_email.strip().lower()
        if not email or "@" not in email:
            self.mensaje = "Ingresa un email valido."
            return
        nueva = self.config_admin_password_nueva.strip()
        confirm = self.config_admin_password_confirm.strip()
        if nueva and nueva != confirm:
            self.mensaje = "Las contraseñas no coinciden."
            return
        with self._tenant_session() as session:
            cfg = session.exec(
                select(ConfigImpresora).where(ConfigImpresora.company_id == self._company_id())
            ).first()
            if cfg is None:
                cfg = ConfigImpresora(company_id=self._company_id())
            cfg.admin_email = email
            if nueva:
                cfg.admin_password_hash = hashlib.sha256(nueva.encode()).hexdigest()
            session.add(cfg)
            session.commit()
        self.config_admin_email = email
        self.config_admin_password_nueva = ""
        self.config_admin_password_confirm = ""
        self.mensaje = "Cuenta del dueño guardada."

    # ─── CRUD Mesas (admin config) ────────────────────────────────────────────

    def cargar_mesas_config(self) -> None:
        with self._tenant_session() as session:
            mesas = session.exec(
                select(Mesa)
                .where(Mesa.company_id == self._company_id())
                .order_by(Mesa.numero)
            ).all()
            self.mesas_config = [
                MesaAdminView(
                    id=m.id or 0,
                    numero=m.numero,
                    nombre=m.nombre or "",
                    capacidad=m.capacidad,
                    activa=m.activa,
                    estado=m.estado,
                )
                for m in mesas
            ]

    def _reset_mesa_config_form(self) -> None:
        self.mesa_config_form_id = 0
        self.mesa_config_form_numero = ""
        self.mesa_config_form_nombre = ""
        self.mesa_config_form_capacidad = "4"

    def cancelar_mesa_config_form(self) -> None:
        self._reset_mesa_config_form()

    def editar_mesa_config(self, mesa_id: int) -> None:
        with self._tenant_session() as session:
            m = session.get(Mesa, mesa_id)
        if m is None or m.company_id != self._company_id():
            return
        self.mesa_config_form_id = m.id or 0
        self.mesa_config_form_numero = str(m.numero)
        self.mesa_config_form_nombre = m.nombre or ""
        self.mesa_config_form_capacidad = str(m.capacidad)

    def guardar_mesa_config(self) -> None:
        try:
            numero = int(self.mesa_config_form_numero.strip())
        except (ValueError, AttributeError):
            self.mensaje = "El numero de mesa debe ser un entero."
            return
        try:
            capacidad = max(1, int(self.mesa_config_form_capacidad.strip() or "4"))
        except ValueError:
            capacidad = 4
        nombre = self.mesa_config_form_nombre.strip()
        with self._tenant_session() as session:
            es_edicion = self.mesa_config_form_id > 0
            if es_edicion:
                m = session.get(Mesa, self.mesa_config_form_id)
                if m is None or m.company_id != self._company_id():
                    self.mensaje = "Mesa no encontrada."
                    return
            else:
                conflicto = session.exec(
                    select(Mesa).where(
                        Mesa.company_id == self._company_id(),
                        Mesa.numero == numero,
                    )
                ).first()
                if conflicto:
                    self.mensaje = f"Ya existe la mesa #{numero}."
                    return
                m = Mesa(company_id=self._company_id(), numero=numero)
            m.numero = numero
            m.nombre = nombre
            m.capacidad = capacidad
            m.updated_at = _utcnow()
            session.add(m)
            session.commit()
        accion = "actualizada" if self.mesa_config_form_id > 0 else "creada"
        self.mensaje = f"Mesa #{numero} {accion}."
        self._reset_mesa_config_form()
        self.cargar_mesas_config()

    def toggle_mesa_activa_config(self, mesa_id: int) -> None:
        with self._tenant_session() as session:
            m = session.get(Mesa, mesa_id)
            if m is None or m.company_id != self._company_id():
                return
            m.activa = not m.activa
            m.updated_at = _utcnow()
            session.add(m)
            session.commit()
        self.cargar_mesas_config()

    def eliminar_mesa_config(self, mesa_id: int) -> None:
        with self._tenant_session() as session:
            m = session.get(Mesa, mesa_id)
            if m is None or m.company_id != self._company_id():
                self.mensaje = "Mesa no encontrada."
                return
            pedido_abierto = session.exec(
                select(Pedido).where(
                    Pedido.mesa_id == mesa_id,
                    Pedido.estado.in_(list(OPEN_ORDER_STATES)),
                )
            ).first()
            if pedido_abierto:
                self.mensaje = f"La mesa #{m.numero} tiene un pedido activo — no se puede eliminar."
                return
            session.delete(m)
            session.commit()
        self.mensaje = f"Mesa #{m.numero} eliminada."
        self.cargar_mesas_config()

    def set_mesa_config_form_numero(self, v: str) -> None:
        self.mesa_config_form_numero = v

    def set_mesa_config_form_nombre(self, v: str) -> None:
        self.mesa_config_form_nombre = v

    def set_mesa_config_form_capacidad(self, v: str) -> None:
        self.mesa_config_form_capacidad = v

    def _ticket_paper_width_mm(self) -> int:
        try:
            with self._tenant_session() as session:
                cfg = session.exec(
                    select(ConfigImpresora).where(ConfigImpresora.company_id == self._company_id())
                ).first()
                if cfg is not None:
                    return cfg.ticket_paper_width_mm
        except Exception:
            pass
        return 80

    # ─── Polling ──────────────────────────────────────────────────────────────

    def _refresh_mozos_slice(self) -> None:
        if self.usuario_actual is None:
            return
        self.cargar_mesas()
        if self.mesa_seleccionada_id:
            self._cargar_carrito_mesa(self.mesa_seleccionada_id)
            self._cargar_historial_mesa(self.mesa_seleccionada_id)

    def _refresh_cocina_slice(self) -> None:
        if self.usuario_actual is None:
            return
        self.cargar_cocina()
        self.cargar_mesas()

    def _refresh_caja_slice(self) -> None:
        if self.usuario_actual is None:
            return
        self.cargar_mesas()
        if self.mesa_seleccionada_id:
            self._cargar_historial_mesa(self.mesa_seleccionada_id)

    def _refresh_mostrador_slice(self) -> None:
        if self.usuario_actual is None:
            return
        self.cargar_pedidos_mostrador_listos()
        self.cargar_pedidos_mostrador_entregados()

    async def _run_polling_loop(self, flag_name: str, interval_seconds: int, refresh_callback) -> None:
        async with self:
            if getattr(self, flag_name):
                return
            setattr(self, flag_name, True)
            refresh_callback()
        while True:
            await asyncio.sleep(interval_seconds)
            try:
                async with self:
                    if not getattr(self, flag_name):
                        break
                    refresh_callback()
            except Exception:
                break

    @rx.event(background=True)
    async def start_mozos_polling(self) -> None:
        await self._run_polling_loop("mozos_polling_enabled", 5, self._refresh_mozos_slice)

    def stop_mozos_polling(self) -> None:
        self.mozos_polling_enabled = False

    @rx.event(background=True)
    async def start_cocina_polling(self) -> None:
        await self._run_polling_loop("cocina_polling_enabled", 5, self._refresh_cocina_slice)

    def stop_cocina_polling(self) -> None:
        self.cocina_polling_enabled = False

    @rx.event(background=True)
    async def start_caja_polling(self) -> None:
        await self._run_polling_loop("caja_polling_enabled", 10, self._refresh_caja_slice)

    def stop_caja_polling(self) -> None:
        self.caja_polling_enabled = False

    @rx.event(background=True)
    async def start_mostrador_polling(self) -> None:
        await self._run_polling_loop("mostrador_polling_enabled", 5, self._refresh_mostrador_slice)

    def stop_mostrador_polling(self) -> None:
        self.mostrador_polling_enabled = False

    # ─── Mesas ───────────────────────────────────────────────────────────────

    def cargar_mesas(self) -> None:
        mesas_ui: list[MesaView] = []
        with self._tenant_session() as session:
            mesas_db = session.exec(
                select(Mesa).where(
                    Mesa.company_id == self._company_id(),
                    Mesa.activa.is_(True),
                ).order_by(Mesa.numero)
            ).all()
            for mesa in mesas_db:
                pedido_abierto = _get_open_order(session, mesa.id or 0, self._company_id())
                # Auto-corregir mesa stuck: si figura OCUPADA pero no hay pedido abierto, volver a LIBRE
                if mesa.estado != EstadoMesa.LIBRE.value and pedido_abierto is None:
                    mesa.estado = EstadoMesa.LIBRE.value
                    mesa.updated_at = _utcnow()
                    session.add(mesa)
                    session.commit()
                total_abierto = _to_decimal(pedido_abierto.total if pedido_abierto else Decimal("0.00"))
                ready_details = _get_ready_details(session, pedido_abierto.id or 0) if pedido_abierto else []
                items_listos_count = sum(d.cantidad for d in ready_details)
                tiene_items_listos = items_listos_count > 0
                items_total_count = 0
                if pedido_abierto is not None:
                    todos_los_detalles = session.exec(
                        select(DetallePedido).where(DetallePedido.pedido_id == pedido_abierto.id)
                    ).all()
                    items_total_count = sum(d.cantidad for d in todos_los_detalles)
                tiempo_abierto_texto = ""
                if pedido_abierto is not None:
                    elapsed_min = max(0, int((_utcnow() - pedido_abierto.created_at).total_seconds() // 60))
                    tiempo_abierto_texto = f"{elapsed_min} min"
                mesas_ui.append(MesaView(
                    id=mesa.id or 0,
                    numero=mesa.numero,
                    label=f"Mesa {mesa.numero}",
                    nombre=mesa.nombre or f"Mesa {mesa.numero}",
                    estado=mesa.estado,
                    estado_label=MESA_LABELS.get(mesa.estado, mesa.estado),
                    badge_bg=MESA_BADGE_BACKGROUNDS.get(mesa.estado, "#E5E7EB"),
                    badge_text=MESA_BADGE_TEXTS.get(mesa.estado, "#111827"),
                    capacidad=mesa.capacidad,
                    total_abierto=float(total_abierto),
                    total_abierto_texto=_money_text(total_abierto),
                    card_bg=MESA_CARD_BACKGROUNDS.get(mesa.estado, "#FFFFFF"),
                    card_border=(
                        READY_ALERT_BORDER if tiene_items_listos
                        else MESA_CARD_BORDERS.get(mesa.estado, "1px solid #E5E7EB")
                    ),
                    tiene_items_listos=tiene_items_listos,
                    items_listos_count=items_listos_count,
                    items_total_count=items_total_count,
                    tiempo_abierto_texto=tiempo_abierto_texto,
                ))
        self.mesas = mesas_ui
        if self.mesa_seleccionada_id and not any(m.id == self.mesa_seleccionada_id for m in self.mesas):
            self.mesa_seleccionada_id = 0
            self.carrito = []
            self.historial_pedido = []

    # ─── Carta ────────────────────────────────────────────────────────────────

    def cargar_menu(self) -> None:
        with self._tenant_session() as session:
            categorias_db = session.exec(
                select(Categoria).where(Categoria.company_id == self._company_id()).order_by(Categoria.orden, Categoria.nombre)
            ).all()
            productos_db = session.exec(
                select(Producto).where(Producto.company_id == self._company_id()).order_by(Producto.nombre)
            ).all()
            categorias_map = {c.id: c.nombre for c in categorias_db}
            conteo_por_categoria: dict[int, int] = {}
            for p in productos_db:
                conteo_por_categoria[p.categoria_id] = conteo_por_categoria.get(p.categoria_id, 0) + 1
            self.categorias = [
                CategoriaView(
                    id=c.id or 0,
                    nombre=c.nombre,
                    descripcion=c.descripcion or "",
                    orden=c.orden,
                    activa=c.activa,
                    productos_count=conteo_por_categoria.get(c.id or 0, 0),
                    emoji=_emoji_para_categoria(c.nombre),
                )
                for c in categorias_db
            ]
            self.productos = [
                ProductoView(
                    id=p.id or 0,
                    categoria_id=p.categoria_id,
                    categoria_nombre=categorias_map.get(p.categoria_id, "General"),
                    nombre=p.nombre,
                    descripcion=p.descripcion or "",
                    precio=float(_to_decimal(p.precio)),
                    precio_texto=_money_text(p.precio),
                    disponible=p.disponible,
                    imagen_url=p.imagen_url or "",
                    emoji=p.emoji or _emoji_para_producto(p.nombre),
                )
                for p in productos_db
            ]

    def seleccionar_categoria(self, categoria_id: int) -> None:
        self.categoria_activa_id = categoria_id

    def seleccionar_mostrador_categoria(self, categoria_id: int) -> None:
        self.mostrador_categoria_activa_id = categoria_id

    # ─── Mozos — Selección de mesa ────────────────────────────────────────────

    def seleccionar_mesa(self, mesa_id: int) -> None:
        self.mesa_seleccionada_id = mesa_id
        self._cargar_carrito_mesa(mesa_id)
        self._cargar_historial_mesa(mesa_id)
        mesa = next((m for m in self.mesas if m.id == mesa_id), None)
        alerta = (
            f" {mesa.items_listos_count} items listos para entregar."
            if mesa and mesa.tiene_items_listos else ""
        )
        self.mensaje = f"{self.mesa_seleccionada_label} seleccionada. {self.cantidad_items_carrito} items pendientes.{alerta}"

    def _cargar_carrito_mesa(self, mesa_id: int) -> None:
        with self._tenant_session() as session:
            pedido = _get_open_order(session, mesa_id, self._company_id())
            if pedido is None:
                self.carrito = []
                return
            detalles = _get_unsent_details(session, pedido.id or 0)
            productos_map = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            self.carrito = [
                CarritoItem(
                    producto_id=d.producto_id,
                    nombre=(productos_map[d.producto_id].nombre if d.producto_id in productos_map else f"Producto {d.producto_id}"),
                    cantidad=d.cantidad,
                    precio_unitario=float(_to_decimal(d.precio_unitario)),
                    subtotal=float(_to_decimal(d.subtotal)),
                    subtotal_texto=_money_text(d.subtotal),
                    nota=d.notas or "",
                )
                for d in detalles
            ]

    def _cargar_historial_mesa(self, mesa_id: int) -> None:
        with self._tenant_session() as session:
            pedido = _get_open_order(session, mesa_id, self._company_id())
            if pedido is None:
                self.historial_pedido = []
                self.mesa_atendida_por_nombre = ""
                self.nota_pedido_mesa = ""
                return
            detalles = session.exec(
                select(DetallePedido).where(
                    DetallePedido.pedido_id == pedido.id,
                    DetallePedido.impreso_cocina.is_(True),
                ).order_by(DetallePedido.enviado_cocina_at, DetallePedido.id)
            ).all()
            productos_map = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            usuarios_map = {u.id: u for u in session.exec(select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())).all()}
            mozo = usuarios_map.get(pedido.mozo_id)
            self.mesa_atendida_por_nombre = _actor_name(mozo.nombre if mozo else "")
            self.nota_pedido_mesa = pedido.notas or ""
            historial: list[HistorialItem] = []
            for d in detalles:
                producto = productos_map.get(d.producto_id)
                preparado_por = usuarios_map.get(d.preparado_por_id)
                enviado_en = d.enviado_cocina_at or d.updated_at
                estado_produccion = d.estado_produccion or EstadoProduccion.PENDIENTE.value
                historial.append(HistorialItem(
                    detalle_id=d.id or 0,
                    nombre=producto.nombre if producto else f"Producto {d.producto_id}",
                    cantidad=d.cantidad,
                    precio_unitario_texto=_money_text(d.precio_unitario),
                    subtotal_texto=_money_text(d.subtotal),
                    nota=d.notas or "",
                    enviado_en_texto=enviado_en.strftime("%H:%M"),
                    estado_clave=estado_produccion,
                    estado_label=PRODUCTION_LABELS.get(estado_produccion, estado_produccion),
                    estado_bg=PRODUCTION_BADGE_BACKGROUNDS.get(estado_produccion, "#E2E8F0"),
                    estado_color=PRODUCTION_BADGE_TEXTS.get(estado_produccion, "#334155"),
                    preparado_por_nombre=_actor_name(preparado_por.nombre if preparado_por else ""),
                    puede_entregar=(estado_produccion == EstadoProduccion.LISTO_PARA_ENTREGAR.value),
                    puede_cancelar=(estado_produccion == EstadoProduccion.PENDIENTE.value),
                ))
            self.historial_pedido = historial

    def agregar_producto(self, producto_id: int) -> None:
        if self.mesa_seleccionada_id == 0:
            self.mensaje = "Selecciona una mesa antes de agregar productos."
            return
        with self._tenant_session() as session:
            mesa = session.get(Mesa, self.mesa_seleccionada_id)
            if mesa is None or mesa.company_id != self._company_id():
                self.mensaje = "La mesa seleccionada ya no existe."
                return
            producto = session.get(Producto, producto_id)
            if producto is None or producto.company_id != self._company_id() or not producto.disponible:
                self.mensaje = "Producto no disponible."
                return
            producto_nombre = producto.nombre
            pedido = _ensure_open_order(session, mesa, self._company_id(), mozo_id=self.usuario_actual.id if self.usuario_actual else None)
            detalle = session.exec(
                select(DetallePedido).where(
                    DetallePedido.pedido_id == pedido.id,
                    DetallePedido.producto_id == producto.id,
                    DetallePedido.impreso_cocina.is_(False),
                ).order_by(DetallePedido.id.desc())
            ).first()
            precio = _to_decimal(producto.precio)
            if detalle is None:
                detalle = DetallePedido(
                    company_id=self._company_id(),
                    pedido_id=pedido.id or 0,
                    producto_id=producto.id or 0,
                    cantidad=1,
                    precio_unitario=precio,
                    subtotal=precio,
                    estado_produccion=EstadoProduccion.PENDIENTE.value,
                    impreso_cocina=False,
                    impreso_caja=False,
                )
            else:
                detalle.cantidad += 1
                detalle.precio_unitario = precio
                detalle.subtotal = precio * detalle.cantidad
            session.add(detalle)
            _recalculate_order_total(session, pedido)
            mesa.estado = EstadoMesa.OCUPADA.value
            mesa.updated_at = _utcnow()
            session.add(mesa)
            session.commit()
        self._cargar_carrito_mesa(self.mesa_seleccionada_id)
        self._cargar_historial_mesa(self.mesa_seleccionada_id)
        self.cargar_mesas()
        self.mensaje = f"{producto_nombre} agregado a {self.mesa_seleccionada_label}."

    def restar_producto(self, producto_id: int) -> None:
        if self.mesa_seleccionada_id == 0:
            self.mensaje = "Selecciona una mesa antes de editar el carrito."
            return
        with self._tenant_session() as session:
            mesa = session.get(Mesa, self.mesa_seleccionada_id)
            if mesa is None:
                self.mensaje = "La mesa seleccionada ya no existe."
                return
            pedido = _get_open_order(session, mesa.id or 0, self._company_id())
            if pedido is None:
                self.mensaje = "No hay pedido abierto para esta mesa."
                return
            detalle = session.exec(
                select(DetallePedido).where(
                    DetallePedido.pedido_id == pedido.id,
                    DetallePedido.producto_id == producto_id,
                    DetallePedido.impreso_cocina.is_(False),
                ).order_by(DetallePedido.id.desc())
            ).first()
            if detalle is None:
                self.mensaje = "Ese producto ya fue enviado o no existe en el carrito."
                return
            detalle.cantidad -= 1
            if detalle.cantidad <= 0:
                session.delete(detalle)
            else:
                detalle.subtotal = _to_decimal(detalle.precio_unitario) * detalle.cantidad
                session.add(detalle)
            self._finalize_cart_cleanup(session, pedido, mesa)
            session.commit()
        self._cargar_carrito_mesa(self.mesa_seleccionada_id)
        self._cargar_historial_mesa(self.mesa_seleccionada_id)
        self.cargar_mesas()
        self.mensaje = "Carrito actualizado."

    def _finalize_cart_cleanup(self, session, pedido: Pedido, mesa: Mesa) -> None:
        detalles_restantes = session.exec(
            select(DetallePedido).where(DetallePedido.pedido_id == pedido.id)
        ).all()
        if not detalles_restantes:
            session.delete(pedido)
            mesa.estado = EstadoMesa.LIBRE.value
            mesa.updated_at = _utcnow()
            session.add(mesa)
            return
        _recalculate_order_total(session, pedido)
        _sync_order_status(session, pedido)
        mesa.estado = EstadoMesa.OCUPADA.value
        mesa.updated_at = _utcnow()
        session.add(mesa)

    def limpiar_carrito(self) -> None:
        if self.mesa_seleccionada_id == 0:
            self.carrito = []
            self.mensaje = "No hay mesa seleccionada."
            return
        with self._tenant_session() as session:
            mesa = session.get(Mesa, self.mesa_seleccionada_id)
            if mesa is None:
                self.mensaje = "La mesa seleccionada ya no existe."
                return
            pedido = _get_open_order(session, mesa.id or 0, self._company_id())
            if pedido is None:
                self.carrito = []
                self.mensaje = "No hay pedido abierto para limpiar."
                return
            for d in _get_unsent_details(session, pedido.id or 0):
                session.delete(d)
            self._finalize_cart_cleanup(session, pedido, mesa)
            session.commit()
        self._cargar_carrito_mesa(self.mesa_seleccionada_id)
        self._cargar_historial_mesa(self.mesa_seleccionada_id)
        self.cargar_mesas()
        self.mensaje = "Items pendientes eliminados."

    # ─── Notas ────────────────────────────────────────────────────────────────

    def set_mozos_tab(self, tab: str) -> None:
        self.mozos_tab_activa = tab

    def abrir_nota_item(self, producto_id: int) -> None:
        item = next((i for i in self.carrito if i.producto_id == producto_id), None)
        self.nota_producto_activo_id = producto_id
        self.nota_input_temporal = item.nota if item else ""

    def set_nota_input_temporal(self, value: str) -> None:
        self.nota_input_temporal = str(value)[:120]

    def guardar_nota_carrito_item(self, producto_id: int) -> None:
        if self.mesa_seleccionada_id == 0:
            self.nota_producto_activo_id = 0
            return
        nota = self.nota_input_temporal.strip()
        with self._tenant_session() as session:
            pedido = _get_open_order(session, self.mesa_seleccionada_id, self._company_id())
            if pedido is None:
                self.nota_producto_activo_id = 0
                return
            detalle = session.exec(
                select(DetallePedido).where(
                    DetallePedido.pedido_id == pedido.id,
                    DetallePedido.producto_id == producto_id,
                    DetallePedido.impreso_cocina.is_(False),
                ).order_by(DetallePedido.id.desc())
            ).first()
            if detalle is None:
                self.mensaje = "El item ya fue enviado a cocina; no se puede editar."
                self.nota_producto_activo_id = 0
                return
            detalle.notas = nota or None
            detalle.updated_at = _utcnow()
            session.add(detalle)
            session.commit()
        self.nota_producto_activo_id = 0
        self.nota_input_temporal = ""
        self._cargar_carrito_mesa(self.mesa_seleccionada_id)
        self.mensaje = "Nota guardada." if nota else "Nota eliminada."

    def cerrar_nota_item(self) -> None:
        self.nota_producto_activo_id = 0
        self.nota_input_temporal = ""

    def set_nota_pedido_mesa(self, value: str) -> None:
        self.nota_pedido_mesa = str(value)[:500]

    def guardar_nota_pedido_mesa(self) -> None:
        if self.mesa_seleccionada_id == 0:
            return
        nota = self.nota_pedido_mesa.strip()
        with self._tenant_session() as session:
            pedido = _get_open_order(session, self.mesa_seleccionada_id, self._company_id())
            if pedido is None:
                return
            pedido.notas = nota or None
            pedido.updated_at = _utcnow()
            session.add(pedido)
            session.commit()
        self.mensaje = "Nota del pedido guardada." if nota else "Nota del pedido eliminada."

    # ─── Enviar a cocina ─────────────────────────────────────────────────────

    def solicitar_cuenta(self) -> None:
        if self.mesa_seleccionada_id == 0:
            self.mensaje = "Selecciona una mesa antes de solicitar cuenta."
            return
        if self.cantidad_items_carrito > 0:
            self.mensaje = "Primero envia a cocina los items pendientes."
            return
        with self._tenant_session() as session:
            mesa = session.get(Mesa, self.mesa_seleccionada_id)
            if mesa is None:
                self.mensaje = "La mesa seleccionada ya no existe."
                return
            pedido = _get_open_order(session, mesa.id or 0, self._company_id())
            if pedido is None or _to_decimal(pedido.total) <= 0:
                self.mensaje = "No hay consumo pendiente en esa mesa."
                return
            if _get_not_delivered_details(session, pedido.id or 0):
                self.mensaje = "Todavia hay items en cocina o listos por entregar."
                return
            mesa.estado = EstadoMesa.ESPERANDO_CUENTA.value
            mesa.updated_at = _utcnow()
            session.add(mesa)
            session.commit()
        self.cargar_mesas()
        self.mensaje = f"{self.mesa_seleccionada_label} marcada para cobrar."

    def enviar_pedido(self) -> None:
        if self.mesa_seleccionada_id == 0:
            self.mensaje = "Selecciona una mesa antes de enviar el pedido."
            return
        pedido_id = 0
        mesa_label = ""
        ticket_lines: list[TicketLine] = []
        with self._tenant_session() as session:
            mesa = session.get(Mesa, self.mesa_seleccionada_id)
            if mesa is None:
                self.mensaje = "La mesa seleccionada ya no existe."
                return
            pedido = _get_open_order(session, mesa.id or 0, self._company_id())
            if pedido is None:
                self.mensaje = "No hay items pendientes para enviar."
                return
            if self.usuario_actual and pedido.mozo_id is None:
                pedido.mozo_id = self.usuario_actual.id
                pedido.updated_at = _utcnow()
                session.add(pedido)
            detalles_pendientes = _get_unsent_details(session, pedido.id or 0)
            if not detalles_pendientes:
                self.mensaje = "No hay items nuevos pendientes de enviar."
                return
            errores_stock = _validar_stock_para_items(
                session, [(d.producto_id, d.cantidad) for d in detalles_pendientes], self._company_id()
            )
            if errores_stock:
                self.mensaje = "Stock insuficiente — " + "; ".join(errores_stock)
                return
            productos_map = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            now = _utcnow()
            for d in detalles_pendientes:
                producto = productos_map.get(d.producto_id)
                ticket_lines.append(TicketLine(
                    name=producto.nombre if producto else f"Producto {d.producto_id}",
                    quantity=d.cantidad,
                    unit_price=float(_to_decimal(d.precio_unitario)),
                    subtotal=float(_to_decimal(d.subtotal)),
                    note=d.notas or "",
                ))
                d.impreso_cocina = True
                d.enviado_cocina_at = now
                d.estado_produccion = EstadoProduccion.PENDIENTE.value
                session.add(d)
            _recalculate_order_total(session, pedido)
            _sync_order_status(session, pedido)
            mesa.estado = EstadoMesa.OCUPADA.value
            mesa.updated_at = now
            session.add(pedido)
            session.add(mesa)
            session.commit()
            pedido_id = pedido.id or 0
            mesa_label = mesa.nombre or f"Mesa {mesa.numero}"
        self.ultimo_pedido_id = pedido_id
        self._cargar_carrito_mesa(self.mesa_seleccionada_id)
        self._cargar_historial_mesa(self.mesa_seleccionada_id)
        self.cargar_mesas()
        self.cargar_cocina()
        html_ticket = generate_kitchen_ticket_html(
            mesa_label=mesa_label, pedido_id=pedido_id, items=ticket_lines,
            notes=self.nota_pedido_mesa, paper_width_mm=self._ticket_paper_width_mm(),
        )
        self.mensaje = f"Pedido #{pedido_id} enviado correctamente."
        return rx.call_script(build_print_script(html_ticket))

    # ─── Cocina (KDS) ────────────────────────────────────────────────────────

    def cargar_cocina(self) -> None:
        with self._tenant_session() as session:
            detalles = session.exec(
                select(DetallePedido).where(
                    DetallePedido.company_id == self._company_id(),
                    DetallePedido.impreso_cocina.is_(True),
                    DetallePedido.estado_produccion.in_(KITCHEN_VISIBLE_STATES),
                ).order_by(DetallePedido.enviado_cocina_at, DetallePedido.id)
            ).all()
            pedido_ids = {d.pedido_id for d in detalles}
            pedidos = {p.id: p for p in session.exec(select(Pedido).where(Pedido.id.in_(pedido_ids))).all()}
            mesas = {m.id: m for m in session.exec(select(Mesa).where(Mesa.company_id == self._company_id())).all()}
            usuarios = {u.id: u for u in session.exec(select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())).all()}
            productos = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            grupos: dict = {}
            for d in detalles:
                pedido = pedidos.get(d.pedido_id)
                if pedido is None:
                    continue
                marca = d.enviado_cocina_at or d.updated_at
                lote = marca.isoformat()
                estado_produccion = d.estado_produccion or EstadoProduccion.PENDIENTE.value
                mozo = usuarios.get(pedido.mozo_id)
                key = (pedido.id or 0, lote, estado_produccion)
                if key not in grupos:
                    grupos[key] = {
                        "pedido_id": pedido.id or 0,
                        "mesa_label": _pedido_kitchen_label(pedido, mesas),
                        "hora_texto": marca.strftime("%H:%M"),
                        "estado_produccion": estado_produccion,
                        "estado_label": PRODUCTION_LABELS.get(estado_produccion, estado_produccion),
                        "estado_bg": PRODUCTION_BADGE_BACKGROUNDS.get(estado_produccion, "#E2E8F0"),
                        "estado_color": PRODUCTION_BADGE_TEXTS.get(estado_produccion, "#334155"),
                        "mozo_nombre": _actor_name(mozo.nombre if mozo else ""),
                        "action_label": (
                            "▶ Iniciar preparación"
                            if estado_produccion == EstadoProduccion.PENDIENTE.value
                            else "✓ Todo listo"
                        ),
                        "accent_bg": KITCHEN_CARD_BACKGROUNDS.get(estado_produccion, "#FFF7ED"),
                        "accent_border": KITCHEN_CARD_BORDERS.get(estado_produccion, "#FCD34D"),
                        "detalle_ids": [],
                        "items_lines": [],
                    }
                producto = productos.get(d.producto_id)
                line = f"{d.cantidad} x {producto.nombre if producto else f'Producto {d.producto_id}'}"
                if d.notas:
                    line = f"{line} · Nota: {d.notas}"
                grupos[key]["items_lines"].append(line)
                grupos[key]["detalle_ids"].append(str(d.id or 0))
                grupos[key]["marca"] = marca
            ahora = _utcnow()
            for data in grupos.values():
                elapsed_seg = max(0, int((ahora - data["marca"]).total_seconds()))
                mins, segs = divmod(elapsed_seg, 60)
                if mins >= 99:
                    horas, mins_rest = divmod(mins, 60)
                    data["minutos_texto"] = f"{horas}h {mins_rest}min"
                else:
                    data["minutos_texto"] = f"{mins:02d}:{segs:02d} min"
                data["demorado"] = mins >= KITCHEN_DEMORADO_MINUTOS
                if data["demorado"]:
                    data["estado_label"] = "⚠ Demorado"
                    data["accent_border"] = KITCHEN_DEMORADO_COLOR
                    data["estado_bg"] = KITCHEN_DEMORADO_COLOR
                    data["estado_color"] = "#FEE2E2"
            self.tickets_cocina = [
                CocinaTicketView(
                    pedido_id=data["pedido_id"],
                    mesa_label=data["mesa_label"],
                    hora_texto=data["hora_texto"],
                    estado_produccion=data["estado_produccion"],
                    estado_label=data["estado_label"],
                    estado_bg=data["estado_bg"],
                    estado_color=data["estado_color"],
                    mozo_nombre=data["mozo_nombre"],
                    action_label=data["action_label"],
                    accent_bg=data["accent_bg"],
                    accent_border=data["accent_border"],
                    detalle_ids_csv=",".join(data["detalle_ids"]),
                    items_lines=data["items_lines"],
                    minutos_texto=data["minutos_texto"],
                    demorado=data["demorado"],
                )
                for _, data in grupos.items()
            ]

    def _transition_ticket_state(self, detalle_ids_csv: str, source_state: str, target_state: str, success_message: str, actor_user_id: int | None = None, actor_field_name: str | None = None) -> None:
        ids = [int(x) for x in detalle_ids_csv.split(",") if x.strip()]
        if not ids:
            self.mensaje = "No se encontro el ticket de cocina."
            return
        with self._tenant_session() as session:
            detalles = session.exec(select(DetallePedido).where(DetallePedido.id.in_(ids))).all()
            actualizables = [d for d in detalles if d.impreso_cocina and d.estado_produccion == source_state]
            if not actualizables:
                self.mensaje = "El ticket ya cambio de estado."
                return
            pedidos_afectados: set[int] = set()
            now = _utcnow()
            for d in actualizables:
                d.estado_produccion = target_state
                d.updated_at = now
                if actor_field_name and actor_user_id is not None:
                    setattr(d, actor_field_name, actor_user_id)
                session.add(d)
                pedidos_afectados.add(d.pedido_id)
            for pedido_id in pedidos_afectados:
                pedido = session.get(Pedido, pedido_id)
                if pedido is not None:
                    _sync_order_status(session, pedido)
            session.commit()
        self.cargar_cocina()
        self.cargar_mesas()
        if self.mesa_seleccionada_id:
            self._cargar_historial_mesa(self.mesa_seleccionada_id)
        self.mensaje = success_message

    def iniciar_preparacion_ticket(self, detalle_ids_csv: str) -> None:
        self._transition_ticket_state(
            detalle_ids_csv, EstadoProduccion.PENDIENTE.value,
            EstadoProduccion.EN_PREPARACION.value, "Ticket movido a preparacion.",
        )

    def marcar_ticket_listo(self, detalle_ids_csv: str) -> None:
        self._transition_ticket_state(
            detalle_ids_csv, EstadoProduccion.EN_PREPARACION.value,
            EstadoProduccion.LISTO_PARA_ENTREGAR.value, "Pedido listo para entregar a salon.",
            actor_user_id=(self.usuario_actual.id if self.usuario_actual else None),
            actor_field_name="preparado_por_id",
        )

    def entregar_item_historial(self, detalle_id: int) -> None:
        with self._tenant_session() as session:
            detalle = session.get(DetallePedido, detalle_id)
            if detalle is None or not detalle.impreso_cocina:
                self.mensaje = "El item indicado ya no existe."
                return
            if detalle.estado_produccion != EstadoProduccion.LISTO_PARA_ENTREGAR.value:
                self.mensaje = "Ese item no esta listo para entrega."
                return
            detalle.estado_produccion = EstadoProduccion.ENTREGADO_AL_CLIENTE.value
            detalle.updated_at = _utcnow()
            session.add(detalle)
            pedido = session.get(Pedido, detalle.pedido_id)
            if pedido is not None:
                _sync_order_status(session, pedido)
            session.commit()
        if self.mesa_seleccionada_id:
            self._cargar_historial_mesa(self.mesa_seleccionada_id)
        self.cargar_mesas()
        self.cargar_cocina()
        self.mensaje = "Item entregado a la mesa."

    def cancelar_item_pedido(self, detalle_id: int) -> None:
        with self._tenant_session() as session:
            detalle = session.get(DetallePedido, detalle_id)
            if detalle is None or detalle.company_id != self._company_id():
                self.mensaje = "El item indicado ya no existe."
                return
            if detalle.estado_produccion != EstadoProduccion.PENDIENTE.value:
                self.mensaje = "Solo se pueden cancelar items aun pendientes en cocina."
                return
            pedido = session.get(Pedido, detalle.pedido_id)
            if pedido is None:
                self.mensaje = "El pedido ya no existe."
                return
            nombre_item = ""
            producto = session.get(Producto, detalle.producto_id)
            if producto:
                nombre_item = producto.nombre
            session.delete(detalle)
            session.flush()
            _recalculate_order_total(session, pedido)
            _sync_order_status(session, pedido)
            session.commit()
        if self.mesa_seleccionada_id:
            self._cargar_historial_mesa(self.mesa_seleccionada_id)
            self._cargar_carrito_mesa(self.mesa_seleccionada_id)
        self.cargar_mesas()
        self.cargar_cocina()
        self.mensaje = f"Item '{nombre_item}' cancelado del pedido."

    # ─── Caja — Flujo de cobro con método de pago ────────────────────────────

    def abrir_cobro_mesa(self, mesa_id: int) -> None:
        mesa = next((m for m in self.mesas if m.id == mesa_id), None)
        if mesa is None or mesa.estado == EstadoMesa.LIBRE.value or mesa.total_abierto <= 0:
            self.mensaje = "Esa mesa no tiene consumo pendiente."
            return
        self.caja_cobro_mesa_id = mesa_id
        self.caja_cobro_metodo = "efectivo"
        self.caja_cobro_propina = ""
        self.caja_cobro_efectivo_recibido = ""
        self.caja_cobro_error = ""
        self.mensaje = ""
        items_ui: list[CajaItemView] = []
        with self._tenant_session() as session:
            pedido_abierto = _get_open_order(session, mesa_id, self._company_id())
            if pedido_abierto is not None:
                detalles = session.exec(
                    select(DetallePedido).where(
                        DetallePedido.pedido_id == pedido_abierto.id
                    ).order_by(DetallePedido.id)
                ).all()
                productos = {
                    p.id: p
                    for p in session.exec(
                        select(Producto).where(Producto.company_id == self._company_id())
                    ).all()
                }
                for d in detalles:
                    prod = productos.get(d.producto_id)
                    items_ui.append(CajaItemView(
                        producto_nombre=prod.nombre if prod else f"Producto {d.producto_id}",
                        cantidad=d.cantidad,
                        precio_unitario_texto=_money_text(_to_decimal(d.precio_unitario)),
                        subtotal_texto=_money_text(_to_decimal(d.subtotal)),
                        notas=d.notas or "",
                    ))
        self.caja_cobro_items = items_ui

    def cancelar_cobro(self) -> None:
        self.caja_cobro_mesa_id = 0
        self.caja_cobro_metodo = "efectivo"
        self.caja_cobro_propina = ""
        self.caja_cobro_descuento = ""
        self.caja_cobro_efectivo_recibido = ""
        self.caja_cobro_cliente_nombre = ""
        self.caja_cobro_cliente_id = 0
        self.caja_cobro_error = ""
        self.caja_cobro_items = []
        self.caja_cobro_dividido = False
        self.caja_pago_staged_metodo = "efectivo"
        self.caja_pago_staged_monto = ""
        self.caja_pagos_staged = []

    def set_caja_cobro_metodo(self, v: str) -> None:
        self.caja_cobro_metodo = v
        self.caja_cobro_efectivo_recibido = ""

    def set_caja_cobro_propina(self, v: str) -> None:
        self.caja_cobro_propina = v

    def set_caja_cobro_descuento(self, v: str) -> None:
        self.caja_cobro_descuento = v

    def set_caja_cobro_efectivo_recibido(self, v: str) -> None:
        self.caja_cobro_efectivo_recibido = v

    # ─── Caja — Cobro dividido / pago mixto ──────────────────────────────────

    @rx.var
    def caja_pagos_total(self) -> float:
        return round(sum(p.monto for p in self.caja_pagos_staged), 2)

    @rx.var
    def caja_pagos_restante(self) -> float:
        return round(max(self.caja_cobro_total_final - self.caja_pagos_total, 0.0), 2)

    @rx.var
    def caja_pagos_restante_texto(self) -> str:
        return _money_text(self.caja_pagos_restante)

    @rx.var
    def caja_pagos_cubierto(self) -> bool:
        return bool(self.caja_pagos_staged) and self.caja_pagos_total >= round(
            self.caja_cobro_total_final, 2
        )

    @rx.var
    def caja_pagos_vuelto_texto(self) -> str:
        vuelto = round(self.caja_pagos_total - self.caja_cobro_total_final, 2)
        return _money_text(vuelto) if vuelto > 0 else ""

    @rx.var
    def caja_pagos_tiene_fiado(self) -> bool:
        return any(p.metodo == "fiado" for p in self.caja_pagos_staged)

    def set_caja_cobro_dividido(self, v: bool) -> None:
        self.caja_cobro_dividido = bool(v)
        self.caja_pagos_staged = []
        self.caja_pago_staged_metodo = "efectivo"
        self.caja_pago_staged_monto = ""
        self.caja_cobro_error = ""

    def set_caja_pago_staged_metodo(self, v: str) -> None:
        self.caja_pago_staged_metodo = v

    def set_caja_pago_staged_monto(self, v: str) -> None:
        self.caja_pago_staged_monto = v

    def agregar_pago_staged(self) -> None:
        self.caja_cobro_error = ""
        raw = (self.caja_pago_staged_monto or "").replace(",", ".").strip()
        if raw:
            try:
                monto = round(float(raw), 2)
            except ValueError:
                self.caja_cobro_error = "Monto de pago inválido."
                return
        else:
            monto = self.caja_pagos_restante  # sin monto: completa lo que falta
        if monto <= 0:
            self.caja_cobro_error = "El pago debe ser mayor a cero."
            return
        labels = {"efectivo": "Efectivo", "tarjeta": "Tarjeta", "qr": "QR / Yape", "fiado": "Fiado / CC"}
        metodo = self.caja_pago_staged_metodo or "efectivo"
        self.caja_pagos_staged.append(PagoStagedView(
            metodo=metodo,
            metodo_label=labels.get(metodo, metodo),
            monto=monto,
            monto_texto=_money_text(monto),
        ))
        self.caja_pago_staged_monto = ""

    def quitar_pago_staged(self, idx: int) -> None:
        if 0 <= idx < len(self.caja_pagos_staged):
            pagos = list(self.caja_pagos_staged)
            pagos.pop(idx)
            self.caja_pagos_staged = pagos

    def confirmar_cobro(self) -> None:
        self.caja_cobro_error = ""
        objetivo = self.caja_cobro_mesa_id
        if objetivo == 0:
            self.caja_cobro_error = "No hay mesa seleccionada para cobrar."
            return
        metodo = self.caja_cobro_metodo or "efectivo"
        try:
            propina_raw = float(self.caja_cobro_propina.replace(",", ".").strip())
            propina = Decimal(str(round(propina_raw, 2))) if propina_raw > 0 else Decimal("0.00")
        except (ValueError, AttributeError, InvalidOperation):
            propina = Decimal("0.00")
        try:
            desc_raw = float(self.caja_cobro_descuento.replace(",", ".").strip())
            descuento = Decimal(str(round(max(desc_raw, 0.0), 2)))
        except (ValueError, AttributeError, InvalidOperation):
            descuento = Decimal("0.00")

        pedido_id = 0
        mesa_label = ""
        attended_by = ""
        total_base = Decimal("0.00")
        ticket_lines: list[TicketLine] = []
        with self._tenant_session() as session:
            mesa = session.get(Mesa, objetivo)
            if mesa is None:
                self.mensaje = "La mesa indicada ya no existe."
                return
            pedido = _get_open_order(session, mesa.id or 0, self._company_id())
            if pedido is None:
                self.mensaje = "No hay pedido abierto para esa mesa."
                return
            if _get_unsent_details(session, pedido.id or 0):
                self.mensaje = "Todavia hay items pendientes de enviar a cocina."
                return
            if _get_not_delivered_details(session, pedido.id or 0):
                self.mensaje = "Todavia hay items en cocina o listos por entregar."
                return
            turno = get_turno_abierto(session, self._company_id())
            if turno is None:
                self.caja_cobro_error = "No hay turno de caja abierto. Abre el turno antes de cobrar."
                return
            detalles = session.exec(select(DetallePedido).where(DetallePedido.pedido_id == pedido.id)).all()
            productos = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            usuarios = {u.id: u for u in session.exec(select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())).all()}
            ticket_lines = [
                TicketLine(
                    name=(productos[d.producto_id].nombre if d.producto_id in productos else f"Producto {d.producto_id}"),
                    quantity=d.cantidad,
                    unit_price=float(_to_decimal(d.precio_unitario)),
                    subtotal=float(_to_decimal(d.subtotal)),
                    note=d.notas or "",
                )
                for d in detalles
            ]
            mozo = usuarios.get(pedido.mozo_id)
            attended_by = _actor_name(
                mozo.nombre if mozo else (self.usuario_actual.nombre if self.usuario_actual else "")
            ) or "Sin asignar"
            total_base = _to_decimal(pedido.total)
            if descuento > total_base:
                self.caja_cobro_error = f"El descuento ({_money_text(descuento)}) no puede superar el total ({_money_text(total_base)})."
                return
            # Construir la lista de pagos: dividido/mixto usa la lista armada
            # en la UI; el cobro simple se normaliza a un único pago para que
            # todo quede registrado igual en food_pagos_pedido.
            if self.caja_cobro_dividido:
                if not self.caja_pagos_staged:
                    self.caja_cobro_error = "Agrega al menos un pago."
                    return
                total_final = max(total_base - descuento + propina, Decimal("0.00"))
                pagos_lista = [
                    (p.metodo, Decimal(str(round(p.monto, 2))))
                    for p in self.caja_pagos_staged
                ]
            else:
                if metodo == "fiado":
                    propina = Decimal("0.00")  # el fiado no registra propina
                total_final = max(total_base - descuento + propina, Decimal("0.00"))
                pagos_lista = [(metodo, total_final)] if total_final > 0 else []
            resultado_pagos = None
            if pagos_lista:
                try:
                    resultado_pagos = validar_pagos(total_final, pagos_lista)
                except ValueError as exc:
                    self.caja_cobro_error = str(exc)
                    return
            total_fiado = resultado_pagos.total_fiado if resultado_pagos else Decimal("0.00")
            if total_fiado > 0 and self.caja_cobro_cliente_id <= 0:
                self.caja_cobro_error = "Selecciona el cliente para registrar el fiado."
                return
            now = _utcnow()
            if self.usuario_actual:
                pedido.cajero_id = self.usuario_actual.id
            pedido.pagado = total_fiado == 0
            pedido.estado = EstadoPedido.COBRADO.value
            pedido.cerrado_en = now
            pedido.updated_at = now
            pedido.metodo_pago = metodo_pago_resumen(pagos_lista) if pagos_lista else metodo
            pedido.turno_caja_id = turno.id
            pedido.propina = propina
            pedido.descuento = descuento
            if self.caja_cobro_cliente_id > 0:
                pedido.cliente_id = self.caja_cobro_cliente_id
            session.add(pedido)
            mesa.estado = EstadoMesa.LIBRE.value
            mesa.updated_at = now
            session.add(mesa)
            _descontar_stock_por_pedido(session, pedido.id or 0, self._company_id())
            if total_fiado > 0:
                try:
                    self._registrar_cargo_cc(
                        session,
                        self.caja_cobro_cliente_id,
                        total_fiado,
                        pedido.id,
                        f"Fiado mesa {mesa.nombre or str(mesa.numero)}",
                    )
                except ValueError as exc:
                    self.caja_cobro_error = str(exc)
                    return
            if resultado_pagos is not None:
                registrar_pagos_pedido(
                    session,
                    pedido,
                    turno.id,
                    self.usuario_actual.id if self.usuario_actual else None,
                    pagos_lista,
                    resultado_pagos,
                )
            metodo_final = pedido.metodo_pago or metodo
            session.commit()
            pedido_id = pedido.id or 0
            mesa_label = mesa.nombre or f"Mesa {mesa.numero}"

        if self.mesa_seleccionada_id == objetivo:
            self.mesa_seleccionada_id = 0
            self.carrito = []
            self.historial_pedido = []
        self.cancelar_cobro()
        self.cargar_mesas()
        self.cargar_historial_ventas()
        total_final = max(total_base - descuento + propina, Decimal("0.00"))
        html_ticket = generate_cashier_ticket_html(
            order_reference=mesa_label,
            pedido_id=pedido_id,
            items=ticket_lines,
            total=float(total_final),
            attended_by=attended_by,
            company_name=self.config_nombre_local or "TUWAYKIFOOD",
            paper_width_mm=self._ticket_paper_width_mm(),
        )
        desc_txt = f" - descuento {_money_text(descuento)}" if descuento > 0 else ""
        propina_txt = f" + propina {_money_text(propina)}" if propina > 0 else ""
        self.mensaje = f"{mesa_label} cobrada ({metodo_final}). Total: {_money_text(total_final)}{desc_txt}{propina_txt}."
        return rx.call_script(build_print_script(html_ticket))

    # ─── Caja — Cobro de mesa ─────────────────────────────────────────────────

    def cobrar_mesa(self, mesa_id: int) -> None:
        objetivo = mesa_id or self.mesa_seleccionada_id
        if objetivo == 0:
            self.mensaje = "Selecciona una mesa antes de cobrar."
            return
        pedido_id = 0
        mesa_label = ""
        attended_by = ""
        total = 0.0
        ticket_lines: list[TicketLine] = []
        with self._tenant_session() as session:
            mesa = session.get(Mesa, objetivo)
            if mesa is None:
                self.mensaje = "La mesa indicada ya no existe."
                return
            pedido = _get_open_order(session, mesa.id or 0, self._company_id())
            if pedido is None:
                self.mensaje = "No hay pedido abierto para esa mesa."
                return
            if _get_unsent_details(session, pedido.id or 0):
                self.mensaje = "Todavia hay items pendientes de enviar a cocina."
                return
            if _get_not_delivered_details(session, pedido.id or 0):
                self.mensaje = "Todavia hay items en cocina o listos por entregar."
                return
            turno = get_turno_abierto(session, self._company_id())
            if turno is None:
                self.mensaje = "No hay turno de caja abierto. Abre el turno antes de cobrar."
                return
            detalles = session.exec(select(DetallePedido).where(DetallePedido.pedido_id == pedido.id)).all()
            productos = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            usuarios = {u.id: u for u in session.exec(select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())).all()}
            ticket_lines = [
                TicketLine(
                    name=(productos[d.producto_id].nombre if d.producto_id in productos else f"Producto {d.producto_id}"),
                    quantity=d.cantidad,
                    unit_price=float(_to_decimal(d.precio_unitario)),
                    subtotal=float(_to_decimal(d.subtotal)),
                    note=d.notas or "",
                )
                for d in detalles
            ]
            mozo = usuarios.get(pedido.mozo_id)
            attended_by = _actor_name(
                mozo.nombre if mozo else (self.usuario_actual.nombre if self.usuario_actual else "")
            ) or "Sin asignar"
            total = float(_to_decimal(pedido.total))
            now = _utcnow()
            if self.usuario_actual:
                pedido.cajero_id = self.usuario_actual.id
            pedido.pagado = True
            pedido.estado = EstadoPedido.COBRADO.value
            pedido.cerrado_en = now
            pedido.updated_at = now
            pedido.turno_caja_id = turno.id
            session.add(pedido)
            mesa.estado = EstadoMesa.LIBRE.value
            mesa.updated_at = now
            session.add(mesa)
            _descontar_stock_por_pedido(session, pedido.id or 0, self._company_id())
            if total > 0:
                session.add(PagoPedido(
                    company_id=self._company_id(),
                    pedido_id=pedido.id or 0,
                    turno_caja_id=turno.id,
                    usuario_id=self.usuario_actual.id if self.usuario_actual else None,
                    metodo="efectivo",
                    monto=Decimal(str(round(total, 2))),
                ))
            session.commit()
            pedido_id = pedido.id or 0
            mesa_label = mesa.nombre or f"Mesa {mesa.numero}"
        if self.mesa_seleccionada_id == objetivo:
            self.mesa_seleccionada_id = 0
            self.carrito = []
            self.historial_pedido = []
        self.cargar_mesas()
        self.cargar_historial_ventas()
        html_ticket = generate_cashier_ticket_html(
            order_reference=mesa_label,
            pedido_id=pedido_id,
            items=ticket_lines,
            total=total,
            attended_by=attended_by,
            company_name=self.config_nombre_local or "TUWAYKIFOOD",
            paper_width_mm=self._ticket_paper_width_mm(),
        )
        self.mensaje = f"{mesa_label} cobrada. Total: {_money_text(total)}."
        return rx.call_script(build_print_script(html_ticket))

    # ─── Mostrador ────────────────────────────────────────────────────────────

    def set_mostrador_cliente_nombre(self, value: str) -> None:
        self.mostrador_cliente_nombre = str(value)[:120]

    def agregar_producto_mostrador(self, producto_id: int) -> None:
        producto = next((p for p in self.productos if p.id == producto_id and p.disponible), None)
        if producto is None:
            self.mensaje = "Producto no disponible para mostrador."
            return
        carrito = list(self.mostrador_carrito)
        for i, item in enumerate(carrito):
            if item.producto_id == producto_id:
                cantidad = item.cantidad + 1
                subtotal = round(producto.precio * cantidad, 2)
                carrito[i] = CarritoItem(
                    producto_id=producto.id,
                    nombre=producto.nombre,
                    cantidad=cantidad,
                    precio_unitario=producto.precio,
                    subtotal=subtotal,
                    subtotal_texto=_money_text(subtotal),
                )
                self.mostrador_carrito = carrito
                self.mensaje = f"{producto.nombre} agregado a mostrador."
                return
        carrito.append(CarritoItem(
            producto_id=producto.id,
            nombre=producto.nombre,
            cantidad=1,
            precio_unitario=producto.precio,
            subtotal=producto.precio,
            subtotal_texto=producto.precio_texto,
        ))
        self.mostrador_carrito = carrito
        self.mensaje = f"{producto.nombre} agregado a mostrador."

    def restar_producto_mostrador(self, producto_id: int) -> None:
        carrito_actualizado: list[CarritoItem] = []
        encontrado = False
        for item in self.mostrador_carrito:
            if item.producto_id != producto_id:
                carrito_actualizado.append(item)
                continue
            encontrado = True
            cantidad = item.cantidad - 1
            if cantidad > 0:
                subtotal = round(item.precio_unitario * cantidad, 2)
                carrito_actualizado.append(CarritoItem(
                    producto_id=item.producto_id,
                    nombre=item.nombre,
                    cantidad=cantidad,
                    precio_unitario=item.precio_unitario,
                    subtotal=subtotal,
                    subtotal_texto=_money_text(subtotal),
                ))
        if not encontrado:
            self.mensaje = "Ese producto no esta en el carrito de mostrador."
            return
        self.mostrador_carrito = carrito_actualizado
        self.mensaje = "Carrito de mostrador actualizado."

    def limpiar_carrito_mostrador(self) -> None:
        self.mostrador_carrito = []
        self.mostrador_metodo_pago = "efectivo"
        self.mensaje = "Carrito de mostrador limpio."

    def seleccionar_mostrador_metodo(self, metodo: str) -> None:
        self.mostrador_metodo_pago = metodo

    def cobrar_y_enviar_mostrador(self) -> None:
        if not self.mostrador_carrito:
            self.mensaje = "Agrega productos antes de cobrar en mostrador."
            return
        if self.usuario_actual is None:
            self.mensaje = "Inicia sesion para registrar la venta de mostrador."
            return
        pedido_id = 0
        total = 0.0
        cliente_nombre = _actor_name(self.mostrador_cliente_nombre) or "Sin nombre"
        ticket_label = f"Para Llevar - Cliente: {cliente_nombre}"
        attended_by = _actor_name(self.usuario_actual.nombre if self.usuario_actual else "") or "Sin asignar"
        ticket_lines: list[TicketLine] = []
        with self._tenant_session() as session:
            productos = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            invalidos = [item.nombre for item in self.mostrador_carrito if item.producto_id not in productos or not productos[item.producto_id].disponible]
            if invalidos:
                self.mensaje = f"Productos no disponibles: {', '.join(invalidos)}"
                return
            errores_stock = _validar_stock_para_items(
                session, [(item.producto_id, item.cantidad) for item in self.mostrador_carrito], self._company_id()
            )
            if errores_stock:
                self.mensaje = "Stock insuficiente — " + "; ".join(errores_stock)
                return
            turno = get_turno_abierto(session, self._company_id())
            if turno is None:
                self.mensaje = "No hay turno de caja abierto. Abre el turno antes de cobrar."
                return
            now = _utcnow()
            pedido = Pedido(
                company_id=self._company_id(),
                mesa_id=None,
                cajero_id=self.usuario_actual.id,
                tipo_pedido=TipoPedido.MOSTRADOR.value,
                nombre_cliente=_actor_name(self.mostrador_cliente_nombre) or None,
                pagado=True,
                estado=EstadoPedido.ENVIADO.value,
                metodo_pago=self.mostrador_metodo_pago,
                total=Decimal("0.00"),
                abierto_en=now,
                cerrado_en=now,
                turno_caja_id=turno.id,
            )
            session.add(pedido)
            session.commit()
            session.refresh(pedido)
            for item in self.mostrador_carrito:
                producto = productos[item.producto_id]
                precio = _to_decimal(producto.precio)
                subtotal = precio * item.cantidad
                detalle = DetallePedido(
                    company_id=self._company_id(),
                    pedido_id=pedido.id or 0,
                    producto_id=producto.id or 0,
                    cantidad=item.cantidad,
                    precio_unitario=precio,
                    subtotal=subtotal,
                    estado_produccion=EstadoProduccion.PENDIENTE.value,
                    impreso_cocina=True,
                    impreso_caja=True,
                    enviado_cocina_at=now,
                )
                session.add(detalle)
                ticket_lines.append(TicketLine(
                    name=producto.nombre,
                    quantity=item.cantidad,
                    unit_price=float(precio),
                    subtotal=float(subtotal),
                    note="",
                ))
            total = float(_recalculate_order_total(session, pedido))
            _sync_order_status(session, pedido)
            session.add(pedido)
            _descontar_stock_por_pedido(session, pedido.id or 0, self._company_id())
            if total > 0:
                session.add(PagoPedido(
                    company_id=self._company_id(),
                    pedido_id=pedido.id or 0,
                    turno_caja_id=turno.id,
                    usuario_id=self.usuario_actual.id,
                    metodo=self.mostrador_metodo_pago or "efectivo",
                    monto=Decimal(str(round(total, 2))),
                ))
            session.commit()
            pedido_id = pedido.id or 0
        self.ultimo_pedido_id = pedido_id
        self.mostrador_carrito = []
        self.mostrador_cliente_nombre = ""
        self.mostrador_metodo_pago = "efectivo"
        self.cargar_cocina()
        self.cargar_historial_ventas()
        paper_width_mm = self._ticket_paper_width_mm()
        html_cocina = generate_kitchen_ticket_html(
            mesa_label=ticket_label, pedido_id=pedido_id, items=ticket_lines,
            paper_width_mm=paper_width_mm,
        )
        html_caja = generate_cashier_ticket_html(
            order_reference=f"Cliente: {cliente_nombre}", pedido_id=pedido_id,
            items=ticket_lines, total=total, attended_by=attended_by,
            company_name=self.config_nombre_local or "TUWAYKIFOOD",
            paper_width_mm=paper_width_mm,
        )
        self.mensaje = f"Pedido de mostrador #{pedido_id} cobrado y enviado."
        return rx.call_script(build_print_script(html_cocina) + build_print_script(html_caja))

    def cargar_pedidos_mostrador_listos(self) -> None:
        with self._tenant_session() as session:
            detalles = session.exec(
                select(DetallePedido).where(
                    DetallePedido.company_id == self._company_id(),
                    DetallePedido.impreso_cocina.is_(True),
                    DetallePedido.estado_produccion == EstadoProduccion.LISTO_PARA_ENTREGAR.value,
                ).order_by(DetallePedido.enviado_cocina_at, DetallePedido.id)
            ).all()
            if not detalles:
                self.pedidos_mostrador_listos = []
                return
            pedido_ids = {d.pedido_id for d in detalles}
            pedidos = {
                p.id: p
                for p in session.exec(select(Pedido).where(Pedido.id.in_(pedido_ids))).all()
                if p.tipo_pedido == TipoPedido.MOSTRADOR.value
            }
            productos = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            grupos: dict = {}
            for d in detalles:
                pedido = pedidos.get(d.pedido_id)
                if pedido is None:
                    continue
                marca = d.updated_at or d.enviado_cocina_at or d.created_at
                if pedido.id not in grupos:
                    grupos[pedido.id] = {
                        "pedido_id": pedido.id or 0,
                        "cliente_nombre": _actor_name(pedido.nombre_cliente) or "Sin nombre",
                        "hora_texto": marca.strftime("%H:%M"),
                        "items_lines": [],
                        "items_count": 0,
                    }
                producto = productos.get(d.producto_id)
                grupos[pedido.id]["items_lines"].append(f"{d.cantidad} x {producto.nombre if producto else f'Producto {d.producto_id}'}")
                grupos[pedido.id]["items_count"] += d.cantidad
            self.pedidos_mostrador_listos = [
                MostradorEntregaView(**data) for data in grupos.values()
            ]

    def cargar_pedidos_mostrador_entregados(self) -> None:
        with self._tenant_session() as session:
            pedidos = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    Pedido.tipo_pedido == TipoPedido.MOSTRADOR.value,
                    Pedido.pagado.is_(True),
                ).order_by(Pedido.updated_at.desc(), Pedido.id.desc())
            ).all()
            if not pedidos:
                self.pedidos_mostrador_entregados = []
                return
            productos = {p.id: p for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()}
            historial: list = []
            for pedido in pedidos:
                detalles = session.exec(
                    select(DetallePedido).where(DetallePedido.pedido_id == pedido.id).order_by(DetallePedido.id)
                ).all()
                if not detalles or any(d.estado_produccion != EstadoProduccion.ENTREGADO_AL_CLIENTE.value for d in detalles):
                    continue
                entregado_en = max(d.updated_at for d in detalles)
                resumen = " · ".join(
                    f"{d.cantidad}x {productos[d.producto_id].nombre if d.producto_id in productos else f'Producto {d.producto_id}'}"
                    for d in detalles
                )
                historial.append((
                    entregado_en,
                    MostradorEntregadoView(
                        pedido_id=pedido.id or 0,
                        cliente_nombre=_actor_name(pedido.nombre_cliente) or "Sin nombre",
                        hora_texto=entregado_en.strftime("%H:%M"),
                        items_resumen=resumen,
                        total_texto=_money_text(pedido.total),
                    ),
                ))
            historial.sort(key=lambda x: x[0], reverse=True)
            self.pedidos_mostrador_entregados = [item for _, item in historial[:10]]

    def entregar_pedido_mostrador(self, pedido_id: int) -> None:
        with self._tenant_session() as session:
            pedido = session.get(Pedido, pedido_id)
            if pedido is None or pedido.tipo_pedido != TipoPedido.MOSTRADOR.value or pedido.company_id != self._company_id():
                self.mensaje = "El pedido de mostrador ya no existe."
                return
            detalles_listos = session.exec(
                select(DetallePedido).where(
                    DetallePedido.pedido_id == pedido_id,
                    DetallePedido.impreso_cocina.is_(True),
                    DetallePedido.estado_produccion == EstadoProduccion.LISTO_PARA_ENTREGAR.value,
                )
            ).all()
            if not detalles_listos:
                self.mensaje = "Ese pedido ya no tiene items listos para entregar."
                return
            now = _utcnow()
            for d in detalles_listos:
                d.estado_produccion = EstadoProduccion.ENTREGADO_AL_CLIENTE.value
                d.updated_at = now
                session.add(d)
            _sync_order_status(session, pedido)
            session.add(pedido)
            session.commit()
        self.cargar_cocina()
        self.cargar_pedidos_mostrador_listos()
        self.cargar_pedidos_mostrador_entregados()
        self.cargar_historial_ventas()
        self.mensaje = "Pedido de mostrador entregado al cliente."

    # ─── Dashboard KPIs ───────────────────────────────────────────────────────

    def cargar_dashboard(self) -> None:
        hoy = _utcnow().date()
        inicio_hoy = datetime(hoy.year, hoy.month, hoy.day)
        fin_hoy = inicio_hoy + timedelta(days=1)
        inicio_ayer = inicio_hoy - timedelta(days=1)
        with self._tenant_session() as session:
            pedidos_hoy = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                    Pedido.cerrado_en >= inicio_hoy,
                    Pedido.cerrado_en < fin_hoy,
                )
            ).all()
            pedidos_ayer = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                    Pedido.cerrado_en >= inicio_ayer,
                    Pedido.cerrado_en < inicio_hoy,
                )
            ).all()
            ventas = sum(
                _to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0))
                for p in pedidos_hoy
            )
            propina_total = sum(_to_decimal(getattr(p, "propina", 0)) for p in pedidos_hoy)
            ventas_ayer = sum(
                _to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0))
                for p in pedidos_ayer
            )
            propina_ayer = sum(_to_decimal(getattr(p, "propina", 0)) for p in pedidos_ayer)
            mesas_no_libres = session.exec(
                select(Mesa).where(
                    Mesa.company_id == self._company_id(),
                    Mesa.estado != EstadoMesa.LIBRE.value,
                    Mesa.activa.is_(True),
                )
            ).all()
            pedido_ids_hoy = {p.id for p in pedidos_hoy}
            top_data: dict[int, dict] = {}
            if pedido_ids_hoy:
                detalles = session.exec(
                    select(DetallePedido).where(DetallePedido.pedido_id.in_(pedido_ids_hoy))
                ).all()
                productos = {
                    p.id: p
                    for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()
                }
                for d in detalles:
                    pid = d.producto_id
                    if pid not in top_data:
                        prod = productos.get(pid)
                        top_data[pid] = {
                            "nombre": prod.nombre if prod else f"Producto {pid}",
                            "cantidad": 0,
                            "total": Decimal("0.00"),
                        }
                    top_data[pid]["cantidad"] += d.cantidad
                    top_data[pid]["total"] += _to_decimal(d.subtotal)
            top_sorted = sorted(top_data.values(), key=lambda x: x["cantidad"], reverse=True)[:5]

        def _pct_change(hoy_val: Decimal, ayer_val: Decimal) -> int:
            if ayer_val <= 0:
                return 100 if hoy_val > 0 else 0
            return int(round((hoy_val - ayer_val) / ayer_val * 100))

        pedidos_count_hoy = len(pedidos_hoy)
        pedidos_count_ayer = len(pedidos_ayer)
        ticket_promedio = ventas / pedidos_count_hoy if pedidos_count_hoy else Decimal("0.00")
        ticket_promedio_ayer = (
            ventas_ayer / pedidos_count_ayer if pedidos_count_ayer else Decimal("0.00")
        )
        self.dashboard_ventas_hoy_texto = _money_text(ventas)
        self.dashboard_pedidos_hoy = pedidos_count_hoy
        self.dashboard_mesas_ocupadas = len(mesas_no_libres)
        self.dashboard_propina_hoy_texto = _money_text(propina_total)
        self.dashboard_ticket_promedio_texto = _money_text(ticket_promedio)
        self.dashboard_ventas_trend_pct = _pct_change(ventas, ventas_ayer)
        self.dashboard_pedidos_trend = pedidos_count_hoy - pedidos_count_ayer
        self.dashboard_ticket_trend_pct = _pct_change(ticket_promedio, ticket_promedio_ayer)
        self.dashboard_propina_trend_pct = _pct_change(propina_total, propina_ayer)
        self.dashboard_top_platos = [
            TopPlatoView(
                nombre=p["nombre"],
                cantidad=p["cantidad"],
                total_generado=float(p["total"]),
                total_texto=_money_text(p["total"]),
            )
            for p in top_sorted
        ]

    # ─── Historial de ventas ──────────────────────────────────────────────────

    def cargar_historial_ventas(self) -> None:
        with self._tenant_session() as session:
            query = (
                select(Pedido)
                .where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                )
            )
            if self.historial_filtro_fecha_desde:
                try:
                    desde = datetime.strptime(self.historial_filtro_fecha_desde, "%Y-%m-%d")
                    query = query.where(Pedido.cerrado_en >= desde)
                except ValueError:
                    pass
            if self.historial_filtro_fecha_hasta:
                try:
                    hasta = datetime.strptime(self.historial_filtro_fecha_hasta, "%Y-%m-%d")
                    hasta = hasta.replace(hour=23, minute=59, second=59)
                    query = query.where(Pedido.cerrado_en <= hasta)
                except ValueError:
                    pass
            if self.historial_filtro_metodo:
                query = query.where(Pedido.metodo_pago == self.historial_filtro_metodo)
            query = query.order_by(Pedido.cerrado_en.desc(), Pedido.id.desc())
            total_count = len(session.exec(query).all())
            self.historial_total = total_count
            offset = self.historial_pagina * self._HISTORIAL_PAGE_SIZE
            pedidos = session.exec(query.offset(offset).limit(self._HISTORIAL_PAGE_SIZE)).all()
            mesas = {m.id: m for m in session.exec(select(Mesa).where(Mesa.company_id == self._company_id())).all()}
            usuarios = {u.id: u for u in session.exec(select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())).all()}
            historial: list[VentaHistorialView] = []
            for p in pedidos:
                total_base = _to_decimal(p.total)
                propina = _to_decimal(getattr(p, "propina", Decimal("0.00")))
                total_con_propina = total_base + propina
                historial.append(VentaHistorialView(
                    pedido_id=p.id or 0,
                    mesa_label=_pedido_sales_label(p, mesas),
                    total=float(total_base),
                    total_texto=_money_text(total_base),
                    propina=float(propina),
                    propina_texto=_money_text(propina) if propina > 0 else "",
                    total_con_propina=float(total_con_propina),
                    total_con_propina_texto=_money_text(total_con_propina),
                    metodo_pago=getattr(p, "metodo_pago", None) or "—",
                    mozo_nombre=_actor_name(usuarios[p.mozo_id].nombre if p.mozo_id in usuarios else "Sin asignar"),
                    cajero_nombre=_actor_name(usuarios[p.cajero_id].nombre if p.cajero_id in usuarios else "Sin asignar"),
                ))
            self.historial_ventas = historial

    def set_historial_filtro_fecha_desde(self, v: str) -> None:
        self.historial_filtro_fecha_desde = v

    def set_historial_filtro_fecha_hasta(self, v: str) -> None:
        self.historial_filtro_fecha_hasta = v

    def set_historial_filtro_metodo(self, v: str) -> None:
        self.historial_filtro_metodo = v

    def aplicar_filtros_historial(self) -> None:
        self.historial_pagina = 0
        self.cargar_historial_ventas()

    def buscar_historial_manual(self) -> None:
        self.historial_filtro_rapido = "personalizado"
        self.aplicar_filtros_historial()

    def limpiar_filtros_historial(self) -> None:
        self.historial_filtro_fecha_desde = ""
        self.historial_filtro_fecha_hasta = ""
        self.historial_filtro_metodo = ""
        self.historial_filtro_rapido = ""
        self.historial_pagina = 0
        self.cargar_historial_ventas()

    def filtro_rapido_hoy(self) -> None:
        hoy = _utcnow().date().isoformat()
        self.historial_filtro_fecha_desde = hoy
        self.historial_filtro_fecha_hasta = hoy
        self.historial_filtro_rapido = "hoy"
        self.aplicar_filtros_historial()

    def filtro_rapido_semana(self) -> None:
        hoy = _utcnow().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        self.historial_filtro_fecha_desde = inicio_semana.isoformat()
        self.historial_filtro_fecha_hasta = hoy.isoformat()
        self.historial_filtro_rapido = "semana"
        self.aplicar_filtros_historial()

    def filtro_rapido_mes(self) -> None:
        hoy = _utcnow().date()
        inicio_mes = hoy.replace(day=1)
        self.historial_filtro_fecha_desde = inicio_mes.isoformat()
        self.historial_filtro_fecha_hasta = hoy.isoformat()
        self.historial_filtro_rapido = "mes"
        self.aplicar_filtros_historial()

    def abrir_detalle_venta(self, pedido_id: int) -> None:
        venta = next((v for v in self.historial_ventas if v.pedido_id == pedido_id), None)
        if venta is None:
            return
        with self._tenant_session() as session:
            detalles = session.exec(
                select(DetallePedido).where(
                    DetallePedido.company_id == self._company_id(),
                    DetallePedido.pedido_id == pedido_id,
                )
            ).all()
            productos = {
                p.id: p for p in session.exec(
                    select(Producto).where(Producto.company_id == self._company_id())
                ).all()
            }
        self.venta_detalle_items = [
            VentaDetalleItemView(
                nombre=productos[d.producto_id].nombre if d.producto_id in productos else f"Producto {d.producto_id}",
                cantidad=d.cantidad,
                precio_unitario_texto=_money_text(d.precio_unitario),
                subtotal_texto=_money_text(d.subtotal),
                notas=d.notas or "",
            )
            for d in detalles
        ]
        self.venta_detalle_pedido_id = pedido_id
        self.venta_detalle_mesa_label = venta.mesa_label
        self.venta_detalle_metodo = venta.metodo_pago
        self.venta_detalle_mozo = venta.mozo_nombre
        self.venta_detalle_cajero = venta.cajero_nombre
        self.venta_detalle_total_texto = venta.total_con_propina_texto
        self.venta_detalle_propina_texto = venta.propina_texto
        self.venta_detalle_visible = True

    def set_venta_detalle_visible(self, v: bool) -> None:
        self.venta_detalle_visible = v

    def historial_pagina_anterior(self) -> None:
        if self.historial_pagina > 0:
            self.historial_pagina -= 1
            self.cargar_historial_ventas()

    def historial_pagina_siguiente(self) -> None:
        if self.historial_tiene_siguiente:
            self.historial_pagina += 1
            self.cargar_historial_ventas()

    def exportar_ventas_excel(self):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        with self._tenant_session() as session:
            query = (
                select(Pedido)
                .where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                )
            )
            if self.historial_filtro_fecha_desde:
                try:
                    desde = datetime.strptime(self.historial_filtro_fecha_desde, "%Y-%m-%d")
                    query = query.where(Pedido.cerrado_en >= desde)
                except ValueError:
                    pass
            if self.historial_filtro_fecha_hasta:
                try:
                    hasta = datetime.strptime(self.historial_filtro_fecha_hasta, "%Y-%m-%d")
                    hasta = hasta.replace(hour=23, minute=59, second=59)
                    query = query.where(Pedido.cerrado_en <= hasta)
                except ValueError:
                    pass
            if self.historial_filtro_metodo:
                query = query.where(Pedido.metodo_pago == self.historial_filtro_metodo)
            query = query.order_by(Pedido.cerrado_en.desc(), Pedido.id.desc())
            pedidos = session.exec(query).all()
            mesas = {m.id: m for m in session.exec(
                select(Mesa).where(Mesa.company_id == self._company_id())
            ).all()}
            usuarios = {u.id: u for u in session.exec(
                select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())
            ).all()}
            pedido_ids = [p.id for p in pedidos if p.id is not None]
            detalles_por_pedido: dict[int, list] = {}
            productos_map: dict[int, Producto] = {}
            if pedido_ids:
                detalles = session.exec(
                    select(DetallePedido).where(DetallePedido.pedido_id.in_(pedido_ids))
                ).all()
                for d in detalles:
                    detalles_por_pedido.setdefault(d.pedido_id, []).append(d)
                productos_map = {
                    pr.id: pr for pr in session.exec(
                        select(Producto).where(Producto.company_id == self._company_id())
                    ).all()
                }

        if not pedidos:
            self.mensaje = "No hay ventas para exportar con estos filtros."
            return None

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Ventas"
        ws1.append([
            "Fecha", "Hora", "Pedido #", "Mesa", "Método de pago",
            "Mozo", "Cajero", "Subtotal", "Propina", "Total",
        ])
        for p in pedidos:
            fecha = p.cerrado_en
            subtotal = float(_to_decimal(p.total))
            propina = float(_to_decimal(getattr(p, "propina", 0)))
            mozo = usuarios.get(p.mozo_id)
            cajero = usuarios.get(p.cajero_id)
            ws1.append([
                fecha.strftime("%Y-%m-%d") if fecha else "",
                fecha.strftime("%H:%M") if fecha else "",
                p.id,
                _pedido_sales_label(p, mesas),
                getattr(p, "metodo_pago", None) or "",
                _actor_name(mozo.nombre) if mozo else "Sin asignar",
                _actor_name(cajero.nombre) if cajero else "Sin asignar",
                subtotal, propina, subtotal + propina,
            ])

        ws2 = wb.create_sheet("Detalle de items")
        ws2.append(["Fecha", "Pedido #", "Mesa", "Producto", "Cantidad",
                     "Precio unitario", "Subtotal"])
        for p in pedidos:
            fecha = p.cerrado_en
            mesa_label = _pedido_sales_label(p, mesas)
            for d in detalles_por_pedido.get(p.id or 0, []):
                prod = productos_map.get(d.producto_id)
                ws2.append([
                    fecha.strftime("%Y-%m-%d") if fecha else "",
                    p.id,
                    mesa_label,
                    prod.nombre if prod else f"Producto {d.producto_id}",
                    d.cantidad,
                    float(_to_decimal(d.precio_unitario)),
                    float(_to_decimal(d.subtotal)),
                ])

        for ws in (ws1, ws2):
            for i, col in enumerate(ws.columns, start=1):
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"ventas_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    # ─── Admin Carta — Categorías ─────────────────────────────────────────────

    def set_categoria_form_nombre(self, v: str) -> None:
        self.categoria_form_nombre = str(v)[:120]

    def set_categoria_form_descripcion(self, v: str) -> None:
        self.categoria_form_descripcion = str(v)[:240]

    def set_categoria_form_orden(self, v: str) -> None:
        self.categoria_form_orden = v

    def guardar_categoria(self) -> None:
        nombre = self.categoria_form_nombre.strip()
        if not nombre:
            self.mensaje = "El nombre de la categoria es obligatorio."
            return
        try:
            orden = int(self.categoria_form_orden)
        except ValueError:
            orden = len(self.categorias) + 1
        with self._tenant_session() as session:
            if self.categoria_form_id:
                cat = session.get(Categoria, self.categoria_form_id)
                if cat is None or cat.company_id != self._company_id():
                    self.mensaje = "Categoria no encontrada."
                    return
                cat.nombre = nombre
                cat.descripcion = self.categoria_form_descripcion.strip() or None
                cat.orden = orden
                cat.updated_at = _utcnow()
                session.add(cat)
            else:
                cat = Categoria(
                    company_id=self._company_id(),
                    nombre=nombre,
                    descripcion=self.categoria_form_descripcion.strip() or None,
                    orden=orden,
                )
                session.add(cat)
            session.commit()
        self.cargar_menu()
        self._reset_categoria_form()
        self.mensaje = "Categoria guardada."

    def editar_categoria(self, categoria_id: int) -> None:
        cat = next((c for c in self.categorias if c.id == categoria_id), None)
        if cat is None:
            return
        self.categoria_form_id = cat.id
        self.categoria_form_nombre = cat.nombre
        self.categoria_form_descripcion = cat.descripcion
        self.categoria_form_orden = str(cat.orden)

    def toggle_categoria_activa(self, categoria_id: int) -> None:
        with self._tenant_session() as session:
            cat = session.get(Categoria, categoria_id)
            if cat is None or cat.company_id != self._company_id():
                return
            cat.activa = not cat.activa
            cat.updated_at = _utcnow()
            session.add(cat)
            session.commit()
        self.cargar_menu()

    def _reset_categoria_form(self) -> None:
        self.categoria_form_id = 0
        self.categoria_form_nombre = ""
        self.categoria_form_descripcion = ""
        self.categoria_form_orden = str(len(self.categorias) + 1)

    def cancelar_categoria_form(self) -> None:
        self._reset_categoria_form()

    # ─── Admin Carta — Productos ──────────────────────────────────────────────

    def set_producto_form_nombre(self, v: str) -> None:
        self.producto_form_nombre = str(v)[:160]

    def set_producto_form_descripcion(self, v: str) -> None:
        self.producto_form_descripcion = str(v)[:240]

    def set_producto_form_precio(self, v: str) -> None:
        self.producto_form_precio = v

    def set_producto_form_categoria(self, v: str) -> None:
        self.producto_form_categoria_nombre = v

    def set_producto_form_disponible(self, v: bool) -> None:
        self.producto_form_disponible = v

    def set_producto_form_emoji(self, v: str) -> None:
        self.producto_form_emoji = str(v)[:8]

    @rx.var
    def producto_form_emoji_sugerido(self) -> str:
        return _emoji_para_producto(self.producto_form_nombre or "")

    def guardar_producto(self) -> None:
        nombre = self.producto_form_nombre.strip()
        if not nombre:
            self.mensaje = "El nombre del producto es obligatorio."
            return
        precio = _parse_positive_price(self.producto_form_precio)
        if precio is None:
            self.mensaje = "El precio debe ser un numero mayor a 0."
            return
        with self._tenant_session() as session:
            cat = session.exec(
                select(Categoria).where(
                    Categoria.company_id == self._company_id(),
                    Categoria.nombre == self.producto_form_categoria_nombre,
                )
            ).first()
            if cat is None:
                self.mensaje = f"Categoria '{self.producto_form_categoria_nombre}' no encontrada."
                return
            if self.producto_form_id:
                prod = session.get(Producto, self.producto_form_id)
                if prod is None or prod.company_id != self._company_id():
                    self.mensaje = "Producto no encontrado."
                    return
                prod.nombre = nombre
                prod.descripcion = self.producto_form_descripcion.strip() or None
                prod.precio = precio
                prod.categoria_id = cat.id or 0
                prod.disponible = self.producto_form_disponible
                prod.imagen_url = self.producto_form_imagen_url or None
                prod.emoji = self.producto_form_emoji.strip() or None
                prod.updated_at = _utcnow()
                session.add(prod)
            else:
                prod = Producto(
                    company_id=self._company_id(),
                    categoria_id=cat.id or 0,
                    nombre=nombre,
                    descripcion=self.producto_form_descripcion.strip() or None,
                    precio=precio,
                    disponible=self.producto_form_disponible,
                    imagen_url=self.producto_form_imagen_url or None,
                    emoji=self.producto_form_emoji.strip() or None,
                )
                session.add(prod)
            session.commit()
        self.cargar_menu()
        self._reset_producto_form()
        self.mensaje = "Producto guardado."

    def editar_producto(self, producto_id: int) -> None:
        prod = next((p for p in self.productos if p.id == producto_id), None)
        if prod is None:
            return
        self.producto_form_id = prod.id
        self.producto_form_nombre = prod.nombre
        self.producto_form_descripcion = prod.descripcion
        self.producto_form_precio = str(prod.precio)
        self.producto_form_categoria_nombre = prod.categoria_nombre
        self.producto_form_disponible = prod.disponible
        self.producto_form_imagen_url = prod.imagen_url
        with self._tenant_session() as session:
            prod_db = session.get(Producto, producto_id)
            self.producto_form_emoji = (prod_db.emoji or "") if prod_db else ""

    def toggle_producto_disponible(self, producto_id: int) -> None:
        with self._tenant_session() as session:
            prod = session.get(Producto, producto_id)
            if prod is None or prod.company_id != self._company_id():
                return
            prod.disponible = not prod.disponible
            prod.updated_at = _utcnow()
            session.add(prod)
            session.commit()
        self.cargar_menu()

    def _reset_producto_form(self) -> None:
        self.producto_form_id = 0
        self.producto_form_nombre = ""
        self.producto_form_descripcion = ""
        self.producto_form_precio = ""
        self.producto_form_disponible = True
        self.producto_form_imagen_url = ""
        self.producto_form_emoji = ""
        if self.categorias:
            self.producto_form_categoria_nombre = self.categorias[0].nombre

    def cancelar_producto_form(self) -> None:
        self._reset_producto_form()

    async def handle_upload_imagen_producto(self, files: list[rx.UploadFile]) -> None:
        for file in files:
            data = await file.read()
            ext = pathlib.Path(file.name).suffix.lower() or ".jpg"
            filename = f"food_prod_{uuid.uuid4().hex[:12]}{ext}"
            upload_dir = pathlib.Path(rx.get_upload_dir()) / "food_productos"
            upload_dir.mkdir(parents=True, exist_ok=True)
            (upload_dir / filename).write_bytes(data)
            self.producto_form_imagen_url = f"{_FOOD_API_URL}/_upload/food_productos/{filename}"
            break  # solo primera imagen

    def quitar_imagen_producto(self) -> None:
        self.producto_form_imagen_url = ""

    async def handle_upload_logo_empresa(self, files: list[rx.UploadFile]) -> None:
        for file in files:
            data = await file.read()
            ext = pathlib.Path(file.name).suffix.lower() or ".jpg"
            filename = f"food_logo_{uuid.uuid4().hex[:12]}{ext}"
            upload_dir = pathlib.Path(rx.get_upload_dir()) / "food_empresas"
            upload_dir.mkdir(parents=True, exist_ok=True)
            (upload_dir / filename).write_bytes(data)
            self.config_logo_url = f"{_FOOD_API_URL}/_upload/food_empresas/{filename}"
            break

    def quitar_logo_empresa(self) -> None:
        self.config_logo_url = ""


    # ─── Inventario ───────────────────────────────────────────────────────────

    def on_load_inventario(self) -> None:
        self.cargar_inventario()
        if not self.productos:
            self.cargar_menu()

    def cargar_inventario(self) -> None:
        with self._tenant_session() as session:
            insumos_db = session.exec(
                select(Insumo)
                .where(Insumo.company_id == self._company_id())
                .order_by(Insumo.nombre)
            ).all()
            views: list[InsumoView] = []
            alertas: list[str] = []
            for ins in insumos_db:
                stock = Decimal(str(ins.stock_actual))
                minimo = Decimal(str(ins.stock_minimo))
                bajo = ins.activo and minimo > 0 and stock <= minimo
                views.append(InsumoView(
                    id=ins.id or 0,
                    nombre=ins.nombre,
                    unidad=ins.unidad,
                    stock_actual=float(stock),
                    stock_minimo=float(minimo),
                    activo=ins.activo,
                    bajo_stock=bajo,
                    stock_texto=f"{stock:.3f} {ins.unidad}".rstrip("0").rstrip(".").strip(),
                    stock_minimo_texto=f"{minimo:.3f} {ins.unidad}".rstrip("0").rstrip(".").strip(),
                ))
                if bajo:
                    alertas.append(ins.nombre)
            self.inv_insumos = views
            self.inv_alertas_bajo_stock = alertas

    def set_inv_form_nombre(self, v: str) -> None:
        self.inv_form_nombre = v

    def set_inv_form_unidad(self, v: str) -> None:
        self.inv_form_unidad = v

    def set_inv_form_stock_actual(self, v: str) -> None:
        self.inv_form_stock_actual = v

    def set_inv_form_stock_minimo(self, v: str) -> None:
        self.inv_form_stock_minimo = v

    def guardar_insumo(self) -> None:
        nombre = self.inv_form_nombre.strip()
        if not nombre:
            self.mensaje = "El nombre del insumo es obligatorio."
            return
        unidad = self.inv_form_unidad.strip() or "unidad"
        try:
            stock_actual = Decimal(self.inv_form_stock_actual.replace(",", ".").strip() or "0")
            stock_minimo = Decimal(self.inv_form_stock_minimo.replace(",", ".").strip() or "0")
            if stock_actual < 0 or stock_minimo < 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            self.mensaje = "Stock inválido. Ingresa números positivos."
            return
        with self._tenant_session() as session:
            if self.inv_form_id == 0:
                existente = session.exec(
                    select(Insumo).where(
                        Insumo.company_id == self._company_id(),
                        Insumo.nombre == nombre,
                    )
                ).first()
                if existente:
                    self.mensaje = f"Ya existe un insumo llamado '{nombre}'."
                    return
                ins = Insumo(
                    company_id=self._company_id(),
                    nombre=nombre,
                    unidad=unidad,
                    stock_actual=stock_actual,
                    stock_minimo=stock_minimo,
                    activo=True,
                )
                session.add(ins)
                self.mensaje = f"Insumo '{nombre}' creado."
            else:
                ins = session.get(Insumo, self.inv_form_id)
                if ins is None or ins.company_id != self._company_id():
                    self.mensaje = "Insumo no encontrado."
                    return
                ins.nombre = nombre
                ins.unidad = unidad
                ins.stock_actual = stock_actual
                ins.stock_minimo = stock_minimo
                ins.updated_at = _utcnow()
                session.add(ins)
                self.mensaje = f"Insumo '{nombre}' actualizado."
            session.commit()
        self.inv_form_id = 0
        self.inv_form_nombre = ""
        self.inv_form_unidad = "unidad"
        self.inv_form_stock_actual = ""
        self.inv_form_stock_minimo = ""
        self.inv_form_editando = False
        self.inv_form_visible = False
        self.cargar_inventario()

    def editar_insumo(self, insumo_id: int) -> None:
        with self._tenant_session() as session:
            ins = session.get(Insumo, insumo_id)
            if ins is None or ins.company_id != self._company_id():
                return
            self.inv_form_id = ins.id or 0
            self.inv_form_nombre = ins.nombre
            self.inv_form_unidad = ins.unidad
            stock = Decimal(str(ins.stock_actual))
            minimo = Decimal(str(ins.stock_minimo))
            self.inv_form_stock_actual = f"{stock:.3f}".rstrip("0").rstrip(".")
            self.inv_form_stock_minimo = f"{minimo:.3f}".rstrip("0").rstrip(".")
        self.inv_form_editando = True
        self.inv_form_visible = True

    def cancelar_insumo_form(self) -> None:
        self.inv_form_id = 0
        self.inv_form_nombre = ""
        self.inv_form_unidad = "unidad"
        self.inv_form_stock_actual = ""
        self.inv_form_stock_minimo = ""
        self.inv_form_editando = False
        self.inv_form_visible = False

    def toggle_insumo_activo(self, insumo_id: int) -> None:
        with self._tenant_session() as session:
            ins = session.get(Insumo, insumo_id)
            if ins is None or ins.company_id != self._company_id():
                return
            ins.activo = not ins.activo
            ins.updated_at = _utcnow()
            session.add(ins)
            session.commit()
        self.cargar_inventario()

    def set_inv_producto_sel_nombre(self, v: str) -> None:
        self.inv_producto_sel_nombre = v
        prod = next((p for p in self.productos if p.nombre == v), None)
        self.inv_producto_sel_id = prod.id if prod else 0
        self.cargar_receta_producto()

    def cargar_receta_producto(self) -> None:
        pid = self.inv_producto_sel_id
        if pid == 0:
            self.inv_receta_items = []
            return
        with self._tenant_session() as session:
            ris = session.exec(
                select(RecetaItem)
                .where(RecetaItem.company_id == self._company_id(), RecetaItem.producto_id == pid)
                .order_by(RecetaItem.id)
            ).all()
            insumos = {
                i.id: i
                for i in session.exec(
                    select(Insumo).where(Insumo.company_id == self._company_id())
                ).all()
            }
            items: list[RecetaItemView] = []
            for ri in ris:
                ins = insumos.get(ri.insumo_id)
                cant = Decimal(str(ri.cantidad))
                unidad = ins.unidad if ins else ""
                items.append(RecetaItemView(
                    id=ri.id or 0,
                    producto_id=pid,
                    insumo_id=ri.insumo_id,
                    insumo_nombre=ins.nombre if ins else "?",
                    insumo_unidad=unidad,
                    cantidad=float(cant),
                    cantidad_texto=f"{cant:.3f}".rstrip("0").rstrip(".") + f" {unidad}",
                ))
            self.inv_receta_items = items

    def set_inv_receta_insumo_sel_nombre(self, v: str) -> None:
        self.inv_receta_insumo_sel_nombre = v

    def set_inv_receta_cantidad(self, v: str) -> None:
        self.inv_receta_cantidad = v

    def guardar_receta_item(self) -> None:
        if self.inv_producto_sel_id == 0:
            self.mensaje = "Selecciona un producto primero."
            return
        insumo_match = next(
            (i for i in self.inv_insumos if i.nombre == self.inv_receta_insumo_sel_nombre), None
        )
        if insumo_match is None:
            self.mensaje = "Selecciona un insumo válido."
            return
        try:
            cantidad = Decimal(self.inv_receta_cantidad.replace(",", ".").strip() or "0")
            if cantidad <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            self.mensaje = "Cantidad inválida. Ingresa un número mayor a 0."
            return
        insumo_id = insumo_match.id
        with self._tenant_session() as session:
            existente = session.exec(
                select(RecetaItem).where(
                    RecetaItem.company_id == self._company_id(),
                    RecetaItem.producto_id == self.inv_producto_sel_id,
                    RecetaItem.insumo_id == insumo_id,
                )
            ).first()
            if existente:
                existente.cantidad = cantidad
                existente.updated_at = _utcnow()
                session.add(existente)
                self.mensaje = "Cantidad actualizada en receta."
            else:
                ri = RecetaItem(
                    company_id=self._company_id(),
                    producto_id=self.inv_producto_sel_id,
                    insumo_id=insumo_id,
                    cantidad=cantidad,
                )
                session.add(ri)
                self.mensaje = "Insumo agregado a la receta."
            session.commit()
        self.inv_receta_cantidad = ""
        self.inv_receta_insumo_sel_nombre = ""
        self.cargar_receta_producto()

    def eliminar_receta_item(self, item_id: int) -> None:
        with self._tenant_session() as session:
            ri = session.get(RecetaItem, item_id)
            if ri is None or ri.company_id != self._company_id():
                return
            session.delete(ri)
            session.commit()
        self.cargar_receta_producto()
        self.mensaje = "Insumo eliminado de la receta."

    # ─── Clientes ─────────────────────────────────────────────────────────────

    def on_load_clientes(self) -> None:
        self.mensaje = ""
        self.cargar_clientes()

    def cargar_clientes(self) -> None:
        hoy = _utcnow()
        with self._tenant_session() as session:
            clientes_db = session.exec(
                select(Cliente)
                .where(Cliente.company_id == self._company_id())
                .order_by(Cliente.nombre)
            ).all()
            pedidos_pagados = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    Pedido.cliente_id.is_not(None),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                )
            ).all()
        stats_por_cliente: dict[int, dict] = {}
        for p in pedidos_pagados:
            cid = p.cliente_id
            if cid is None:
                continue
            stats = stats_por_cliente.setdefault(cid, {"visitas": 0, "gastado": Decimal("0.00"), "ultima": None})
            stats["visitas"] += 1
            stats["gastado"] += _to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0))
            fecha = p.cerrado_en
            if fecha and (stats["ultima"] is None or fecha > stats["ultima"]):
                stats["ultima"] = fecha
        hoy_fecha = hoy.date()
        views: list[ClienteView] = []
        for c in clientes_db:
            fn = c.fecha_nacimiento
            cumple_hoy = False
            cumple_pronto = False
            dias = 999
            nac_texto = ""
            nac_iso = ""
            if fn:
                nac_iso = fn.isoformat()
                meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
                nac_texto = f"{fn.day} {meses[fn.month - 1]}"
                cumple_este_anio = datetime(hoy.year, fn.month, fn.day)
                if cumple_este_anio < hoy.replace(hour=0, minute=0, second=0, microsecond=0):
                    cumple_este_anio = datetime(hoy.year + 1, fn.month, fn.day)
                dias = (cumple_este_anio - hoy.replace(hour=0, minute=0, second=0, microsecond=0)).days
                cumple_hoy = dias == 0
                cumple_pronto = 0 < dias <= 7
            stats = stats_por_cliente.get(c.id or 0, {"visitas": 0, "gastado": Decimal("0.00"), "ultima": None})
            ultima = stats["ultima"]
            if ultima is None:
                ultima_texto = "Sin visitas"
            else:
                dias_desde = (hoy_fecha - ultima.date()).days
                if dias_desde <= 0:
                    ultima_texto = "Hoy"
                elif dias_desde == 1:
                    ultima_texto = "Ayer"
                else:
                    ultima_texto = f"Hace {dias_desde} días"
            views.append(ClienteView(
                id=c.id or 0,
                nombre=c.nombre,
                telefono=c.telefono or "",
                email=c.email or "",
                fecha_nac_iso=nac_iso,
                fecha_nac_texto=nac_texto,
                notas=c.notas or "",
                puntos=c.puntos,
                activo=c.activo,
                cumple_hoy=cumple_hoy,
                cumple_pronto=cumple_pronto,
                dias_para_cumple=dias,
                visitas_count=stats["visitas"],
                gastado_texto=_money_text(stats["gastado"]),
                ultima_visita_texto=ultima_texto,
                es_vip=stats["visitas"] >= CLIENTE_VIP_VISITAS_MIN,
            ))
        self.clientes_lista = views

    def set_cli_busqueda(self, v: str) -> None:
        self.cli_busqueda = v

    def set_cli_form_nombre(self, v: str) -> None:
        self.cli_form_nombre = v

    def set_cli_form_telefono(self, v: str) -> None:
        self.cli_form_telefono = v

    def set_cli_form_email(self, v: str) -> None:
        self.cli_form_email = v

    def set_cli_form_fecha_nac(self, v: str) -> None:
        self.cli_form_fecha_nac = v

    def set_cli_form_notas(self, v: str) -> None:
        self.cli_form_notas = v

    def set_caja_cobro_cliente_nombre(self, v: str) -> None:
        self.caja_cobro_cliente_nombre = v
        cli = next((c for c in self.clientes_lista if c.nombre == v), None)
        self.caja_cobro_cliente_id = cli.id if cli else 0

    def guardar_cliente(self) -> None:
        nombre = self.cli_form_nombre.strip()
        if not nombre:
            self.mensaje = "El nombre del cliente es obligatorio."
            return
        tel = self.cli_form_telefono.strip() or None
        email = self.cli_form_email.strip() or None
        notas = self.cli_form_notas.strip() or None
        fn: date | None = None
        if self.cli_form_fecha_nac:
            try:
                from datetime import date as _date
                parts = self.cli_form_fecha_nac.split("-")
                fn = _date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, IndexError):
                self.mensaje = "Fecha de nacimiento inválida. Usa el formato AAAA-MM-DD."
                return
        with self._tenant_session() as session:
            if self.cli_form_id == 0:
                existente = session.exec(
                    select(Cliente).where(
                        Cliente.company_id == self._company_id(),
                        Cliente.nombre == nombre,
                    )
                ).first()
                if existente:
                    self.mensaje = f"Ya existe un cliente con ese nombre."
                    return
                c = Cliente(
                    company_id=self._company_id(),
                    nombre=nombre,
                    telefono=tel,
                    email=email,
                    fecha_nacimiento=fn,
                    notas=notas,
                    activo=True,
                )
                session.add(c)
                self.mensaje = f"Cliente '{nombre}' registrado."
            else:
                c = session.get(Cliente, self.cli_form_id)
                if c is None or c.company_id != self._company_id():
                    self.mensaje = "Cliente no encontrado."
                    return
                c.nombre = nombre
                c.telefono = tel
                c.email = email
                c.fecha_nacimiento = fn
                c.notas = notas
                c.updated_at = _utcnow()
                session.add(c)
                self.mensaje = f"Cliente '{nombre}' actualizado."
            session.commit()
        self.cli_form_id = 0
        self.cli_form_nombre = ""
        self.cli_form_telefono = ""
        self.cli_form_email = ""
        self.cli_form_fecha_nac = ""
        self.cli_form_notas = ""
        self.cli_form_editando = False
        self.cli_form_visible = False
        self.cargar_clientes()

    def abrir_nuevo_cliente(self) -> None:
        self.cli_form_id = 0
        self.cli_form_nombre = ""
        self.cli_form_telefono = ""
        self.cli_form_email = ""
        self.cli_form_fecha_nac = ""
        self.cli_form_notas = ""
        self.cli_form_editando = False
        self.cli_form_visible = True

    def editar_cliente(self, cliente_id: int) -> None:
        with self._tenant_session() as session:
            c = session.get(Cliente, cliente_id)
            if c is None or c.company_id != self._company_id():
                return
            self.cli_form_id = c.id or 0
            self.cli_form_nombre = c.nombre
            self.cli_form_telefono = c.telefono or ""
            self.cli_form_email = c.email or ""
            self.cli_form_fecha_nac = c.fecha_nacimiento.isoformat() if c.fecha_nacimiento else ""
            self.cli_form_notas = c.notas or ""
        self.cli_form_editando = True
        self.cli_form_visible = True

    def cancelar_cli_form(self) -> None:
        self.cli_form_id = 0
        self.cli_form_nombre = ""
        self.cli_form_telefono = ""
        self.cli_form_email = ""
        self.cli_form_fecha_nac = ""
        self.cli_form_notas = ""
        self.cli_form_editando = False
        self.cli_form_visible = False

    def set_cli_form_visible(self, v: bool) -> None:
        self.cli_form_visible = v

    def toggle_cliente_activo(self, cliente_id: int) -> None:
        with self._tenant_session() as session:
            c = session.get(Cliente, cliente_id)
            if c is None or c.company_id != self._company_id():
                return
            c.activo = not c.activo
            c.updated_at = _utcnow()
            session.add(c)
            session.commit()
        self.cargar_clientes()

    # ─── Cuentas corrientes ───────────────────────────────────────────────────

    def on_load_cuentas(self) -> None:
        self.mensaje = ""
        if not self.clientes_lista:
            self.cargar_clientes()
        self.cargar_cuentas()

    def cargar_cuentas(self) -> None:
        with self._tenant_session() as session:
            cuentas_db = session.exec(
                select(CuentaCorriente)
                .where(CuentaCorriente.company_id == self._company_id())
                .order_by(CuentaCorriente.saldo_deuda.desc())
            ).all()
            clientes_map = {
                c.id: c
                for c in session.exec(
                    select(Cliente).where(Cliente.company_id == self._company_id())
                ).all()
            }
        views: list[CuentaView] = []
        for cc in cuentas_db:
            cli = clientes_map.get(cc.cliente_id)
            saldo = Decimal(str(cc.saldo_deuda))
            views.append(CuentaView(
                id=cc.id or 0,
                cliente_id=cc.cliente_id,
                cliente_nombre=cli.nombre if cli else "?",
                cliente_telefono=cli.telefono or "" if cli else "",
                saldo_deuda=float(saldo),
                saldo_texto=_money_text(saldo),
                limite_credito=float(Decimal(str(cc.limite_credito))),
            ))
        self.cuentas_lista = views

    def set_cc_cliente_sel_nombre(self, v: str) -> None:
        self.cc_cliente_sel_nombre = v
        cli = next((c for c in self.clientes_lista if c.nombre == v), None)
        if cli:
            self._ver_o_crear_cuenta(cli.id)

    def _ver_o_crear_cuenta(self, cliente_id: int) -> None:
        with self._tenant_session() as session:
            cc = session.exec(
                select(CuentaCorriente).where(
                    CuentaCorriente.company_id == self._company_id(),
                    CuentaCorriente.cliente_id == cliente_id,
                )
            ).first()
            if cc:
                self.cuenta_sel_id = cc.id or 0
                movs = session.exec(
                    select(MovimientoCuenta)
                    .where(MovimientoCuenta.cuenta_id == cc.id)
                    .order_by(MovimientoCuenta.created_at.desc())
                ).all()
                self.cuenta_movimientos = [
                    MovimientoView(
                        id=m.id or 0,
                        tipo=m.tipo,
                        tipo_label="Cargo" if m.tipo == "cargo" else "Pago",
                        monto=float(Decimal(str(m.monto))),
                        monto_texto=_money_text(m.monto),
                        descripcion=m.descripcion or "",
                        fecha_texto=m.created_at.strftime("%d/%m %H:%M"),
                    )
                    for m in movs
                ]
            else:
                self.cuenta_sel_id = 0
                self.cuenta_movimientos = []

    def set_cc_pago_monto(self, v: str) -> None:
        self.cc_pago_monto = v

    def set_cc_pago_descripcion(self, v: str) -> None:
        self.cc_pago_descripcion = v

    def registrar_pago_cc(self) -> None:
        if self.cuenta_sel_id == 0:
            self.mensaje = "Selecciona un cliente con cuenta corriente."
            return
        try:
            monto = Decimal(self.cc_pago_monto.replace(",", ".").strip() or "0")
            if monto <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            self.mensaje = "Monto de pago inválido."
            return
        with self._tenant_session() as session:
            cc = session.get(CuentaCorriente, self.cuenta_sel_id)
            if cc is None or cc.company_id != self._company_id():
                self.mensaje = "Cuenta no encontrada."
                return
            pago = MovimientoCuenta(
                company_id=self._company_id(),
                cuenta_id=cc.id or 0,
                tipo="pago",
                monto=monto,
                descripcion=self.cc_pago_descripcion.strip() or "Pago en caja",
            )
            session.add(pago)
            saldo_actual = Decimal(str(cc.saldo_deuda))
            cc.saldo_deuda = max(saldo_actual - monto, Decimal("0.00"))
            cc.updated_at = _utcnow()
            session.add(cc)
            session.commit()
            cliente_id = cc.cliente_id
        self.cc_pago_monto = ""
        self.cc_pago_descripcion = ""
        self.mensaje = f"Pago de {_money_text(monto)} registrado."
        self.cargar_cuentas()
        self._ver_o_crear_cuenta(cliente_id)

    def _registrar_cargo_cc(self, session, cliente_id: int, monto: Decimal, pedido_id: int | None, descripcion: str) -> None:
        cc = session.exec(
            select(CuentaCorriente).where(
                CuentaCorriente.company_id == self._company_id(),
                CuentaCorriente.cliente_id == cliente_id,
            )
        ).first()
        if cc is None:
            cc = CuentaCorriente(
                company_id=self._company_id(),
                cliente_id=cliente_id,
                saldo_deuda=Decimal("0.00"),
                limite_credito=Decimal("0.00"),
            )
            session.add(cc)
            session.flush()
        saldo_actual = Decimal(str(cc.saldo_deuda))
        limite = Decimal(str(cc.limite_credito))
        if limite > Decimal("0.00") and saldo_actual + monto > limite:
            raise ValueError(
                f"Límite de crédito excedido. Deuda actual: {_money_text(saldo_actual)}, "
                f"límite: {_money_text(limite)}, cargo solicitado: {_money_text(monto)}."
            )
        cargo = MovimientoCuenta(
            company_id=self._company_id(),
            cuenta_id=cc.id or 0,
            pedido_id=pedido_id,
            tipo="cargo",
            monto=monto,
            descripcion=descripcion,
        )
        session.add(cargo)
        cc.saldo_deuda = saldo_actual + monto
        cc.updated_at = _utcnow()
        session.add(cc)

    def exportar_cuentas_excel(self):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        with self._tenant_session() as session:
            cuentas_db = session.exec(
                select(CuentaCorriente)
                .where(CuentaCorriente.company_id == self._company_id())
                .order_by(CuentaCorriente.saldo_deuda.desc())
            ).all()
            clientes_map = {
                c.id: c
                for c in session.exec(
                    select(Cliente).where(Cliente.company_id == self._company_id())
                ).all()
            }
            cuenta_ids = [c.id for c in cuentas_db if c.id is not None]
            movimientos: list = []
            if cuenta_ids:
                movimientos = session.exec(
                    select(MovimientoCuenta)
                    .where(MovimientoCuenta.cuenta_id.in_(cuenta_ids))
                    .order_by(MovimientoCuenta.created_at.desc())
                ).all()

        if not cuentas_db and not movimientos:
            self.mensaje = "No hay cuentas corrientes para exportar."
            return None

        cuentas_por_id = {c.id: c for c in cuentas_db}

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Cuentas"
        ws1.append(["Cliente", "Teléfono", "Saldo deuda", "Límite de crédito"])
        for cc in cuentas_db:
            cli = clientes_map.get(cc.cliente_id)
            ws1.append([
                cli.nombre if cli else "?",
                (cli.telefono or "") if cli else "",
                float(Decimal(str(cc.saldo_deuda))),
                float(Decimal(str(cc.limite_credito))),
            ])

        ws2 = wb.create_sheet("Movimientos")
        ws2.append(["Fecha", "Hora", "Cliente", "Tipo", "Monto", "Descripción"])
        for m in movimientos:
            cc = cuentas_por_id.get(m.cuenta_id)
            cli = clientes_map.get(cc.cliente_id) if cc else None
            ws2.append([
                m.created_at.strftime("%Y-%m-%d") if m.created_at else "",
                m.created_at.strftime("%H:%M") if m.created_at else "",
                cli.nombre if cli else "?",
                "Cargo" if m.tipo == "cargo" else "Pago",
                float(Decimal(str(m.monto))),
                m.descripcion or "",
            ])

        for ws in (ws1, ws2):
            for i, col in enumerate(ws.columns, start=1):
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"cuentas_corrientes_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    # ─── Promociones ──────────────────────────────────────────────────────────

    def on_load_promociones(self) -> None:
        self.mensaje = ""
        self.cargar_promociones()

    def cargar_promociones(self) -> None:
        now = _utcnow()
        hora_actual = now.strftime("%H:%M")
        tipo_labels = {
            TipoPromocion.PORCENTAJE.value: "% Descuento",
            TipoPromocion.MONTO_FIJO.value: "Monto fijo",
            TipoPromocion.HAPPY_HOUR.value: "Happy Hour",
        }
        with self._tenant_session() as session:
            promos_db = session.exec(
                select(Promocion)
                .where(Promocion.company_id == self._company_id())
                .order_by(Promocion.activa.desc(), Promocion.nombre)
            ).all()
        views: list[PromocionView] = []
        for p in promos_db:
            aplica = p.activa
            if aplica and p.hora_inicio and p.hora_fin:
                if p.hora_inicio <= p.hora_fin:
                    aplica = p.hora_inicio <= hora_actual <= p.hora_fin
                else:
                    # El horario cruza la medianoche (ej. 22:00 a 02:00):
                    # esta comparado lexicografica simple nunca seria True
                    # para ninguna hora si no se maneja este caso aparte.
                    aplica = hora_actual >= p.hora_inicio or hora_actual <= p.hora_fin
            val = Decimal(str(p.valor))
            if p.tipo in (TipoPromocion.PORCENTAJE.value, TipoPromocion.HAPPY_HOUR.value):
                desc_txt = f"{val:.0f}% off"
            else:
                desc_txt = f"- {_money_text(val)}"
            if p.hora_inicio and p.hora_fin:
                horario = f"{p.hora_inicio} – {p.hora_fin}"
            else:
                horario = "Todo el día"
            views.append(PromocionView(
                id=p.id or 0,
                nombre=p.nombre,
                tipo=p.tipo,
                tipo_label=tipo_labels.get(p.tipo, p.tipo),
                valor=float(val),
                descripcion=p.descripcion or "",
                hora_inicio=p.hora_inicio or "",
                hora_fin=p.hora_fin or "",
                activa=p.activa,
                aplica_ahora=aplica,
                descuento_texto=desc_txt,
                horario_texto=horario,
            ))
        self.promociones_lista = views

    def set_promo_form_nombre(self, v: str) -> None:
        self.promo_form_nombre = v

    def set_promo_form_tipo(self, v: str) -> None:
        self.promo_form_tipo = v

    def set_promo_form_valor(self, v: str) -> None:
        self.promo_form_valor = v

    def set_promo_form_descripcion(self, v: str) -> None:
        self.promo_form_descripcion = v

    def set_promo_form_hora_inicio(self, v: str) -> None:
        self.promo_form_hora_inicio = v

    def set_promo_form_hora_fin(self, v: str) -> None:
        self.promo_form_hora_fin = v

    def guardar_promocion(self) -> None:
        nombre = self.promo_form_nombre.strip()
        if not nombre:
            self.mensaje = "El nombre de la promoción es obligatorio."
            return
        try:
            valor = Decimal(self.promo_form_valor.replace(",", ".").strip() or "0")
            if valor <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            self.mensaje = "Valor inválido. Ingresa un número mayor a 0."
            return
        hora_ini = self.promo_form_hora_inicio.strip() or None
        hora_fin = self.promo_form_hora_fin.strip() or None
        with self._tenant_session() as session:
            if self.promo_form_id == 0:
                p = Promocion(
                    company_id=self._company_id(),
                    nombre=nombre,
                    tipo=self.promo_form_tipo,
                    valor=valor,
                    descripcion=self.promo_form_descripcion.strip() or None,
                    hora_inicio=hora_ini,
                    hora_fin=hora_fin,
                    activa=True,
                )
                session.add(p)
                self.mensaje = f"Promoción '{nombre}' creada."
            else:
                p = session.get(Promocion, self.promo_form_id)
                if p is None or p.company_id != self._company_id():
                    self.mensaje = "Promoción no encontrada."
                    return
                p.nombre = nombre
                p.tipo = self.promo_form_tipo
                p.valor = valor
                p.descripcion = self.promo_form_descripcion.strip() or None
                p.hora_inicio = hora_ini
                p.hora_fin = hora_fin
                p.updated_at = _utcnow()
                session.add(p)
                self.mensaje = f"Promoción '{nombre}' actualizada."
            session.commit()
        self.promo_form_id = 0
        self.promo_form_nombre = ""
        self.promo_form_tipo = TipoPromocion.PORCENTAJE.value
        self.promo_form_valor = ""
        self.promo_form_descripcion = ""
        self.promo_form_hora_inicio = ""
        self.promo_form_hora_fin = ""
        self.promo_form_editando = False
        self.promo_form_visible = False
        self.cargar_promociones()

    def abrir_nueva_promo(self) -> None:
        self.promo_form_id = 0
        self.promo_form_nombre = ""
        self.promo_form_tipo = TipoPromocion.PORCENTAJE.value
        self.promo_form_valor = ""
        self.promo_form_descripcion = ""
        self.promo_form_hora_inicio = ""
        self.promo_form_hora_fin = ""
        self.promo_form_editando = False
        self.promo_form_visible = True

    def set_promo_form_visible(self, v: bool) -> None:
        self.promo_form_visible = v

    def editar_promocion(self, promo_id: int) -> None:
        with self._tenant_session() as session:
            p = session.get(Promocion, promo_id)
            if p is None or p.company_id != self._company_id():
                return
            self.promo_form_id = p.id or 0
            self.promo_form_nombre = p.nombre
            self.promo_form_tipo = p.tipo
            self.promo_form_valor = str(Decimal(str(p.valor)).normalize())
            self.promo_form_descripcion = p.descripcion or ""
            self.promo_form_hora_inicio = p.hora_inicio or ""
            self.promo_form_hora_fin = p.hora_fin or ""
        self.promo_form_editando = True
        self.promo_form_visible = True

    def cancelar_promo_form(self) -> None:
        self.promo_form_id = 0
        self.promo_form_nombre = ""
        self.promo_form_tipo = TipoPromocion.PORCENTAJE.value
        self.promo_form_valor = ""
        self.promo_form_descripcion = ""
        self.promo_form_hora_inicio = ""
        self.promo_form_hora_fin = ""
        self.promo_form_editando = False
        self.promo_form_visible = False

    def toggle_promo_activa(self, promo_id: int) -> None:
        with self._tenant_session() as session:
            p = session.get(Promocion, promo_id)
            if p is None or p.company_id != self._company_id():
                return
            p.activa = not p.activa
            p.updated_at = _utcnow()
            session.add(p)
            session.commit()
        self.cargar_promociones()

    def aplicar_promo_al_cobro(self) -> None:
        desc = self.promo_activa_descuento_sugerido
        if desc > 0:
            self.caja_cobro_descuento = str(round(desc, 2))

    def refrescar_promos(self) -> None:
        self.cargar_promociones()


# ─── Estado público (sin auth) ────────────────────────────────────────────────

class ProductoPublicoView(BaseModel):
    nombre: str
    descripcion: str
    precio_texto: str
    imagen_url: str
    emoji: str = "🍽️"


class CategoriaPublicaView(BaseModel):
    nombre: str
    productos: list[ProductoPublicoView]
    emoji: str = "🍽️"


class MenuPublicoState(rx.State):
    """Estado de la carta pública — no requiere sesión."""

    nombre_local: str = ""
    categorias_menu: list[CategoriaPublicaView] = []
    cargando: bool = True
    no_encontrado: bool = False

    def on_load(self) -> None:
        slug = self.router.page.params.get("slug", "")
        self.cargando = True
        self.no_encontrado = False
        self.nombre_local = ""
        self.categorias_menu = []

        if not slug:
            self.no_encontrado = True
            self.cargando = False
            return

        # Único punto legítimamente cross-tenant de todo el sistema: la carta
        # pública se resuelve por slug, no por sesión — ningún contexto tenant
        # está armado todavía cuando llega esta request.
        with tenant_bypass(), get_session() as session:
            cfg = session.exec(
                select(ConfigImpresora).where(ConfigImpresora.slug == slug)
            ).first()
            if cfg is None:
                self.no_encontrado = True
                self.cargando = False
                return

            company_id = cfg.company_id
            self.nombre_local = cfg.nombre_local

            cats = session.exec(
                select(Categoria)
                .where(Categoria.company_id == company_id, Categoria.activa.is_(True))
                .order_by(Categoria.orden, Categoria.nombre)
            ).all()

            result: list[CategoriaPublicaView] = []
            for cat in cats:
                prods = session.exec(
                    select(Producto)
                    .where(
                        Producto.company_id == company_id,
                        Producto.categoria_id == cat.id,
                        Producto.disponible.is_(True),
                    )
                    .order_by(Producto.nombre)
                ).all()
                if prods:
                    result.append(
                        CategoriaPublicaView(
                            nombre=cat.nombre,
                            emoji=_emoji_para_categoria(cat.nombre),
                            productos=[
                                ProductoPublicoView(
                                    nombre=p.nombre,
                                    descripcion=p.descripcion or "",
                                    precio_texto=_money_text(p.precio),
                                    imagen_url=p.imagen_url or "",
                                    emoji=p.emoji or _emoji_para_producto(p.nombre),
                                )
                                for p in prods
                            ],
                        )
                    )

            self.categorias_menu = result

        self.cargando = False


class AdminLocalState(rx.State):
    """Estado para login de dueño del local vía email+contraseña (independiente del PIN)."""

    autenticado: bool = False
    email_input: str = ""
    password_input: str = ""
    error_msg: str = ""
    show_password: bool = False
    login_empresa_nombre: str = ""
    login_empresa_logo: str = ""
    login_empresa_slug: str = ""

    def set_email_input(self, v: str) -> None:
        self.email_input = v

    def set_password_input(self, v: str) -> None:
        self.password_input = v

    def toggle_show_password(self) -> None:
        self.show_password = not self.show_password

    def on_load_dono_login(self):
        self.error_msg = ""
        self.login_empresa_nombre = ""
        self.login_empresa_logo = ""
        self.login_empresa_slug = ""
        slug = self.router.page.params.get("empresa", "")
        if slug:
            with tenant_bypass():
                with get_session() as session:
                    empresa = session.exec(
                        select(Company).where(Company.slug == slug, Company.is_active.is_(True))
                    ).first()
            if empresa is not None:
                self.login_empresa_nombre = empresa.name
                self.login_empresa_logo = empresa.logo_url or ""
                self.login_empresa_slug = empresa.slug
        if self.autenticado:
            return rx.redirect("/admin")
        return None

    async def on_load_dono(self):
        if self.autenticado:
            return None
        # Una sesion PIN de rol Admin (login rapido en /login) tambien
        # habilita el Dashboard, no solo el login por email/contraseña.
        food_state = await self.get_state(FoodState)
        if (
            food_state.usuario_actual is not None
            and food_state.usuario_actual.rol == RolUsuario.ADMIN.value
        ):
            self.autenticado = True
            return None
        return rx.redirect("/admin/login")

    def login_on_enter(self, key: str) -> None:
        if key == "Enter":
            return self.login_admin_local()

    async def login_admin_local(self) -> None:
        import hashlib
        email = self.email_input.strip().lower()
        password = self.password_input.strip()
        self.error_msg = ""
        if not email or not password:
            self.error_msg = "Ingresa email y contraseña."
            return
        # El email de admin es único globalmente (migración 0014), así que esta
        # búsqueda es la única legítimamente cross-tenant: todavía no sabemos a
        # qué empresa pertenece este login hasta encontrar el email.
        with tenant_bypass():
            with get_session() as session:
                cfg = session.exec(
                    select(ConfigImpresora).where(
                        ConfigImpresora.admin_email == email,
                    )
                ).first()
        if cfg is None or not cfg.admin_password_hash:
            self.error_msg = "Credenciales incorrectas."
            return
        hashed = hashlib.sha256(password.encode()).hexdigest()
        if hashed != cfg.admin_password_hash:
            self.error_msg = "Credenciales incorrectas."
            return
        self.autenticado = True
        self.password_input = ""
        company_id = cfg.company_id
        # Vincular esta sesion con FoodState.usuario_actual: Carta, Reportes,
        # Usuarios y Configuracion validan acceso via usuario_actual.rol, no
        # via AdminLocalState.autenticado, asi que sin esto el dueño podia
        # ver el Dashboard pero quedaba bloqueado en todos sus sub-modulos.
        food_state = await self.get_state(FoodState)
        set_tenant_context(company_id, None)
        with get_session() as session:
            session.info["tenant_bypass"] = True
            admin_usuario = session.exec(
                select(UsuarioFood).where(
                    UsuarioFood.company_id == company_id,
                    UsuarioFood.rol == RolUsuario.ADMIN.value,
                    UsuarioFood.activo.is_(True),
                )
            ).first()
        if admin_usuario is not None:
            food_state.usuario_actual = UsuarioSesion(
                id=admin_usuario.id or 0,
                nombre=admin_usuario.nombre,
                rol=admin_usuario.rol,
                company_id=company_id,
            )
        else:
            food_state.usuario_actual = UsuarioSesion(
                id=0, nombre=email, rol=RolUsuario.ADMIN.value, company_id=company_id,
            )
        return rx.redirect("/admin")

    async def logout_admin_local(self) -> None:
        self.autenticado = False
        self.email_input = ""
        self.password_input = ""
        food_state = await self.get_state(FoodState)
        food_state.usuario_actual = None
        return rx.redirect("/admin/login")
