"""Clientes y cuentas corrientes — registro de clientes, cumpleaños y fiado.

Gestiona el CRUD de clientes (con consulta DNI/RUC vía API externa),
lógica de cumpleaños y la operativa de cuentas corrientes (cargos, pagos,
exportación Excel).

La lógica vive en métodos del mixin; FoodState provee en runtime
``_tenant_session()``, ``_company_id()``, ``usuario_actual``, ``mensaje``
y las listas de ViewModels compartidas.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

import reflex as rx
from sqlalchemy import or_
from sqlmodel import select

from app.models.food import (
    Cliente,
    CuentaCorriente,
    EstadoPedido,
    MovimientoCuenta,
    Pedido,
)
from app.states.food_state import (
    CLIENTE_VIP_VISITAS_MIN,
    ClienteView,
    CuentaView,
    MovimientoView,
    _money_text,
    _to_decimal,
    _utcnow,
)


# ─── Mixin de estado ─────────────────────────────────────────────────────────

class ClientesCuentasMixin(rx.State, mixin=True):
    """Estado de clientes y cuentas corrientes (fiado)."""

    # ── Clientes ────────────────────────────────────────────────────────────
    clientes_lista: list[ClienteView] = []  # noqa: F821  (forward ref — resolved at runtime via FoodState)
    cli_busqueda: str = ""
    cli_form_id: int = 0
    cli_form_nombre: str = ""
    cli_form_telefono: str = ""
    cli_form_email: str = ""
    cli_form_fecha_nac: str = ""
    cli_form_notas: str = ""
    cli_form_editando: bool = False
    cli_form_visible: bool = False
    cli_dni_ruc: str = ""
    cli_dni_ruc_buscando: bool = False
    cli_dni_ruc_error: str = ""

    # ── Cuentas corrientes ──────────────────────────────────────────────────
    cuentas_lista: list[CuentaView] = []  # noqa: F821
    cuenta_sel_id: int = 0
    cuenta_movimientos: list[MovimientoView] = []  # noqa: F821
    cc_pago_monto: str = ""
    cc_pago_descripcion: str = ""
    cc_cliente_sel_nombre: str = ""

    # ── Computed vars — Clientes ────────────────────────────────────────────

    @rx.var
    def clientes_filtrados(self) -> list[ClienteView]:  # noqa: F821
        q = self.cli_busqueda.lower().strip()
        if not q:
            return self.clientes_lista
        return [c for c in self.clientes_lista if q in c.nombre.lower() or q in c.telefono]

    @rx.var
    def clientes_activos_nombres(self) -> list[str]:
        """Opciones para el selector de fiado: "Nombre — tel" para evitar homónimos."""
        opciones = []
        for c in self.clientes_lista:
            if not c.activo:
                continue
            label = c.nombre + (f" — {c.telefono}" if c.telefono else "")
            opciones.append(label)
        return opciones

    @rx.var
    def clientes_cumpleanos_hoy(self) -> list[ClienteView]:  # noqa: F821
        return [c for c in self.clientes_lista if c.cumple_hoy]

    @rx.var
    def clientes_cumpleanos_pronto(self) -> list[ClienteView]:  # noqa: F821
        return [c for c in self.clientes_lista if c.cumple_pronto and not c.cumple_hoy]

    # ── Computed vars — Cuentas corrientes ──────────────────────────────────

    @rx.var
    def cuentas_con_deuda(self) -> list[CuentaView]:  # noqa: F821
        return [c for c in self.cuentas_lista if c.saldo_deuda > 0]

    @rx.var
    def cuentas_total_deuda_texto(self) -> str:
        total = sum(Decimal(str(c.saldo_deuda)) for c in self.cuentas_lista)
        return _money_text(total)

    @rx.var
    def cuenta_sel_nombre(self) -> str:
        c = next((x for x in self.cuentas_lista if x.id == self.cuenta_sel_id), None)
        return c.cliente_nombre if c else ""

    @rx.var
    def cuenta_sel_saldo(self) -> str:
        c = next((x for x in self.cuentas_lista if x.id == self.cuenta_sel_id), None)
        return c.saldo_texto if c else "S/ 0.00"

    def set_caja_cobro_cliente_nombre(self, v: str) -> None:
        self.caja_cobro_cliente_nombre = v
        nombre_parte = v.split(" — ")[0].strip()
        tel_parte = v.split(" — ")[1].strip() if " — " in v else ""
        cli = next(
            (c for c in self.clientes_lista
             if c.nombre == nombre_parte and (not tel_parte or c.telefono == tel_parte)),
            None,
        )
        self.caja_cobro_cliente_id = cli.id if cli else 0

    # ── Clientes — carga y CRUD ─────────────────────────────────────────────

    def on_load_clientes(self):
        result = self._route_access_result("clientes")
        if result is not None:
            return result
        self.cargar_clientes()

    def cargar_clientes(self) -> None:
        hoy = _utcnow()
        with self._tenant_session() as session:
            clientes_db = session.exec(
                select(Cliente)
                .where(Cliente.company_id == self._company_id())
                .order_by(Cliente.nombre)
            ).all()
            pedidos_pagados = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    Pedido.cliente_id.is_not(None),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                )
            ).all()
        stats_por_cliente: dict[int, dict] = {}
        for p in pedidos_pagados:
            cid = p.cliente_id
            if cid is None:
                continue
            stats = stats_por_cliente.setdefault(cid, {"visitas": 0, "gastado": Decimal("0.00"), "ultima": None})
            stats["visitas"] += 1
            stats["gastado"] += _to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0))
            fecha = p.cerrado_en
            if fecha and (stats["ultima"] is None or fecha > stats["ultima"]):
                stats["ultima"] = fecha
        hoy_fecha = hoy.date()
        views: list[ClienteView] = []
        for c in clientes_db:
            fn = c.fecha_nacimiento
            cumple_hoy = False
            cumple_pronto = False
            dias = 999
            nac_texto = ""
            nac_iso = ""
            if fn:
                nac_iso = fn.isoformat()
                meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
                nac_texto = f"{fn.day} {meses[fn.month - 1]}"
                cumple_este_anio = datetime(hoy.year, fn.month, fn.day)
                if cumple_este_anio < hoy.replace(hour=0, minute=0, second=0, microsecond=0):
                    cumple_este_anio = datetime(hoy.year + 1, fn.month, fn.day)
                dias = (cumple_este_anio - hoy.replace(hour=0, minute=0, second=0, microsecond=0)).days
                cumple_hoy = dias == 0
                cumple_pronto = 0 < dias <= 7
            stats = stats_por_cliente.get(c.id or 0, {"visitas": 0, "gastado": Decimal("0.00"), "ultima": None})
            ultima = stats["ultima"]
            if ultima is None:
                ultima_texto = "Sin visitas"
            else:
                dias_desde = (hoy_fecha - ultima.date()).days
                if dias_desde <= 0:
                    ultima_texto = "Hoy"
                elif dias_desde == 1:
                    ultima_texto = "Ayer"
                else:
                    ultima_texto = f"Hace {dias_desde} días"
            views.append(ClienteView(
                id=c.id or 0,
                nombre=c.nombre,
                telefono=c.telefono or "",
                email=c.email or "",
                fecha_nac_iso=nac_iso,
                fecha_nac_texto=nac_texto,
                notas=c.notas or "",
                puntos=c.puntos,
                activo=c.activo,
                cumple_hoy=cumple_hoy,
                cumple_pronto=cumple_pronto,
                dias_para_cumple=dias,
                visitas_count=stats["visitas"],
                gastado_texto=_money_text(stats["gastado"]),
                ultima_visita_texto=ultima_texto,
                es_vip=stats["visitas"] >= CLIENTE_VIP_VISITAS_MIN,
            ))
        self.clientes_lista = views

    def set_cli_busqueda(self, v: str) -> None:
        self.cli_busqueda = v

    def set_cli_form_nombre(self, v: str) -> None:
        self.cli_form_nombre = v

    def set_cli_form_telefono(self, v: str) -> None:
        self.cli_form_telefono = v

    def set_cli_form_email(self, v: str) -> None:
        self.cli_form_email = v

    def set_cli_form_fecha_nac(self, v: str) -> None:
        self.cli_form_fecha_nac = v

    def set_cli_form_notas(self, v: str) -> None:
        self.cli_form_notas = v

    def guardar_cliente(self) -> None:
        nombre = self.cli_form_nombre.strip()
        if not nombre:
            return rx.toast.error("El nombre del cliente es obligatorio.")
        tel = self.cli_form_telefono.strip() or None
        email = self.cli_form_email.strip() or None
        notas = self.cli_form_notas.strip() or None
        fn: date | None = None
        if self.cli_form_fecha_nac:
            try:
                from datetime import date as _date
                parts = self.cli_form_fecha_nac.split("-")
                fn = _date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, IndexError):
                return rx.toast.error("Fecha de nacimiento inválida. Usa el formato AAAA-MM-DD.")
        with self._tenant_session() as session:
            if self.cli_form_id == 0:
                existente = session.exec(
                    select(Cliente).where(
                        Cliente.company_id == self._company_id(),
                        Cliente.nombre == nombre,
                    )
                ).first()
                if existente:
                    return rx.toast.error("Ya existe un cliente con ese nombre.")
                c = Cliente(
                    company_id=self._company_id(),
                    nombre=nombre,
                    telefono=tel,
                    email=email,
                    fecha_nacimiento=fn,
                    notas=notas,
                    activo=True,
                )
                session.add(c)
                _msg = f"Cliente '{nombre}' registrado."
            else:
                c = session.get(Cliente, self.cli_form_id)
                if c is None or c.company_id != self._company_id():
                    return rx.toast.error("Cliente no encontrado.")
                c.nombre = nombre
                c.telefono = tel
                c.email = email
                c.fecha_nacimiento = fn
                c.notas = notas
                c.updated_at = _utcnow()
                session.add(c)
                _msg = f"Cliente '{nombre}' actualizado."
            session.commit()
        self.cli_form_id = 0
        self.cli_form_nombre = ""
        self.cli_form_telefono = ""
        self.cli_form_email = ""
        self.cli_form_fecha_nac = ""
        self.cli_form_notas = ""
        self.cli_form_editando = False
        self.cli_form_visible = False
        self.cargar_clientes()
        return rx.toast.success(_msg)

    def abrir_nuevo_cliente(self) -> None:
        self.cli_form_id = 0
        self.cli_form_nombre = ""
        self.cli_form_telefono = ""
        self.cli_form_email = ""
        self.cli_form_fecha_nac = ""
        self.cli_form_notas = ""
        self.cli_form_editando = False
        self.cli_form_visible = True
        self.cli_dni_ruc = ""
        self.cli_dni_ruc_error = ""

    def editar_cliente(self, cliente_id: int) -> None:
        with self._tenant_session() as session:
            c = session.get(Cliente, cliente_id)
            if c is None or c.company_id != self._company_id():
                return
            self.cli_form_id = c.id or 0
            self.cli_form_nombre = c.nombre
            self.cli_form_telefono = c.telefono or ""
            self.cli_form_email = c.email or ""
            self.cli_form_fecha_nac = c.fecha_nacimiento.isoformat() if c.fecha_nacimiento else ""
            self.cli_form_notas = c.notas or ""
        self.cli_form_editando = True
        self.cli_form_visible = True

    def cancelar_cli_form(self) -> None:
        self.cli_form_id = 0
        self.cli_form_nombre = ""
        self.cli_form_telefono = ""
        self.cli_form_email = ""
        self.cli_form_fecha_nac = ""
        self.cli_form_notas = ""
        self.cli_form_editando = False
        self.cli_form_visible = False
        self.cli_dni_ruc = ""
        self.cli_dni_ruc_error = ""

    def set_cli_dni_ruc(self, v: str) -> None:
        self.cli_dni_ruc = re.sub(r"\D", "", v)[:11]
        self.cli_dni_ruc_error = ""

    async def buscar_dni_ruc(self) -> None:
        import os
        numero = self.cli_dni_ruc.strip()
        if not numero:
            self.cli_dni_ruc_error = "Ingresa un DNI (8 dígitos) o RUC (11 dígitos)."
            return
        if len(numero) not in (8, 11):
            self.cli_dni_ruc_error = "DNI debe tener 8 dígitos y RUC 11 dígitos."
            return
        token = (os.getenv("RENIEC_API_TOKEN") or "").strip()
        if not token:
            self.cli_dni_ruc_error = "Servicio no configurado (falta RENIEC_API_TOKEN)."
            return
        self.cli_dni_ruc_buscando = True
        self.cli_dni_ruc_error = ""
        yield
        try:
            import httpx
            tipo = "dni" if len(numero) == 8 else "ruc"
            url = f"https://apiperu.dev/api/{tipo}/{numero}"
            async with httpx.AsyncClient(timeout=8.0) as client:
                resp = await client.get(url, headers={"Authorization": f"Bearer {token}"})
            if resp.status_code == 200:
                data = resp.json()
                if tipo == "dni":
                    nombre = data.get("nombre_completo") or data.get("nombre") or ""
                else:
                    nombre = data.get("nombre_o_razon_social") or data.get("razon_social") or ""
                if nombre:
                    self.cli_form_nombre = nombre
                    self.cli_dni_ruc_error = ""
                else:
                    self.cli_dni_ruc_error = "No se encontraron datos para ese número."
            elif resp.status_code == 404:
                self.cli_dni_ruc_error = "Número no encontrado en el registro."
            else:
                self.cli_dni_ruc_error = f"Error del servicio ({resp.status_code})."
        except Exception:
            self.cli_dni_ruc_error = "No se pudo conectar al servicio de consulta."
        finally:
            self.cli_dni_ruc_buscando = False

    def set_cli_form_visible(self, v: bool) -> None:
        self.cli_form_visible = v

    def toggle_cliente_activo(self, cliente_id: int) -> None:
        with self._tenant_session() as session:
            c = session.get(Cliente, cliente_id)
            if c is None or c.company_id != self._company_id():
                return
            c.activo = not c.activo
            c.updated_at = _utcnow()
            session.add(c)
            session.commit()
        self.cargar_clientes()

    # ── Cuentas corrientes ──────────────────────────────────────────────────

    def on_load_cuentas(self):
        result = self._route_access_result("cuentas")
        if result is not None:
            return result
        if not self.clientes_lista:
            self.cargar_clientes()
        self.cargar_cuentas()

    def cargar_cuentas(self) -> None:
        with self._tenant_session() as session:
            cuentas_db = session.exec(
                select(CuentaCorriente)
                .where(CuentaCorriente.company_id == self._company_id())
                .order_by(CuentaCorriente.saldo_deuda.desc())
            ).all()
            clientes_map = {
                c.id: c
                for c in session.exec(
                    select(Cliente).where(Cliente.company_id == self._company_id())
                ).all()
            }
        views: list[CuentaView] = []
        for cc in cuentas_db:
            cli = clientes_map.get(cc.cliente_id)
            saldo = Decimal(str(cc.saldo_deuda))
            views.append(CuentaView(
                id=cc.id or 0,
                cliente_id=cc.cliente_id,
                cliente_nombre=cli.nombre if cli else "?",
                cliente_telefono=cli.telefono or "" if cli else "",
                saldo_deuda=float(saldo),
                saldo_texto=_money_text(saldo),
                limite_credito=float(Decimal(str(cc.limite_credito))),
            ))
        self.cuentas_lista = views

    def set_cc_cliente_sel_nombre(self, v: str) -> None:
        self.cc_cliente_sel_nombre = v
        cli = next((c for c in self.clientes_lista if c.nombre == v), None)
        if cli:
            self._ver_o_crear_cuenta(cli.id)

    def _ver_o_crear_cuenta(self, cliente_id: int) -> None:
        with self._tenant_session() as session:
            cc = session.exec(
                select(CuentaCorriente).where(
                    CuentaCorriente.company_id == self._company_id(),
                    CuentaCorriente.cliente_id == cliente_id,
                )
            ).first()
            if cc:
                self.cuenta_sel_id = cc.id or 0
                movs = session.exec(
                    select(MovimientoCuenta)
                    .where(MovimientoCuenta.cuenta_id == cc.id)
                    .order_by(MovimientoCuenta.created_at.desc())
                ).all()
                self.cuenta_movimientos = [
                    MovimientoView(
                        id=m.id or 0,
                        tipo=m.tipo,
                        tipo_label="Cargo" if m.tipo == "cargo" else "Pago",
                        monto=float(Decimal(str(m.monto))),
                        monto_texto=_money_text(m.monto),
                        descripcion=m.descripcion or "",
                        fecha_texto=m.created_at.strftime("%d/%m %H:%M"),
                    )
                    for m in movs
                ]
            else:
                self.cuenta_sel_id = 0
                self.cuenta_movimientos = []

    def set_cc_pago_monto(self, v: str) -> None:
        self.cc_pago_monto = v

    def set_cc_pago_descripcion(self, v: str) -> None:
        self.cc_pago_descripcion = v

    def registrar_pago_cc(self) -> None:
        if self.cuenta_sel_id == 0:
            return rx.toast.error("Selecciona un cliente con cuenta corriente.")
        try:
            monto = Decimal(self.cc_pago_monto.replace(",", ".").strip() or "0")
            if monto <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            return rx.toast.error("Monto de pago inválido.")
        with self._tenant_session() as session:
            cc = session.get(CuentaCorriente, self.cuenta_sel_id)
            if cc is None or cc.company_id != self._company_id():
                return rx.toast.error("Cuenta no encontrada.")
            saldo_actual = Decimal(str(cc.saldo_deuda))
            if monto > saldo_actual:
                return rx.toast.error(
                    f"El pago ({_money_text(monto)}) supera la deuda "
                    f"({_money_text(saldo_actual)}). Máximo: {_money_text(saldo_actual)}."
                )
            pago = MovimientoCuenta(
                company_id=self._company_id(),
                cuenta_id=cc.id or 0,
                tipo="pago",
                monto=monto,
                descripcion=self.cc_pago_descripcion.strip() or "Pago en caja",
            )
            session.add(pago)
            cc.saldo_deuda = saldo_actual - monto
            cc.updated_at = _utcnow()
            session.add(cc)
            session.commit()
            cliente_id = cc.cliente_id
        self.cc_pago_monto = ""
        self.cc_pago_descripcion = ""
        self.cargar_cuentas()
        self._ver_o_crear_cuenta(cliente_id)
        return rx.toast.success(f"Pago de {_money_text(monto)} registrado.")

    def _registrar_cargo_cc(self, session, cliente_id: int, monto: Decimal, pedido_id: int | None, descripcion: str) -> None:
        cc = session.exec(
            select(CuentaCorriente).where(
                CuentaCorriente.company_id == self._company_id(),
                CuentaCorriente.cliente_id == cliente_id,
            )
        ).first()
        if cc is None:
            cc = CuentaCorriente(
                company_id=self._company_id(),
                cliente_id=cliente_id,
                saldo_deuda=Decimal("0.00"),
                limite_credito=Decimal("0.00"),
            )
            session.add(cc)
            session.flush()
        saldo_actual = Decimal(str(cc.saldo_deuda))
        limite = Decimal(str(cc.limite_credito))
        if limite > Decimal("0.00") and saldo_actual + monto > limite:
            raise ValueError(
                f"Límite de crédito excedido. Deuda actual: {_money_text(saldo_actual)}, "
                f"límite: {_money_text(limite)}, cargo solicitado: {_money_text(monto)}."
            )
        cargo = MovimientoCuenta(
            company_id=self._company_id(),
            cuenta_id=cc.id or 0,
            pedido_id=pedido_id,
            tipo="cargo",
            monto=monto,
            descripcion=descripcion,
        )
        session.add(cargo)
        cc.saldo_deuda = saldo_actual + monto
        cc.updated_at = _utcnow()
        session.add(cc)

    def exportar_cuentas_excel(self):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        with self._tenant_session() as session:
            cuentas_db = session.exec(
                select(CuentaCorriente)
                .where(CuentaCorriente.company_id == self._company_id())
                .order_by(CuentaCorriente.saldo_deuda.desc())
            ).all()
            clientes_map = {
                c.id: c
                for c in session.exec(
                    select(Cliente).where(Cliente.company_id == self._company_id())
                ).all()
            }
            cuenta_ids = [c.id for c in cuentas_db if c.id is not None]
            movimientos: list = []
            if cuenta_ids:
                movimientos = session.exec(
                    select(MovimientoCuenta)
                    .where(MovimientoCuenta.cuenta_id.in_(cuenta_ids))
                    .order_by(MovimientoCuenta.created_at.desc())
                ).all()

        if not cuentas_db and not movimientos:
            return rx.toast.error("No hay cuentas corrientes para exportar.")

        cuentas_por_id = {c.id: c for c in cuentas_db}

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Cuentas"
        ws1.append(["Cliente", "Teléfono", "Saldo deuda", "Límite de crédito"])
        for cc in cuentas_db:
            cli = clientes_map.get(cc.cliente_id)
            ws1.append([
                cli.nombre if cli else "?",
                (cli.telefono or "") if cli else "",
                float(Decimal(str(cc.saldo_deuda))),
                float(Decimal(str(cc.limite_credito))),
            ])

        ws2 = wb.create_sheet("Movimientos")
        ws2.append(["Fecha", "Hora", "Cliente", "Tipo", "Monto", "Descripción"])
        for m in movimientos:
            cc = cuentas_por_id.get(m.cuenta_id)
            cli = clientes_map.get(cc.cliente_id) if cc else None
            ws2.append([
                m.created_at.strftime("%Y-%m-%d") if m.created_at else "",
                m.created_at.strftime("%H:%M") if m.created_at else "",
                cli.nombre if cli else "?",
                "Cargo" if m.tipo == "cargo" else "Pago",
                float(Decimal(str(m.monto))),
                m.descripcion or "",
            ])

        for ws in (ws1, ws2):
            for i, col in enumerate(ws.columns, start=1):
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"cuentas_corrientes_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)
