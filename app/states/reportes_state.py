"""Substate independiente de reportes — dashboard KPIs, historial, analítica, Excel.

Hereda de rx.State (NO de FoodState). Accede a datos de tenant vía
`await self.get_state(FoodState)` para obtener `_company_id()`,
`_tenant_session()`, y vars compartidas.

Beneficio: las ~10 páginas que no usan reportes (mozos, cocina, mostrador,
login, menú público, inventario, clientes, cupones, promociones, usuarios)
ya no serializan ~50 vars de reportes por WebSocket.
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta
from decimal import Decimal

import reflex as rx
from sqlalchemy import and_, func, or_
from sqlmodel import select

from app.models.food import (
    DetallePedido,
    EstadoMesa,
    EstadoPedido,
    EstadoTurnoCaja,
    Mesa,
    PagoPedido,
    Pedido,
    Producto,
    Reserva,
    TurnoCaja,
    UsuarioFood,
)
from app.services.analitica_service import (
    margen_por_plato,
    matriz_estrella_perro,
    reporte_productos_stock,
    ventas_por_hora,
    ventas_por_mozo,
)
from app.services.finanzas_service import (
    pyl_mensual,
    reporte_anulaciones,
    reporte_descuentos,
    reporte_mermas,
    reporte_reversiones,
    resumen_igv_mensual,
)
from app.services.metodos_pago_service import mapa_nombres, mapa_tipos
from app.services.plan_service import plan_permite
from app.states.caja_turno_mixin import _bucket_tipo
from tuwayki_core.utils.timezone import (
    country_today_date,
    format_local_datetime,
    local_day_bounds_utc_naive,
)
from app.states.food_state import (
    AnulacionView,
    DescuentoRankView,
    ReversionView,
    FranjaHoraView,
    MargenPlatoView,
    MatrizProductoView,
    MermaCategoriaView,
    MermaInsumoView,
    MozoRankView,
    PylLineView,
    TopPlatoView,
    VentaDetalleItemView,
    VentaHistorialView,
    _actor_name,
    _money_text,
    _pedido_sales_label,
    _to_decimal,
    _utcnow,
)


from app.utils.excel_export import (
    escribir_encabezado as _excel_titulo,
    finalizar_hoja as _excel_finalizar,
)


class ReportesState(rx.State):
    """Dashboard KPIs, historial de ventas, analítica y exportación Excel."""

    # ── Filtro sucursal (MT-03b) ───────────────────────────────────────────
    reportes_sucursal_id: int = 0

    # ── Dashboard KPIs ──────────────────────────────────────────────────────
    dashboard_ventas_hoy_texto: str = "S/ 0.00"
    dashboard_pedidos_hoy: int = 0
    dashboard_mesas_ocupadas: int = 0
    dashboard_total_mesas: int = 0
    dashboard_items_en_cocina: int = 0
    dashboard_reservas_hoy: int = 0
    dashboard_propina_hoy_texto: str = "S/ 0.00"
    dashboard_ticket_promedio_texto: str = "S/ 0.00"
    dashboard_top_platos: list[TopPlatoView] = []
    dashboard_ventas_trend_pct: int = 0
    dashboard_pedidos_trend: int = 0
    dashboard_ticket_trend_pct: int = 0
    dashboard_propina_trend_pct: int = 0
    # Etiqueta del período activo de los KPIs (sigue al filtro Hoy/Semana/Mes de
    # la página /reportes). El panel /admin lo ancla siempre a "Hoy".
    dashboard_periodo_label: str = "Hoy"

    # ── Analítica ────────────────────────────────────────────────────────────
    reporte_mozos: list[MozoRankView] = []
    reporte_horas: list[FranjaHoraView] = []
    reporte_margen: list[MargenPlatoView] = []
    reporte_metodos: list[dict[str, object]] = []

    # ── Comparativa entre períodos ──────────────────────────────────────────
    comp_label_actual: str = "Período actual"
    comp_label_anterior: str = "Período anterior"
    comp_ventas_actual: str = "S/ 0.00"
    comp_ventas_anterior: str = "S/ 0.00"
    comp_ventas_pct: int = 0
    comp_pedidos_actual: int = 0
    comp_pedidos_anterior: int = 0
    comp_pedidos_diff: int = 0
    comp_ticket_actual: str = "S/ 0.00"
    comp_ticket_anterior: str = "S/ 0.00"
    comp_ticket_pct: int = 0

    # País de la empresa (ISO-2) para agrupar por día local. Se refresca desde la
    # config al cargar dashboard/historial.
    reportes_pais: str = "PE"

    # ── Historial ────────────────────────────────────────────────────────────
    historial_ventas: list[VentaHistorialView] = []
    historial_filtro_fecha_desde: str = ""
    historial_filtro_fecha_hasta: str = ""
    historial_filtro_metodo: str = ""
    historial_filtro_rapido: str = "hoy"
    historial_pagina: int = 0
    historial_total: int = 0
    _HISTORIAL_PAGE_SIZE: int = 50

    # ── P&L mensual (ADM-01) ────────────────────────────────────────────────
    pyl_lineas: list[PylLineView] = []
    pyl_anio: int = 0
    pyl_mes: int = 0

    # ── Descuentos y anulaciones (ADM-02) ────────────────────────────────────
    descuentos_rank: list[DescuentoRankView] = []
    anulaciones_lista: list[AnulacionView] = []
    anulaciones_total_texto: str = "S/ 0.00"
    descuentos_total_texto: str = "S/ 0.00"
    reversiones_lista: list[ReversionView] = []
    reversiones_total_texto: str = "S/ 0.00"

    # ── Propinas por mozo (ADM-04) ──────────────────────────────────────────
    reporte_propinas_total_texto: str = "S/ 0.00"

    # ── Resumen IGV (ADM-05) ────────────────────────────────────────────────
    igv_base_imponible_texto: str = "S/ 0.00"
    igv_monto_texto: str = "S/ 0.00"
    igv_ventas_netas_texto: str = "S/ 0.00"
    igv_porcentaje: float = 18.0
    igv_pedidos: int = 0
    igv_anio: int = 0
    igv_mes: int = 0

    # ── Mermas (ADM-03) ─────────────────────────────────────────────────────
    mermas_por_categoria: list[MermaCategoriaView] = []
    mermas_por_insumo: list[MermaInsumoView] = []
    mermas_total_texto: str = "S/ 0.00"

    # ── Matriz estrella/perro (ADM-06) ──────────────────────────────────────
    matriz_productos: list[MatrizProductoView] = []
    matriz_estrellas: int = 0
    matriz_vacas: int = 0
    matriz_puzzles: int = 0
    matriz_perros: int = 0

    # ── Detalle de venta (modal) ────────────────────────────────────────────
    venta_detalle_visible: bool = False
    venta_detalle_pedido_id: int = 0
    venta_detalle_mesa_label: str = ""
    venta_detalle_metodo: str = ""
    venta_detalle_mozo: str = ""
    venta_detalle_cajero: str = ""
    venta_detalle_total_texto: str = ""
    venta_detalle_propina_texto: str = ""
    venta_detalle_items: list[VentaDetalleItemView] = []

    # ── Computed vars ────────────────────────────────────────────────────────

    @rx.var
    def historial_ventas_recientes(self) -> list[VentaHistorialView]:
        return self.historial_ventas[:5]

    @rx.var
    def reporte_horas_chart(self) -> list[dict[str, object]]:
        return [{"hora_label": h.hora_label, "total": h.total, "pedidos": h.pedidos} for h in self.reporte_horas]

    @rx.var
    def reporte_mozos_chart(self) -> list[dict[str, object]]:
        return [{"nombre": m.nombre, "total": m.total, "propinas": m.propinas} for m in self.reporte_mozos]

    @rx.var
    def historial_filtro_activo(self) -> bool:
        return bool(
            self.historial_filtro_fecha_desde
            or self.historial_filtro_fecha_hasta
            or self.historial_filtro_metodo
        )

    @rx.var
    def historial_tiene_anterior(self) -> bool:
        return self.historial_pagina > 0

    @rx.var
    def historial_tiene_siguiente(self) -> bool:
        return (self.historial_pagina + 1) * self._HISTORIAL_PAGE_SIZE < self.historial_total

    @rx.var
    def historial_pagina_label(self) -> str:
        if self.historial_total == 0:
            return "Sin resultados"
        desde = self.historial_pagina * self._HISTORIAL_PAGE_SIZE + 1
        hasta = min((self.historial_pagina + 1) * self._HISTORIAL_PAGE_SIZE, self.historial_total)
        return f"{desde}–{hasta} de {self.historial_total}"

    # ── Helpers internos ─────────────────────────────────────────────────────

    async def _food(self):
        from app.states.food_state import FoodState
        return await self.get_state(FoodState)

    def _sucursal_q(self, model_class, query):
        sid = self.reportes_sucursal_id
        if sid and hasattr(model_class, "sucursal_id"):
            return query.where(model_class.sucursal_id == sid)
        return query

    async def cambiar_sucursal_reportes(self, v: str) -> None:
        try:
            self.reportes_sucursal_id = int(v)
        except ValueError:
            self.reportes_sucursal_id = 0
        food = await self._food()
        self._do_cargar_dashboard(food)
        self._do_cargar_historial(food)

    def _rango_filtros_historial(self) -> tuple[datetime | None, datetime | None]:
        """Filtros de fecha (día LOCAL) convertidos a límites UTC-naive.

        Los timestamps se guardan en UTC-naive; para que "un día" coincida con el
        día local del país de la empresa (no el día UTC), convertimos los bordes
        con la zona horaria del país. Así un turno de noche no se parte en dos.
        """
        cc = self.reportes_pais
        desde = hasta = None
        if self.historial_filtro_fecha_desde:
            try:
                desde, _ = local_day_bounds_utc_naive(
                    self.historial_filtro_fecha_desde, cc
                )
            except ValueError:
                pass
        if self.historial_filtro_fecha_hasta:
            try:
                _, hasta = local_day_bounds_utc_naive(
                    self.historial_filtro_fecha_hasta, cc
                )
            except ValueError:
                pass
        return desde, hasta

    def _simbolo_moneda(self) -> str:
        return "S/" if self.reportes_pais == "PE" else "$"

    def _rango_export_texto(self) -> str:
        """Texto legible del rango de fechas aplicado, para el encabezado del Excel."""
        d = (self.historial_filtro_fecha_desde or "").strip()
        h = (self.historial_filtro_fecha_hasta or "").strip()
        if not d and not h:
            return "Todo el histórico (sin filtro de fecha)"
        if d and h:
            return f"{d} a {h}"
        if d:
            return f"Desde {d}"
        return f"Hasta {h}"

    async def _excel_meta(self, extra: list[tuple[str, str]] | None = None):
        """Metadatos comunes del encabezado de los Excel: Empresa + Generado, más
        las líneas ``extra`` (período/rango) que aporte cada reporte."""
        food = await self._food()
        self.reportes_pais = food._country_code()
        empresa = getattr(food, "empresa_nombre", "TUWAYKIFOOD") or "TUWAYKIFOOD"
        generado = format_local_datetime(_utcnow(), "%d/%m/%Y %H:%M", self.reportes_pais)
        meta = [("Empresa", empresa), ("Generado", generado)]
        if extra:
            meta.extend(extra)
        return meta

    # ── on_load ──────────────────────────────────────────────────────────────

    async def on_load_reportes(self):
        food = await self._food()
        food.pagina_cargada = False
        result = food._route_access_result(
            "reportes",
            also_allowed=food.usuario_actual is not None and food.usuario_actual.perm_reportes,
        )
        if result is not None:
            return result
        self.reportes_pais = food._country_code()
        self.historial_filtro_fecha_desde = country_today_date(
            self.reportes_pais
        ).strftime("%Y-%m-%d")
        self.historial_filtro_rapido = "hoy"
        self._do_cargar_dashboard(food)
        self._do_cargar_historial(food)
        self._do_cargar_analitica(food)
        if plan_permite(food.empresa_plan, "reportes_avanzados"):
            self._do_cargar_pyl(food)
            self._do_cargar_igv(food)
            self._do_cargar_descuentos_anulaciones(food)
            self._do_cargar_mermas(food)
            self._do_cargar_matriz(food)
        food.pagina_cargada = True
        return None

    async def init_reportes_dono(self):
        """Inicializa reportes para el panel dueño."""
        food = await self._food()
        self.reportes_pais = food._country_code()
        self.historial_filtro_fecha_desde = country_today_date(
            self.reportes_pais
        ).strftime("%Y-%m-%d")
        self.historial_filtro_fecha_hasta = ""
        self.historial_filtro_metodo = ""
        # El panel /admin siempre muestra el día de hoy; ancla el filtro para que
        # los KPIs no arrastren un período (Semana/Mes) de una visita a /reportes.
        self.historial_filtro_rapido = "hoy"
        self._do_cargar_dashboard(food)
        self._do_cargar_historial(food)

    # ── Dashboard KPIs ───────────────────────────────────────────────────────

    async def cargar_dashboard(self) -> None:
        food = await self._food()
        self._do_cargar_dashboard(food)

    def _rango_dashboard(self) -> tuple[datetime, datetime, datetime, datetime, str]:
        """Rango de los KPIs del dashboard según el filtro rápido activo.

        Devuelve (inicio, fin, inicio_prev, fin_prev, etiqueta) con 'fin'
        exclusivo. El período anterior es una ventana de la MISMA duración
        inmediatamente anterior (hoy→ayer, semana→semana previa, etc.), para que
        los trends comparen manzanas con manzanas. Sin filtro (o "hoy") el rango
        es el día de hoy — así el panel /admin sigue mostrando "hoy"."""
        cc = self.reportes_pais
        hoy = country_today_date(cc)

        def _parse(s: str):
            try:
                return datetime.strptime(s, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                return None

        desde = _parse(self.historial_filtro_fecha_desde)
        hasta = _parse(self.historial_filtro_fecha_hasta)
        rapido = self.historial_filtro_rapido
        if rapido in ("", "hoy") or desde is None:
            dia_ini = dia_fin = hoy
            etiqueta = "Hoy"
        else:
            if hasta is None or hasta < desde:
                hasta = desde
            dia_ini, dia_fin = desde, hasta
            etiqueta = {
                "semana": "Esta semana",
                "mes": "Este mes",
            }.get(rapido, "Período personalizado")
        # Bordes del día LOCAL convertidos a UTC-naive (inicio inclusivo, fin
        # exclusivo = inicio del día siguiente al último día del rango).
        inicio, _ = local_day_bounds_utc_naive(dia_ini, cc)
        fin, _ = local_day_bounds_utc_naive(dia_fin + timedelta(days=1), cc)
        dur = fin - inicio
        return inicio, fin, inicio - dur, inicio, etiqueta

    def _do_cargar_dashboard(self, food) -> None:
        self.reportes_pais = food._country_code()
        hoy = country_today_date(self.reportes_pais)
        inicio_hoy, fin_hoy, inicio_ayer, fin_ayer, periodo_label = self._rango_dashboard()
        with food._tenant_session() as session:
            # PERF-04: agregación en SQL (func.count/sum) en vez de cargar todos
            # los pedidos del período como objetos y sumar en Python.
            cobrado = or_(
                Pedido.pagado.is_(True),
                Pedido.estado == EstadoPedido.COBRADO.value,
            )

            def _agg_ventas(inicio, fin):
                q = select(
                    func.count(Pedido.id),
                    func.coalesce(func.sum(Pedido.total), 0),
                    func.coalesce(func.sum(Pedido.propina), 0),
                    func.coalesce(func.sum(Pedido.descuento), 0),
                    func.coalesce(func.sum(Pedido.recargo), 0),
                ).where(
                    Pedido.company_id == food._company_id(),
                    cobrado,
                    Pedido.cerrado_en >= inicio,
                    Pedido.cerrado_en < fin,
                )
                cnt, tot, prop, desc, rec = session.exec(self._sucursal_q(Pedido, q)).one()
                # Venta neta = Σ(subtotal + recargo − descuento). La propina NO es
                # venta (es del personal); se reporta aparte.
                neta = _to_decimal(tot or 0) + _to_decimal(rec or 0) - _to_decimal(desc or 0)
                return int(cnt or 0), neta, _to_decimal(prop or 0)

            pedidos_count_hoy, sum_total_hoy, propina_total = _agg_ventas(inicio_hoy, fin_hoy)
            pedidos_count_ayer, sum_total_ayer, propina_ayer = _agg_ventas(inicio_ayer, inicio_hoy)
            # "Ventas" = venta neta (sin propina). Ver _agg_ventas.
            ventas = sum_total_hoy
            ventas_ayer = sum_total_ayer
            q_mesas = select(Mesa).where(
                Mesa.company_id == food._company_id(),
                Mesa.estado != EstadoMesa.LIBRE.value,
                Mesa.activa.is_(True),
            )
            mesas_no_libres = session.exec(self._sucursal_q(Mesa, q_mesas)).all()
            q_total = select(func.count(Mesa.id)).where(
                Mesa.company_id == food._company_id(),
                Mesa.activa.is_(True),
            )
            total_mesas_activas = session.exec(self._sucursal_q(Mesa, q_total)).one()
            # "En cocina" = líneas realmente en producción, con el MISMO criterio
            # que el KDS (cargar_cocina): enviadas a cocina, que requieren
            # preparación y cuyo pedido no está cancelado ni cobrado. Sin el JOIN
            # a Pedido, las líneas 'pendiente' de pedidos cancelados quedaban
            # contadas para siempre (contador fantasma en el Resumen).
            q_cocina = (
                select(func.count(DetallePedido.id))
                .join(Pedido, DetallePedido.pedido_id == Pedido.id)
                .where(
                    DetallePedido.company_id == food._company_id(),
                    DetallePedido.impreso_cocina.is_(True),
                    DetallePedido.requiere_preparacion.is_(True),
                    DetallePedido.estado_produccion.in_(["pendiente", "en_preparacion"]),
                    Pedido.estado.notin_(
                        [EstadoPedido.CANCELADO.value, EstadoPedido.COBRADO.value]
                    ),
                )
            )
            items_en_cocina = session.exec(q_cocina).one()
            q_reservas = select(func.count(Reserva.id)).where(
                Reserva.company_id == food._company_id(),
                Reserva.fecha == hoy,
                Reserva.estado.in_(["pendiente", "confirmada"]),
            )
            reservas_hoy_count = session.exec(self._sucursal_q(Reserva, q_reservas)).one()
            # Top 5 platos del día — agregado en SQL con JOIN (PERF-04): evita
            # cargar todos los detalles y todos los productos del tenant.
            top_q = (
                select(
                    DetallePedido.producto_id,
                    func.sum(DetallePedido.cantidad).label("cantidad"),
                    func.sum(DetallePedido.subtotal).label("total"),
                )
                .join(Pedido, DetallePedido.pedido_id == Pedido.id)
                .where(
                    Pedido.company_id == food._company_id(),
                    cobrado,
                    Pedido.cerrado_en >= inicio_hoy,
                    Pedido.cerrado_en < fin_hoy,
                )
                .group_by(DetallePedido.producto_id)
                .order_by(func.sum(DetallePedido.cantidad).desc())
                .limit(5)
            )
            top_rows = session.exec(self._sucursal_q(Pedido, top_q)).all()
            top_pids = [r[0] for r in top_rows]
            nombres_top: dict[int, str] = {}
            if top_pids:
                nombres_top = {
                    p.id: p.nombre
                    for p in session.exec(
                        select(Producto).where(Producto.id.in_(top_pids))
                    ).all()
                }
            top_sorted = [
                {
                    "nombre": nombres_top.get(pid, f"Producto {pid}"),
                    "cantidad": int(cant or 0),
                    "total": _to_decimal(tot or 0),
                }
                for pid, cant, tot in top_rows
            ]

        def _pct_change(hoy_val: Decimal, ayer_val: Decimal) -> int:
            if ayer_val <= 0:
                return 100 if hoy_val > 0 else 0
            return int(round((hoy_val - ayer_val) / ayer_val * 100))

        ticket_promedio = ventas / pedidos_count_hoy if pedidos_count_hoy else Decimal("0.00")
        ticket_promedio_ayer = (
            ventas_ayer / pedidos_count_ayer if pedidos_count_ayer else Decimal("0.00")
        )
        self.dashboard_periodo_label = periodo_label
        self.dashboard_ventas_hoy_texto = _money_text(ventas)
        self.dashboard_pedidos_hoy = pedidos_count_hoy
        self.dashboard_mesas_ocupadas = len(mesas_no_libres)
        self.dashboard_total_mesas = total_mesas_activas or 0
        self.dashboard_items_en_cocina = items_en_cocina or 0
        self.dashboard_reservas_hoy = reservas_hoy_count or 0
        self.dashboard_propina_hoy_texto = _money_text(propina_total)
        self.dashboard_ticket_promedio_texto = _money_text(ticket_promedio)
        self.dashboard_ventas_trend_pct = _pct_change(ventas, ventas_ayer)
        self.dashboard_pedidos_trend = pedidos_count_hoy - pedidos_count_ayer
        self.dashboard_ticket_trend_pct = _pct_change(ticket_promedio, ticket_promedio_ayer)
        self.dashboard_propina_trend_pct = _pct_change(propina_total, propina_ayer)
        self.dashboard_top_platos = [
            TopPlatoView(
                nombre=p["nombre"],
                cantidad=p["cantidad"],
                total_generado=float(p["total"]),
                total_texto=_money_text(p["total"]),
            )
            for p in top_sorted
        ]

    # ── Historial de ventas ──────────────────────────────────────────────────

    async def cargar_historial_ventas(self) -> None:
        food = await self._food()
        self._do_cargar_historial(food)

    def _do_cargar_historial(self, food) -> None:
        self.reportes_pais = food._country_code()
        with food._tenant_session() as session:
            query = (
                select(Pedido)
                .where(
                    Pedido.company_id == food._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                        and_(
                            Pedido.estado == EstadoPedido.CANCELADO.value,
                            Pedido.cerrado_en.isnot(None),
                        ),
                    ),
                )
            )
            query = self._sucursal_q(Pedido, query)
            desde_utc, hasta_utc = self._rango_filtros_historial()
            if desde_utc is not None:
                query = query.where(Pedido.cerrado_en >= desde_utc)
            if hasta_utc is not None:
                query = query.where(Pedido.cerrado_en <= hasta_utc)
            if self.historial_filtro_metodo:
                from app.models.food import PagoPedido as _PP
                query = query.where(
                    or_(
                        Pedido.metodo_pago == self.historial_filtro_metodo,
                        Pedido.id.in_(
                            select(_PP.pedido_id).where(
                                _PP.company_id == food._company_id(),
                                _PP.metodo == self.historial_filtro_metodo,
                            )
                        ),
                    )
                )
            # PERF-04: contar en SQL en vez de materializar todo el período; solo
            # la página visible se carga como objetos.
            total_count = session.exec(
                select(func.count()).select_from(query.subquery())
            ).one()
            self.historial_total = int(total_count or 0)
            query = query.order_by(Pedido.cerrado_en.desc(), Pedido.id.desc())
            offset = self.historial_pagina * self._HISTORIAL_PAGE_SIZE
            pedidos = session.exec(query.offset(offset).limit(self._HISTORIAL_PAGE_SIZE)).all()
            mesas = {m.id: m for m in session.exec(select(Mesa).where(Mesa.company_id == food._company_id())).all()}
            usuarios = {u.id: u for u in session.exec(select(UsuarioFood).where(UsuarioFood.company_id == food._company_id())).all()}
            # Turnos abiertos de la empresa: una venta es "corregible/anulable en
            # caja" solo si pertenece a un turno todavía abierto (misma regla que
            # Caja). Guardamos los ids para marcar cada fila.
            turnos_abiertos = {
                t.id for t in session.exec(
                    select(TurnoCaja).where(
                        TurnoCaja.company_id == food._company_id(),
                        TurnoCaja.estado == EstadoTurnoCaja.ABIERTO.value,
                    )
                ).all()
            }
            historial: list[VentaHistorialView] = []
            for p in pedidos:
                # Venta neta = subtotal + recargo (envases/delivery) − descuento.
                total_base = (
                    _to_decimal(p.total)
                    + _to_decimal(getattr(p, "recargo", Decimal("0.00")))
                    - _to_decimal(getattr(p, "descuento", Decimal("0.00")))
                )
                propina = _to_decimal(getattr(p, "propina", Decimal("0.00")))
                total_con_propina = total_base + propina
                anulada = p.estado == EstadoPedido.CANCELADO.value
                anulacion_texto = ""
                if anulada:
                    quien = usuarios.get(p.cancelado_por_id)
                    anulacion_texto = (p.motivo_cancelacion or "Sin motivo") + (
                        f" — {quien.nombre}" if quien else ""
                    )
                metodo_cod = getattr(p, "metodo_pago", None) or "—"
                historial.append(VentaHistorialView(
                    pedido_id=p.id or 0,
                    mesa_label=_pedido_sales_label(p, mesas),
                    total=float(total_base),
                    total_texto=_money_text(total_base),
                    propina=float(propina),
                    propina_texto=_money_text(propina) if propina > 0 else "",
                    total_con_propina=float(total_con_propina),
                    total_con_propina_texto=_money_text(total_con_propina),
                    metodo_pago=metodo_cod,
                    mozo_nombre=_actor_name(usuarios[p.mozo_id].nombre if p.mozo_id in usuarios else "Sin asignar"),
                    cajero_nombre=_actor_name(usuarios[p.cajero_id].nombre if p.cajero_id in usuarios else "Sin asignar"),
                    anulada=anulada,
                    anulacion_texto=anulacion_texto,
                    fecha_texto=format_local_datetime(p.cerrado_en, "%d/%m/%Y", self.reportes_pais) if p.cerrado_en else "",
                    hora_texto=format_local_datetime(p.cerrado_en, "%H:%M", self.reportes_pais) if p.cerrado_en else "",
                    comprobante_impreso=getattr(p, "comprobante_impreso_at", None) is not None,
                    es_turno_actual=bool(p.turno_caja_id and p.turno_caja_id in turnos_abiertos),
                ))
            self.historial_ventas = historial

    # ── Filtros de historial ─────────────────────────────────────────────────

    def set_historial_filtro_fecha_desde(self, v: str) -> None:
        self.historial_filtro_fecha_desde = v

    def set_historial_filtro_fecha_hasta(self, v: str) -> None:
        self.historial_filtro_fecha_hasta = v

    def set_historial_filtro_metodo(self, v: str) -> None:
        self.historial_filtro_metodo = v

    async def aplicar_filtros_historial(self) -> None:
        food = await self._food()
        self.historial_pagina = 0
        # Los KPIs superiores deben seguir al mismo período que los gráficos
        # (antes quedaban fijos en "hoy" y contradecían al gráfico de abajo).
        self._do_cargar_dashboard(food)
        self._do_cargar_historial(food)
        self._do_cargar_analitica(food)
        self._do_cargar_descuentos_anulaciones(food)
        self._do_cargar_mermas(food)

    async def buscar_historial_manual(self) -> None:
        self.historial_filtro_rapido = "personalizado"
        await self.aplicar_filtros_historial()

    async def limpiar_filtros_historial(self) -> None:
        food = await self._food()
        self.historial_filtro_fecha_desde = ""
        self.historial_filtro_fecha_hasta = ""
        self.historial_filtro_metodo = ""
        self.historial_filtro_rapido = ""
        self.historial_pagina = 0
        self._do_cargar_dashboard(food)
        self._do_cargar_historial(food)
        self._do_cargar_analitica(food)

    async def filtro_rapido_hoy(self) -> None:
        hoy = _utcnow().date().isoformat()
        self.historial_filtro_fecha_desde = hoy
        self.historial_filtro_fecha_hasta = hoy
        self.historial_filtro_rapido = "hoy"
        await self.aplicar_filtros_historial()

    async def filtro_rapido_semana(self) -> None:
        hoy = _utcnow().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        self.historial_filtro_fecha_desde = inicio_semana.isoformat()
        self.historial_filtro_fecha_hasta = hoy.isoformat()
        self.historial_filtro_rapido = "semana"
        await self.aplicar_filtros_historial()

    async def filtro_rapido_mes(self) -> None:
        hoy = _utcnow().date()
        inicio_mes = hoy.replace(day=1)
        self.historial_filtro_fecha_desde = inicio_mes.isoformat()
        self.historial_filtro_fecha_hasta = hoy.isoformat()
        self.historial_filtro_rapido = "mes"
        await self.aplicar_filtros_historial()

    # ── Detalle de venta ─────────────────────────────────────────────────────

    async def abrir_detalle_venta(self, pedido_id: int) -> None:
        venta = next((v for v in self.historial_ventas if v.pedido_id == pedido_id), None)
        if venta is None:
            return
        food = await self._food()
        with food._tenant_session() as session:
            detalles = session.exec(
                select(DetallePedido).where(
                    DetallePedido.company_id == food._company_id(),
                    DetallePedido.pedido_id == pedido_id,
                )
            ).all()
            productos = {
                p.id: p for p in session.exec(
                    select(Producto).where(Producto.company_id == food._company_id())
                ).all()
            }
        self.venta_detalle_items = [
            VentaDetalleItemView(
                nombre=productos[d.producto_id].nombre if d.producto_id in productos else f"Producto {d.producto_id}",
                cantidad=d.cantidad,
                precio_unitario_texto=_money_text(d.precio_unitario),
                subtotal_texto=_money_text(d.subtotal),
                notas=d.notas or "",
            )
            for d in detalles
        ]
        self.venta_detalle_pedido_id = pedido_id
        self.venta_detalle_mesa_label = venta.mesa_label
        self.venta_detalle_metodo = venta.metodo_pago
        self.venta_detalle_mozo = venta.mozo_nombre
        self.venta_detalle_cajero = venta.cajero_nombre
        self.venta_detalle_total_texto = venta.total_con_propina_texto
        self.venta_detalle_propina_texto = venta.propina_texto
        self.venta_detalle_visible = True

    def set_venta_detalle_visible(self, v: bool) -> None:
        self.venta_detalle_visible = v

    # ── Paginación ───────────────────────────────────────────────────────────

    async def historial_pagina_anterior(self) -> None:
        if self.historial_pagina > 0:
            self.historial_pagina -= 1
            food = await self._food()
            self._do_cargar_historial(food)

    async def historial_pagina_siguiente(self) -> None:
        if self.historial_tiene_siguiente:
            self.historial_pagina += 1
            food = await self._food()
            self._do_cargar_historial(food)

    # ── Analítica ────────────────────────────────────────────────────────────

    async def cargar_analitica(self) -> None:
        food = await self._food()
        self._do_cargar_analitica(food)

    def _do_cargar_analitica(self, food) -> None:
        self.reportes_pais = food._country_code()
        desde, hasta = self._rango_filtros_historial()
        pagos_por_metodo: dict[str, dict[str, object]] = {}
        with food._tenant_session() as session:
            mozos = ventas_por_mozo(session, food._company_id(), desde, hasta)
            horas = ventas_por_hora(
                session, food._company_id(), desde, hasta, self.reportes_pais
            )
            margenes = margen_por_plato(session, food._company_id())
            q_pagos = select(PagoPedido).where(PagoPedido.company_id == food._company_id())
            if desde is not None:
                q_pagos = q_pagos.where(PagoPedido.created_at >= desde)
            if hasta is not None:
                q_pagos = q_pagos.where(PagoPedido.created_at <= hasta)
            for pago in session.exec(q_pagos).all():
                m = pago.metodo or "otro"
                if m not in pagos_por_metodo:
                    pagos_por_metodo[m] = {"total": 0.0, "count": 0}
                pagos_por_metodo[m]["total"] = round(float(pagos_por_metodo[m]["total"]) + float(pago.monto), 2)
                pagos_por_metodo[m]["count"] = int(pagos_por_metodo[m]["count"]) + 1
        self.reporte_mozos = [
            MozoRankView(
                nombre=f["nombre"],
                pedidos=f["pedidos"],
                total=float(f["total"]),
                total_texto=_money_text(f["total"]),
                propinas=float(f["propinas"]),
                propinas_texto=_money_text(f["propinas"]) if f["propinas"] > 0 else "",
            )
            for f in mozos[:10]
        ]
        total_propinas = sum(float(f["propinas"]) for f in mozos)
        self.reporte_propinas_total_texto = _money_text(Decimal(str(total_propinas)))
        max_total = max((float(f["total"]) for f in horas), default=0.0)
        self.reporte_horas = [
            FranjaHoraView(
                hora_label=f"{f['hora']:02d}:00",
                pedidos=f["pedidos"],
                total=float(f["total"]),
                total_texto=_money_text(f["total"]),
                barra_pct=int(float(f["total"]) / max_total * 100) if max_total > 0 else 0,
            )
            for f in horas
        ]
        vistas_margen: list[MargenPlatoView] = []
        for f in margenes:
            pct = f["margen_pct"]
            if not f["costo_completo"]:
                color = "#94A3B8"
            elif pct < 30:
                color = "#DC2626"
            elif pct < 60:
                color = "#D97706"
            else:
                color = "#16A34A"
            vistas_margen.append(MargenPlatoView(
                nombre=f["nombre"],
                precio_texto=_money_text(f["precio"]),
                costo_texto=_money_text(f["costo"]),
                margen_texto=_money_text(f["margen"]),
                margen_pct_texto=f"{pct:.1f}%",
                color=color,
                costo_completo=f["costo_completo"],
                precio=float(f["precio"]),
                costo=float(f["costo"]),
                margen=float(f["margen"]),
                margen_pct=float(pct),
            ))
        self.reporte_margen = vistas_margen
        labels = {"efectivo": "Efectivo", "tarjeta": "Tarjeta", "qr": "QR / Yape", "fiado": "Fiado"}
        self.reporte_metodos = [
            {"metodo": labels.get(m, m.title()), "total": v["total"], "count": v["count"]}
            for m, v in sorted(pagos_por_metodo.items(), key=lambda x: -float(x[1]["total"]))
        ]
        self._do_cargar_comparativa(food)

    def _do_cargar_comparativa(self, food) -> None:
        desde, hasta = self._rango_filtros_historial()
        if desde is None:
            desde = datetime.combine(_utcnow().date(), datetime.min.time())
        if hasta is None:
            hasta = datetime.combine(_utcnow().date(), datetime.max.time())
        delta = hasta - desde
        anterior_hasta = desde - timedelta(seconds=1)
        anterior_desde = anterior_hasta - delta

        dias_actual = max(delta.days, 1)
        if dias_actual <= 1:
            self.comp_label_actual = "Hoy"
            self.comp_label_anterior = "Ayer"
        elif dias_actual <= 7:
            self.comp_label_actual = "Esta semana"
            self.comp_label_anterior = "Semana anterior"
        else:
            self.comp_label_actual = f"Últimos {dias_actual} días"
            self.comp_label_anterior = f"{dias_actual} días previos"

        cobrado_filter = or_(
            Pedido.pagado.is_(True),
            Pedido.estado == EstadoPedido.COBRADO.value,
        )
        with food._tenant_session() as session:
            # PERF-04: agregación en SQL en vez de cargar los pedidos de ambos
            # períodos como objetos.
            def _agg_comp(d0, d1):
                q = select(
                    func.count(Pedido.id),
                    func.coalesce(func.sum(Pedido.total), 0),
                    func.coalesce(func.sum(Pedido.descuento), 0),
                    func.coalesce(func.sum(Pedido.recargo), 0),
                ).where(
                    Pedido.company_id == food._company_id(),
                    cobrado_filter,
                    Pedido.cerrado_en >= d0,
                    Pedido.cerrado_en <= d1,
                )
                cnt, tot, desc, rec = session.exec(q).one()
                # Venta neta (sin propina), consistente con el KPI del dashboard.
                neta = _to_decimal(tot or 0) + _to_decimal(rec or 0) - _to_decimal(desc or 0)
                return int(cnt or 0), neta

            count_act, ventas_act = _agg_comp(desde, hasta)
            count_ant, ventas_ant = _agg_comp(anterior_desde, anterior_hasta)

        ticket_act = ventas_act / count_act if count_act else Decimal("0.00")
        ticket_ant = ventas_ant / count_ant if count_ant else Decimal("0.00")

        def _pct(a: Decimal, b: Decimal) -> int:
            if b <= 0:
                return 100 if a > 0 else 0
            return int(round((a - b) / b * 100))

        self.comp_ventas_actual = _money_text(ventas_act)
        self.comp_ventas_anterior = _money_text(ventas_ant)
        self.comp_ventas_pct = _pct(ventas_act, ventas_ant)
        self.comp_pedidos_actual = count_act
        self.comp_pedidos_anterior = count_ant
        self.comp_pedidos_diff = count_act - count_ant
        self.comp_ticket_actual = _money_text(ticket_act)
        self.comp_ticket_anterior = _money_text(ticket_ant)
        self.comp_ticket_pct = _pct(ticket_act, ticket_ant)

    # ── Excel exports ────────────────────────────────────────────────────────

    async def exportar_ventas_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        food = await self._food()
        self.reportes_pais = food._country_code()
        with food._tenant_session() as session:
            query = (
                select(Pedido)
                .where(
                    Pedido.company_id == food._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                )
            )
            desde_utc, hasta_utc = self._rango_filtros_historial()
            if desde_utc is not None:
                query = query.where(Pedido.cerrado_en >= desde_utc)
            if hasta_utc is not None:
                query = query.where(Pedido.cerrado_en <= hasta_utc)
            if self.historial_filtro_metodo:
                from app.models.food import PagoPedido as _PP
                query = query.where(
                    or_(
                        Pedido.metodo_pago == self.historial_filtro_metodo,
                        Pedido.id.in_(
                            select(_PP.pedido_id).where(
                                _PP.company_id == food._company_id(),
                                _PP.metodo == self.historial_filtro_metodo,
                            )
                        ),
                    )
                )
            query = query.order_by(Pedido.cerrado_en.desc(), Pedido.id.desc())
            pedidos = session.exec(query).all()
            mesas = {m.id: m for m in session.exec(
                select(Mesa).where(Mesa.company_id == food._company_id())
            ).all()}
            usuarios = {u.id: u for u in session.exec(
                select(UsuarioFood).where(UsuarioFood.company_id == food._company_id())
            ).all()}
            pedido_ids = [p.id for p in pedidos if p.id is not None]
            detalles_por_pedido: dict[int, list] = {}
            productos_map: dict[int, Producto] = {}
            # Pagos persistidos por pedido, para el desglose por método (montos
            # netos de vuelto). Se clasifican por tipo configurado de la empresa.
            tipos_metodo = mapa_tipos(session, food._company_id())
            nombres_metodo = mapa_nombres(session, food._company_id())
            pagos_por_pedido: dict[int, list] = {}
            if pedido_ids:
                detalles = session.exec(
                    select(DetallePedido).where(DetallePedido.pedido_id.in_(pedido_ids))
                ).all()
                for d in detalles:
                    detalles_por_pedido.setdefault(d.pedido_id, []).append(d)
                productos_map = {
                    pr.id: pr for pr in session.exec(
                        select(Producto).where(Producto.company_id == food._company_id())
                    ).all()
                }
                for pg in session.exec(
                    select(PagoPedido).where(PagoPedido.pedido_id.in_(pedido_ids))
                ).all():
                    pagos_por_pedido.setdefault(pg.pedido_id, []).append(pg)

        if not pedidos:
            return rx.toast.error("No hay ventas para exportar con estos filtros.")

        empresa = getattr(food, "empresa_nombre", "TUWAYKIFOOD") or "TUWAYKIFOOD"
        generado = format_local_datetime(_utcnow(), "%d/%m/%Y %H:%M", self.reportes_pais)
        meta = [("Empresa", empresa), ("Generado", generado),
                ("Rango de fechas", self._rango_export_texto())]

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Ventas"
        # Cada pedido se descompone en TODOS sus conceptos para que el reporte
        # cuadre con el cierre: subtotal de ítems + recargo (envases/delivery)
        # − descuento = venta neta; + propina = total cobrado (≡ Σ pagos). El
        # desglose por método usa los montos netos de vuelto realmente cobrados.
        cols1 = [
            "Fecha", "Hora", "Pedido #", "Tipo", "Mesa / Cliente",
            "Subtotal ítems", "Recargo", "Concepto recargo", "Descuento",
            "Venta neta", "Propina", "Total cobrado",
            "Efectivo", "Tarjeta", "QR/Yape", "Fiado",
            "Métodos de pago (detalle)",
            "Mozo", "Cajero",
        ]
        hdr1 = _excel_titulo(
            ws1,
            "Reporte de ventas — una fila por venta cobrada",
            "Cada fila es un pedido cobrado con su desglose: subtotal de ítems + "
            "recargo − descuento = venta neta; + propina = total cobrado "
            "(= suma de los métodos de pago). «Efectivo/Tarjeta/QR/Fiado» agrupan "
            "por tipo; «Métodos de pago (detalle)» lista el método exacto y su "
            "importe (Yape, Plin, etc.). El detalle de QUÉ productos se vendieron "
            "en cada pedido está en la hoja «Detalle de items».",
            meta, n_cols=len(cols1),
        )
        ws1.append(cols1)
        for c in ws1[hdr1]:
            c.font = Font(bold=True)
        tot = {k: Decimal("0.00") for k in (
            "subtotal", "recargo", "descuento", "neta", "propina", "cobrado",
            "efectivo", "tarjeta", "qr", "fiado",
        )}
        for p in pedidos:
            fecha = p.cerrado_en
            subtotal = _to_decimal(p.total)
            recargo = _to_decimal(getattr(p, "recargo", 0))
            descuento = _to_decimal(getattr(p, "descuento", 0))
            propina = _to_decimal(getattr(p, "propina", 0))
            venta_neta = subtotal + recargo - descuento
            # Desglose por método a partir de los pagos persistidos (netos de
            # vuelto). El "total cobrado" es lo realmente recibido: la suma de
            # las columnas por método, que es lo que cuadra con el cierre.
            buckets = {"efectivo": Decimal("0.00"), "tarjeta": Decimal("0.00"),
                       "qr": Decimal("0.00"), "fiado": Decimal("0.00")}
            # Detalle exacto por método (Yape, Plin, …) para no perder el "con qué
            # se pagó" al agrupar por tipo en las columnas de arriba.
            detalle_metodos: dict[str, Decimal] = {}
            pagos_ped = pagos_por_pedido.get(p.id or 0, [])
            if pagos_ped:
                for pg in pagos_ped:
                    cod = (pg.metodo or "efectivo").lower()
                    b = _bucket_tipo(tipos_metodo.get(cod, "digital"))
                    buckets[b] += _to_decimal(pg.monto)
                    nombre_m = nombres_metodo.get(cod, cod.title())
                    detalle_metodos[nombre_m] = detalle_metodos.get(
                        nombre_m, Decimal("0.00")) + _to_decimal(pg.monto)
                total_cobrado = sum(buckets.values(), Decimal("0.00"))
            else:
                # Pedido legacy sin filas de pago: el cobrado se asume la venta
                # neta + propina, cargada al método declarado.
                total_cobrado = venta_neta + propina
                cod = (getattr(p, "metodo_pago", None) or "efectivo").lower()
                buckets[_bucket_tipo(tipos_metodo.get(cod, "digital"))] += total_cobrado
                nombre_m = nombres_metodo.get(cod, cod.title())
                detalle_metodos[nombre_m] = detalle_metodos.get(
                    nombre_m, Decimal("0.00")) + total_cobrado
            metodos_texto = "; ".join(
                f"{n}: {self._simbolo_moneda()} {float(v):,.2f}"
                for n, v in detalle_metodos.items()
            )
            mozo = usuarios.get(p.mozo_id)
            cajero = usuarios.get(p.cajero_id)
            ws1.append([
                format_local_datetime(fecha, "%Y-%m-%d", self.reportes_pais) if fecha else "",
                format_local_datetime(fecha, "%H:%M", self.reportes_pais) if fecha else "",
                p.id,
                (getattr(p, "tipo_pedido", None) or "").capitalize(),
                _pedido_sales_label(p, mesas),
                float(subtotal), float(recargo),
                getattr(p, "recargo_concepto", None) or "",
                float(descuento), float(venta_neta), float(propina),
                float(total_cobrado),
                float(buckets["efectivo"]), float(buckets["tarjeta"]),
                float(buckets["qr"]), float(buckets["fiado"]),
                metodos_texto,
                _actor_name(mozo.nombre) if mozo else "Sin asignar",
                _actor_name(cajero.nombre) if cajero else "Sin asignar",
            ])
            tot["subtotal"] += subtotal
            tot["recargo"] += recargo
            tot["descuento"] += descuento
            tot["neta"] += venta_neta
            tot["propina"] += propina
            tot["cobrado"] += total_cobrado
            tot["efectivo"] += buckets["efectivo"]
            tot["tarjeta"] += buckets["tarjeta"]
            tot["qr"] += buckets["qr"]
            tot["fiado"] += buckets["fiado"]
        fila_tot = ws1.max_row + 1
        ws1.append([
            "", "", "", "", "TOTALES",
            float(tot["subtotal"]), float(tot["recargo"]), "",
            float(tot["descuento"]), float(tot["neta"]), float(tot["propina"]),
            float(tot["cobrado"]), float(tot["efectivo"]), float(tot["tarjeta"]),
            float(tot["qr"]), float(tot["fiado"]), "", "", "",
        ])
        for c in ws1[fila_tot]:
            c.font = Font(bold=True)

        ws2 = wb.create_sheet("Detalle de items")
        cols2 = ["Fecha", "Hora", "Pedido #", "Tipo", "Mesa / Cliente",
                 "Producto", "Cantidad", "Precio unitario", "Subtotal",
                 "Mozo", "Cajero", "Notas"]
        hdr2 = _excel_titulo(
            ws2,
            "Detalle de ítems vendidos — qué se vendió, cuánto y quién",
            "Una fila por producto dentro de cada venta cobrada: qué producto, "
            "cuántas unidades, a qué precio, el mozo/cajero que atendió y las "
            "notas del ítem. Se cruza con la hoja «Ventas» por «Pedido #».",
            meta, n_cols=len(cols2),
        )
        ws2.append(cols2)
        for c in ws2[hdr2]:
            c.font = Font(bold=True)
        tot_items_cant = 0
        tot_items_imp = Decimal("0.00")
        for p in pedidos:
            fecha = p.cerrado_en
            mesa_label = _pedido_sales_label(p, mesas)
            tipo = (getattr(p, "tipo_pedido", None) or "").capitalize()
            mozo = usuarios.get(p.mozo_id)
            cajero = usuarios.get(p.cajero_id)
            for d in detalles_por_pedido.get(p.id or 0, []):
                prod = productos_map.get(d.producto_id)
                tot_items_cant += int(d.cantidad or 0)
                tot_items_imp += _to_decimal(d.subtotal)
                ws2.append([
                    format_local_datetime(fecha, "%Y-%m-%d", self.reportes_pais) if fecha else "",
                    format_local_datetime(fecha, "%H:%M", self.reportes_pais) if fecha else "",
                    p.id,
                    tipo,
                    mesa_label,
                    prod.nombre if prod else f"Producto {d.producto_id}",
                    d.cantidad,
                    float(_to_decimal(d.precio_unitario)),
                    float(_to_decimal(d.subtotal)),
                    _actor_name(mozo.nombre) if mozo else "Sin asignar",
                    _actor_name(cajero.nombre) if cajero else "Sin asignar",
                    getattr(d, "notas", None) or "",
                ])
        fila_tot2 = ws2.max_row + 1
        ws2.append(["", "", "", "", "", "TOTALES", tot_items_cant, "",
                    float(tot_items_imp), "", "", ""])
        for c in ws2[fila_tot2]:
            c.font = Font(bold=True)

        simbolo = self._simbolo_moneda()
        _excel_finalizar(ws1, hdr1, money_cols=[6, 7, 9, 10, 11, 12, 13, 14, 15, 16], simbolo=simbolo)
        _excel_finalizar(ws2, hdr2, money_cols=[8, 9], simbolo=simbolo)

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"ventas_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    async def exportar_pyl_excel(self):
        if not self.pyl_lineas:
            return rx.toast.error("Sin datos de P&L para exportar.")
        from openpyxl import Workbook
        from openpyxl.styles import Font

        meta = await self._excel_meta([("Período", f"{self.pyl_mes:02d}/{self.pyl_anio}")])
        wb = Workbook()
        ws = wb.active
        ws.title = "P&L"
        cols = ["Concepto", "Valor", "Margen %"]
        hdr = _excel_titulo(
            ws, "Estado de resultados (P&L) mensual",
            "Ingresos, costos y utilidad del mes indicado. 'Margen %' es sobre las "
            "ventas. Los valores salen ya formateados como texto.",
            meta, n_cols=len(cols))
        ws.append(cols)
        for c in ws[hdr]:
            c.font = Font(bold=True)
        for ln in self.pyl_lineas:
            ws.append([ln.concepto, ln.valor_texto, ln.margen_pct_texto or ""])
        _excel_finalizar(ws, hdr)
        buf = io.BytesIO()
        wb.save(buf)
        filename = f"pyl_{self.pyl_anio}_{self.pyl_mes:02d}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    async def exportar_igv_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        meta = await self._excel_meta([("Período", f"{self.igv_mes:02d}/{self.igv_anio}")])
        wb = Workbook()
        ws = wb.active
        ws.title = "IGV"
        cols = ["Concepto", "Valor"]
        hdr = _excel_titulo(
            ws, "Resumen de IGV mensual",
            "Ventas netas, base imponible e IGV del mes, para la declaración de "
            "impuestos.",
            meta, n_cols=len(cols))
        ws.append(cols)
        for c in ws[hdr]:
            c.font = Font(bold=True)
        ws.append(["Ventas netas", self.igv_ventas_netas_texto])
        ws.append(["Base imponible", self.igv_base_imponible_texto])
        ws.append([f"IGV ({self.igv_porcentaje}%)", self.igv_monto_texto])
        ws.append(["Pedidos", self.igv_pedidos])
        _excel_finalizar(ws, hdr)
        buf = io.BytesIO()
        wb.save(buf)
        filename = f"igv_{self.igv_anio}_{self.igv_mes:02d}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    async def exportar_descuentos_excel(self):
        if not self.descuentos_rank and not self.anulaciones_lista:
            return rx.toast.error("Sin descuentos ni anulaciones para exportar.")
        from openpyxl import Workbook
        from openpyxl.styles import Font

        meta = await self._excel_meta([("Rango de fechas", self._rango_export_texto())])
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Descuentos"
        cols1 = ["Cajero", "Pedidos", "Total descuento", "Total ventas", "% Descuento"]
        h1 = _excel_titulo(
            ws1, "Descuentos por cajero",
            "Cuánto descontó cada cajero en el rango: cantidad de pedidos, monto "
            "descontado y su peso sobre las ventas.",
            meta, n_cols=len(cols1))
        ws1.append(cols1)
        for c in ws1[h1]:
            c.font = Font(bold=True)
        for d in self.descuentos_rank:
            ws1.append([d.cajero, d.pedidos, d.total_descuento_texto,
                        d.total_ventas_texto, d.pct_descuento_texto])

        ws2 = wb.create_sheet("Anulaciones")
        cols2 = ["Pedido #", "Total", "Motivo", "Anulado por", "Fecha", "Cajero original"]
        h2 = _excel_titulo(
            ws2, "Anulaciones de pedidos",
            "Cada pedido anulado en el rango, con su motivo, quién lo anuló y el "
            "cajero que lo había cobrado.",
            meta, n_cols=len(cols2))
        ws2.append(cols2)
        for c in ws2[h2]:
            c.font = Font(bold=True)
        for a in self.anulaciones_lista:
            ws2.append([a.pedido_id, a.total_texto, a.motivo,
                        a.cancelado_por, a.cancelado_en_texto, a.cajero_original])

        _excel_finalizar(ws1, h1)
        _excel_finalizar(ws2, h2)
        buf = io.BytesIO()
        wb.save(buf)
        filename = f"descuentos_anulaciones_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    async def exportar_mermas_excel(self):
        if not self.mermas_por_categoria and not self.mermas_por_insumo:
            return rx.toast.error("Sin mermas para exportar.")
        from openpyxl import Workbook
        from openpyxl.styles import Font

        meta = await self._excel_meta([("Rango de fechas", self._rango_export_texto())])
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Mermas por categoría"
        cols1 = ["Categoría", "Registros", "Valor"]
        h1 = _excel_titulo(
            ws1, "Mermas por categoría",
            "Pérdidas de insumos (mermas) agrupadas por categoría en el rango: "
            "cantidad de registros y valor perdido.",
            meta, n_cols=len(cols1))
        ws1.append(cols1)
        for c in ws1[h1]:
            c.font = Font(bold=True)
        for c in self.mermas_por_categoria:
            ws1.append([c.categoria, c.registros, c.valor_texto])

        ws2 = wb.create_sheet("Mermas por insumo")
        cols2 = ["Insumo", "Unidad", "Cantidad", "Valor", "Registros"]
        h2 = _excel_titulo(
            ws2, "Mermas por insumo",
            "El detalle de cada insumo mermado en el rango: cuánto se perdió, su "
            "valor y cuántas veces se registró.",
            meta, n_cols=len(cols2))
        ws2.append(cols2)
        for c in ws2[h2]:
            c.font = Font(bold=True)
        for i in self.mermas_por_insumo:
            ws2.append([i.nombre, i.unidad, i.cantidad_texto, i.valor_texto, i.registros])

        _excel_finalizar(ws1, h1)
        _excel_finalizar(ws2, h2)
        buf = io.BytesIO()
        wb.save(buf)
        filename = f"mermas_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    async def exportar_matriz_excel(self):
        if not self.matriz_productos:
            return rx.toast.error("Sin datos de matriz para exportar.")
        from openpyxl import Workbook
        from openpyxl.styles import Font

        meta = await self._excel_meta([("Rango de fechas", self._rango_export_texto())])
        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz productos"
        cols = ["Producto", "Unidades", "Ingreso", "Margen %", "Categoría"]
        hdr = _excel_titulo(
            ws, "Matriz de productos (estrella / perro)",
            "Cada producto por unidades vendidas e ingreso, con su margen % (precio "
            "vs costo de receta). La 'Categoría' es el cuadrante BCG: estrella, vaca, "
            "puzzle o perro. Ingreso y Margen % van como número para poder ordenar "
            "y sumar.",
            meta, n_cols=len(cols))
        ws.append(cols)
        for c in ws[hdr]:
            c.font = Font(bold=True)
        for p in self.matriz_productos:
            ws.append([p.nombre, p.unidades, round(p.ingreso, 2),
                       round(p.margen_pct, 1), p.categoria.capitalize()])
        for row in ws.iter_rows(min_row=hdr + 1, min_col=4, max_col=4):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0"%"'
        _excel_finalizar(ws, hdr, money_cols=[3], simbolo=self._simbolo_moneda())
        buf = io.BytesIO()
        wb.save(buf)
        filename = f"matriz_productos_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    async def exportar_margen_excel(self):
        if not self.reporte_margen:
            return rx.toast.error("Sin datos de margen para exportar.")
        from openpyxl import Workbook
        from openpyxl.styles import Font

        meta = await self._excel_meta()
        wb = Workbook()
        ws = wb.active
        ws.title = "Margen por plato"
        cols = ["Producto", "Precio", "Costo", "Margen", "Margen %", "Costo completo"]
        hdr = _excel_titulo(
            ws, "Margen por plato",
            "Para cada plato con receta: precio de venta, costo de insumos, margen "
            "en dinero y en %. 'Costo completo' = No cuando falta cargar algún "
            "insumo de la receta (el margen mostrado es parcial). Requiere tener "
            "cargadas las recetas en Inventario.",
            meta, n_cols=len(cols))
        ws.append(cols)
        for c in ws[hdr]:
            c.font = Font(bold=True)
        for m in self.reporte_margen:
            ws.append([m.nombre, round(m.precio, 2), round(m.costo, 2),
                       round(m.margen, 2), round(m.margen_pct, 1),
                       "Sí" if m.costo_completo else "No"])
        for row in ws.iter_rows(min_row=hdr + 1, min_col=5, max_col=5):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0"%"'
        _excel_finalizar(ws, hdr, money_cols=[2, 3, 4], simbolo=self._simbolo_moneda())
        buf = io.BytesIO()
        wb.save(buf)
        filename = f"margen_platos_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    async def exportar_productos_stock_excel(self):
        """Excel de control de productos de la carta: stock actual + unidades
        vendidas + ingreso, por producto. Disponible en Standard.

        Respeta los filtros de fecha del Historial (Desde/Hasta). Sin filtro =
        todo el histórico. Los montos van como número para poder sumarlos en Excel.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font

        food = await self._food()
        self.reportes_pais = food._country_code()
        desde, hasta = self._rango_filtros_historial()
        if hasta is not None:
            hasta = hasta + timedelta(seconds=1)
        with food._tenant_session() as session:
            filas = reporte_productos_stock(session, food._company_id(), desde, hasta)
        if not filas:
            return rx.toast.error("No hay productos para exportar.")

        empresa = getattr(food, "empresa_nombre", "TUWAYKIFOOD") or "TUWAYKIFOOD"
        generado = format_local_datetime(_utcnow(), "%d/%m/%Y %H:%M", self.reportes_pais)
        simbolo = "S/" if self.reportes_pais == "PE" else "$"

        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        encabezados = ["Categoría", "Producto", "Stock actual",
                       "Unidades vendidas", "Ingreso"]
        hdr = _excel_titulo(
            ws,
            "Productos — vendidos y stock",
            "Una fila por producto de la carta. 'Stock actual' es lo que queda "
            "(o 'Sin control' si el producto no lleva stock). 'Unidades vendidas' "
            "e 'Ingreso' son del rango indicado abajo. Ordenado de más a menos vendido.",
            [("Empresa", empresa), ("Generado", generado),
             ("Rango de fechas", self._rango_export_texto())],
            n_cols=len(encabezados),
        )
        ws.append(encabezados)
        for c in ws[hdr]:
            c.font = Font(bold=True)

        total_unidades = 0
        total_ingreso = Decimal("0.00")
        for f in filas:
            stock = f["stock"]
            stock_cell = "Sin control" if stock is None else int(stock)
            unidades = int(f["unidades"] or 0)
            ingreso = f["ingreso"] or Decimal("0.00")
            total_unidades += unidades
            total_ingreso += ingreso
            ws.append([f["categoria"], f["nombre"], stock_cell,
                       unidades, float(ingreso)])

        # Fila de totales
        fila_total = ws.max_row + 1
        ws.cell(row=fila_total, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=fila_total, column=4, value=total_unidades).font = Font(bold=True)
        ws.cell(row=fila_total, column=5, value=float(total_ingreso)).font = Font(bold=True)

        _excel_finalizar(ws, hdr, money_cols=[5], simbolo=simbolo)

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"productos_vendidos_stock_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    # ── P&L mensual (ADM-01) ────────────────────────────────────────────────

    async def cargar_pyl(self) -> None:
        food = await self._food()
        self._do_cargar_pyl(food)

    def _do_cargar_pyl(self, food) -> None:
        now = _utcnow()
        anio = self.pyl_anio or now.year
        mes = self.pyl_mes or now.month
        self.pyl_anio = anio
        self.pyl_mes = mes

        with food._tenant_session() as session:
            data = pyl_mensual(session, food._company_id(), anio, mes, food._country_code())

        meses_nombre = [
            "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]
        self.pyl_lineas = [
            PylLineView(
                concepto="Ventas brutas",
                valor_texto=_money_text(data["ventas_brutas"]),
            ),
            PylLineView(
                concepto="Descuentos otorgados",
                valor_texto=f"- {_money_text(data['descuentos'])}",
                es_negativo=True,
            ),
            PylLineView(
                concepto="Ventas netas",
                valor_texto=_money_text(data["ventas_netas"]),
                es_total=True,
            ),
            PylLineView(
                concepto="Costo de insumos consumidos",
                valor_texto=f"- {_money_text(data['costo_consumo'])}",
                es_negativo=True,
            ),
            PylLineView(
                concepto="Egresos de caja",
                valor_texto=f"- {_money_text(data['egresos_caja'])}",
                es_negativo=True,
            ),
            PylLineView(
                concepto="Utilidad bruta operativa",
                valor_texto=_money_text(data["utilidad"]),
                es_total=True,
                es_negativo=data["utilidad"] < 0,
                margen_pct_texto=f"{data['margen_pct']}%",
            ),
        ]

    def set_pyl_anio(self, v: str) -> None:
        try:
            self.pyl_anio = int(v)
        except (ValueError, TypeError):
            pass

    def set_pyl_mes(self, v: str) -> None:
        try:
            self.pyl_mes = int(v)
        except (ValueError, TypeError):
            pass

    async def actualizar_pyl(self) -> None:
        await self.cargar_pyl()

    # ── IGV mensual (ADM-05) ────────────────────────────────────────────────

    async def cargar_igv(self) -> None:
        food = await self._food()
        self._do_cargar_igv(food)

    def _do_cargar_igv(self, food) -> None:
        now = _utcnow()
        anio = self.igv_anio or now.year
        mes = self.igv_mes or now.month
        self.igv_anio = anio
        self.igv_mes = mes
        try:
            pct = float(food.config_porcentaje_iva or "18.0")
        except (ValueError, AttributeError):
            pct = 18.0
        with food._tenant_session() as session:
            data = resumen_igv_mensual(
                session, food._company_id(), anio, mes, pct, food._country_code()
            )
        self.igv_base_imponible_texto = _money_text(data["base_imponible"])
        self.igv_monto_texto = _money_text(data["igv"])
        self.igv_ventas_netas_texto = _money_text(data["ventas_netas"])
        self.igv_porcentaje = data["porcentaje_igv"]
        self.igv_pedidos = data["pedidos"]

    def set_igv_anio(self, v: str) -> None:
        try:
            self.igv_anio = int(v)
        except (ValueError, TypeError):
            pass

    def set_igv_mes(self, v: str) -> None:
        try:
            self.igv_mes = int(v)
        except (ValueError, TypeError):
            pass

    async def actualizar_igv(self) -> None:
        await self.cargar_igv()

    # ── Matriz estrella/perro (ADM-06) ─────────────────────────────────────

    async def cargar_matriz_productos(self) -> None:
        food = await self._food()
        self._do_cargar_matriz(food)

    def _do_cargar_matriz(self, food) -> None:
        self.reportes_pais = food._country_code()
        desde, hasta = self._rango_filtros_historial()
        if hasta is not None:
            hasta = hasta + timedelta(seconds=1)
        _CAT_EMOJI = {
            "estrella": "⭐",
            "vaca": "\U0001F42E",
            "puzzle": "\U0001F9E9",
            "perro": "\U0001F415",
        }
        with food._tenant_session() as session:
            data = matriz_estrella_perro(session, food._company_id(), desde, hasta)
        vistas = []
        counts = {"estrella": 0, "vaca": 0, "puzzle": 0, "perro": 0}
        for f in data:
            cat = f["categoria"]
            counts[cat] = counts.get(cat, 0) + 1
            vistas.append(MatrizProductoView(
                nombre=f["nombre"],
                unidades=f["unidades"],
                ingreso_texto=_money_text(f["ingreso"]),
                margen_pct_texto=f"{f['margen_pct']:.1f}%",
                categoria=cat,
                categoria_emoji=_CAT_EMOJI.get(cat, ""),
                ingreso=float(f["ingreso"]),
                margen_pct=float(f["margen_pct"]),
            ))
        self.matriz_productos = vistas
        self.matriz_estrellas = counts["estrella"]
        self.matriz_vacas = counts["vaca"]
        self.matriz_puzzles = counts["puzzle"]
        self.matriz_perros = counts["perro"]

    # ── Descuentos y anulaciones (ADM-02) ────────────────────────────────────

    async def cargar_descuentos_anulaciones(self) -> None:
        food = await self._food()
        self._do_cargar_descuentos_anulaciones(food)

    def _do_cargar_descuentos_anulaciones(self, food) -> None:
        self.reportes_pais = food._country_code()
        desde, hasta = self._rango_filtros_historial()
        if hasta is not None:
            hasta = hasta + timedelta(seconds=1)

        with food._tenant_session() as session:
            desc_data = reporte_descuentos(session, food._company_id(), desde, hasta)
            anul_data = reporte_anulaciones(session, food._company_id(), desde, hasta)
            rev_data = reporte_reversiones(session, food._company_id(), desde, hasta)

        self.descuentos_rank = [
            DescuentoRankView(
                cajero=d["cajero"],
                pedidos=d["pedidos"],
                total_descuento_texto=_money_text(d["total_descuento"]),
                total_ventas_texto=_money_text(d["total_ventas"]),
                pct_descuento_texto=f"{d['pct_descuento']}%",
            )
            for d in desc_data
        ]
        total_desc = sum(d["total_descuento"] for d in desc_data)
        self.descuentos_total_texto = _money_text(total_desc)

        from tuwayki_core.utils.timezone import format_local_datetime
        self.anulaciones_lista = [
            AnulacionView(
                pedido_id=a["pedido_id"],
                total_texto=_money_text(a["total"]),
                motivo=a["motivo"],
                cancelado_por=a["cancelado_por"],
                cancelado_en_texto=format_local_datetime(a["cancelado_en"], "%d/%m %H:%M", self.reportes_pais) if a["cancelado_en"] else "",
                cajero_original=a["cajero_original"],
            )
            for a in anul_data
        ]
        total_anul = sum(a["total"] for a in anul_data)
        self.anulaciones_total_texto = _money_text(total_anul)

        self.reversiones_lista = [
            ReversionView(
                pedido_id=r["pedido_id"],
                total_texto=_money_text(r["total"]),
                motivo=r["motivo"],
                revertido_por=r["revertido_por"],
                revertido_en_texto=format_local_datetime(r["revertido_en"], "%d/%m %H:%M", self.reportes_pais) if r["revertido_en"] else "",
            )
            for r in rev_data
        ]
        total_rev = sum(r["total"] for r in rev_data)
        self.reversiones_total_texto = _money_text(total_rev)

    # ── Mermas (ADM-03) ─────────────────────────────────────────────────────

    async def cargar_mermas(self) -> None:
        food = await self._food()
        self._do_cargar_mermas(food)

    def _do_cargar_mermas(self, food) -> None:
        self.reportes_pais = food._country_code()
        desde, hasta = self._rango_filtros_historial()
        if hasta is not None:
            hasta = hasta + timedelta(seconds=1)

        with food._tenant_session() as session:
            data = reporte_mermas(session, food._company_id(), desde, hasta)

        self.mermas_por_categoria = [
            MermaCategoriaView(
                categoria=c["categoria"],
                registros=c["registros"],
                valor_texto=_money_text(c["valor"]),
            )
            for c in data["por_categoria"]
        ]
        self.mermas_por_insumo = [
            MermaInsumoView(
                nombre=i["nombre"],
                unidad=i["unidad"],
                cantidad_texto=f"{i['cantidad_total']:.3f}",
                valor_texto=_money_text(i["valor"]),
                registros=i["registros"],
            )
            for i in data["por_insumo"][:20]
        ]
        self.mermas_total_texto = _money_text(data["total"])

    # ── PDF ejecutivo ────────────────────────────────────────────────────────

    async def exportar_pdf_ejecutivo(self):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        food = await self._food()
        self.reportes_pais = food._country_code()
        empresa_nombre = getattr(food, "empresa_nombre", "TUWAYKIFOOD")
        fecha_str = format_local_datetime(_utcnow(), "%d/%m/%Y %H:%M", self.reportes_pais)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm,
            topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            "TitlePDF", parent=styles["Title"],
            fontSize=18, spaceAfter=4 * mm,
        ))
        styles.add(ParagraphStyle(
            "SubtitlePDF", parent=styles["Normal"],
            fontSize=10, textColor=colors.grey, spaceAfter=8 * mm,
        ))
        styles.add(ParagraphStyle(
            "SectionHeader", parent=styles["Heading2"],
            fontSize=13, spaceBefore=6 * mm, spaceAfter=3 * mm,
            textColor=colors.HexColor("#1E293B"),
        ))

        elements = []

        elements.append(Paragraph(f"Reporte Ejecutivo — {empresa_nombre}", styles["TitlePDF"]))
        elements.append(Paragraph(f"Generado: {fecha_str}", styles["SubtitlePDF"]))

        # KPIs
        elements.append(Paragraph("Resumen del Día", styles["SectionHeader"]))
        kpi_data = [
            ["Ventas", "Pedidos", "Ticket Prom.", "Propinas"],
            [
                self.dashboard_ventas_hoy_texto,
                str(self.dashboard_pedidos_hoy),
                self.dashboard_ticket_promedio_texto,
                self.dashboard_propina_hoy_texto,
            ],
        ]
        kpi_table = Table(kpi_data, colWidths=[4.2 * cm] * 4)
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, 1), 12),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 4 * mm))

        # Ocupación
        ocup_data = [
            ["Mesas Ocupadas", "Total Mesas", "Ítems en Cocina", "Reservas Hoy"],
            [
                str(self.dashboard_mesas_ocupadas),
                str(self.dashboard_total_mesas),
                str(self.dashboard_items_en_cocina),
                str(self.dashboard_reservas_hoy),
            ],
        ]
        ocup_table = Table(ocup_data, colWidths=[4.2 * cm] * 4)
        ocup_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, 1), 12),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(ocup_table)
        elements.append(Spacer(1, 6 * mm))

        # Top platos
        if self.dashboard_top_platos:
            elements.append(Paragraph("Top Platos del Día", styles["SectionHeader"]))
            top_data = [["#", "Producto", "Unidades", "Total"]]
            for i, p in enumerate(self.dashboard_top_platos[:10], 1):
                top_data.append([str(i), p.nombre, str(p.cantidad), p.total_texto])
            top_table = Table(top_data, colWidths=[1 * cm, 8 * cm, 3 * cm, 4 * cm])
            top_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (2, 0), (3, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(top_table)
            elements.append(Spacer(1, 6 * mm))

        # Mozos ranking
        if self.reporte_mozos:
            elements.append(Paragraph("Ranking Mozos", styles["SectionHeader"]))
            mozo_data = [["Mozo", "Pedidos", "Total Vendido"]]
            for m in self.reporte_mozos[:10]:
                mozo_data.append([m.nombre, str(m.pedidos), m.total_texto])
            mozo_table = Table(mozo_data, colWidths=[7 * cm, 3 * cm, 5 * cm])
            mozo_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (2, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(mozo_table)
            elements.append(Spacer(1, 6 * mm))

        # Tendencias
        elements.append(Paragraph("Tendencias vs. Ayer", styles["SectionHeader"]))
        trend_data = [
            ["Métrica", "Variación"],
            ["Ventas", f"{'+' if self.dashboard_ventas_trend_pct >= 0 else ''}{self.dashboard_ventas_trend_pct}%"],
            ["Pedidos", f"{'+' if self.dashboard_pedidos_trend >= 0 else ''}{self.dashboard_pedidos_trend}"],
            ["Ticket Prom.", f"{'+' if self.dashboard_ticket_trend_pct >= 0 else ''}{self.dashboard_ticket_trend_pct}%"],
            ["Propinas", f"{'+' if self.dashboard_propina_trend_pct >= 0 else ''}{self.dashboard_propina_trend_pct}%"],
        ]
        trend_table = Table(trend_data, colWidths=[8 * cm, 5 * cm])
        trend_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(trend_table)

        doc.build(elements)
        filename = f"reporte_ejecutivo_{_utcnow().strftime('%Y%m%d_%H%M')}.pdf"
        return rx.download(data=buf.getvalue(), filename=filename)
