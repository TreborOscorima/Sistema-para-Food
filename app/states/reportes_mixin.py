"""Reportes y analítica — dashboard KPIs, historial de ventas y exportación.

Mixin extraído de FoodState. En runtime, FoodState hereda este mixin y provee
`_tenant_session()`, `_company_id()`, `mensaje`, `config_nombre_local`,
`_ticket_paper_width_mm()`, `usuario_actual` y los helpers del módulo
(`_utcnow`, `_to_decimal`, `_money_text`, `_actor_name`, `_pedido_sales_label`).
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta
from decimal import Decimal

import reflex as rx
from sqlalchemy import and_, or_
from sqlmodel import select

from app.models.food import (
    DetallePedido,
    EstadoMesa,
    EstadoPedido,
    Mesa,
    PagoPedido,
    Pedido,
    Producto,
    UsuarioFood,
)
from app.services.analitica_service import (
    margen_por_plato,
    ventas_por_hora,
    ventas_por_mozo,
)
from app.services.finanzas_service import (
    pyl_mensual,
    reporte_anulaciones,
    reporte_descuentos,
    reporte_mermas,
)
from app.states.food_state import (
    AnulacionView,
    DescuentoRankView,
    FranjaHoraView,
    MargenPlatoView,
    MermaCategoriaView,
    MermaInsumoView,
    MozoRankView,
    PylLineView,
    TopPlatoView,
    VentaDetalleItemView,
    VentaHistorialView,
    _actor_name,
    _money_text,
    _pedido_sales_label,
    _to_decimal,
    _utcnow,
)


class ReportesMixin(rx.State, mixin=True):
    """Dashboard KPIs, historial de ventas, analítica y exportación Excel."""

    # ── State vars — Dashboard KPIs ──────────────────────────────────────────
    dashboard_ventas_hoy_texto: str = "S/ 0.00"
    dashboard_pedidos_hoy: int = 0
    dashboard_mesas_ocupadas: int = 0
    dashboard_propina_hoy_texto: str = "S/ 0.00"
    dashboard_ticket_promedio_texto: str = "S/ 0.00"
    dashboard_top_platos: list[TopPlatoView] = []
    dashboard_ventas_trend_pct: int = 0
    dashboard_pedidos_trend: int = 0
    dashboard_ticket_trend_pct: int = 0
    dashboard_propina_trend_pct: int = 0

    # ── State vars — Analítica de reportes ───────────────────────────────────
    reporte_mozos: list[MozoRankView] = []
    reporte_horas: list[FranjaHoraView] = []
    reporte_margen: list[MargenPlatoView] = []
    reporte_metodos: list[dict[str, object]] = []

    # ── State vars — Comparativa entre períodos ──────────────────────────────
    comp_label_actual: str = "Período actual"
    comp_label_anterior: str = "Período anterior"
    comp_ventas_actual: str = "S/ 0.00"
    comp_ventas_anterior: str = "S/ 0.00"
    comp_ventas_pct: int = 0
    comp_pedidos_actual: int = 0
    comp_pedidos_anterior: int = 0
    comp_pedidos_diff: int = 0
    comp_ticket_actual: str = "S/ 0.00"
    comp_ticket_anterior: str = "S/ 0.00"
    comp_ticket_pct: int = 0

    # ── State vars — Historial filtros ───────────────────────────────────────
    historial_filtro_fecha_desde: str = ""
    historial_filtro_fecha_hasta: str = ""
    historial_filtro_metodo: str = ""
    historial_filtro_rapido: str = "hoy"
    historial_pagina: int = 0
    historial_total: int = 0
    _HISTORIAL_PAGE_SIZE: int = 50

    # ── State vars — P&L mensual (ADM-01) ──────────────────────────────────
    pyl_lineas: list[PylLineView] = []
    pyl_anio: int = 0
    pyl_mes: int = 0

    # ── State vars — Descuentos y anulaciones (ADM-02) ───────────────────
    descuentos_rank: list[DescuentoRankView] = []
    anulaciones_lista: list[AnulacionView] = []
    anulaciones_total_texto: str = "S/ 0.00"
    descuentos_total_texto: str = "S/ 0.00"

    # ── State vars — Mermas valorizado (ADM-03) ─────────────────────────
    mermas_por_categoria: list[MermaCategoriaView] = []
    mermas_por_insumo: list[MermaInsumoView] = []
    mermas_total_texto: str = "S/ 0.00"

    # ── State vars — Detalle de venta (modal) ────────────────────────────────
    venta_detalle_visible: bool = False
    venta_detalle_pedido_id: int = 0
    venta_detalle_mesa_label: str = ""
    venta_detalle_metodo: str = ""
    venta_detalle_mozo: str = ""
    venta_detalle_cajero: str = ""
    venta_detalle_total_texto: str = ""
    venta_detalle_propina_texto: str = ""
    venta_detalle_items: list[VentaDetalleItemView] = []

    # ── Computed vars ────────────────────────────────────────────────────────

    @rx.var
    def historial_ventas_recientes(self) -> list[VentaHistorialView]:
        return self.historial_ventas[:5]

    @rx.var
    def reporte_horas_chart(self) -> list[dict[str, object]]:
        return [{"hora_label": h.hora_label, "total": h.total, "pedidos": h.pedidos} for h in self.reporte_horas]

    @rx.var
    def reporte_mozos_chart(self) -> list[dict[str, object]]:
        return [{"nombre": m.nombre, "total": m.total, "propinas": m.propinas} for m in self.reporte_mozos]

    @rx.var
    def historial_filtro_activo(self) -> bool:
        return bool(
            self.historial_filtro_fecha_desde
            or self.historial_filtro_fecha_hasta
            or self.historial_filtro_metodo
        )

    @rx.var
    def historial_tiene_anterior(self) -> bool:
        return self.historial_pagina > 0

    @rx.var
    def historial_tiene_siguiente(self) -> bool:
        return (self.historial_pagina + 1) * self._HISTORIAL_PAGE_SIZE < self.historial_total

    @rx.var
    def historial_pagina_label(self) -> str:
        if self.historial_total == 0:
            return "Sin resultados"
        desde = self.historial_pagina * self._HISTORIAL_PAGE_SIZE + 1
        hasta = min((self.historial_pagina + 1) * self._HISTORIAL_PAGE_SIZE, self.historial_total)
        return f"{desde}–{hasta} de {self.historial_total}"

    # ─── Dashboard KPIs ───────────────────────────────────────────────────────

    def cargar_dashboard(self) -> None:
        hoy = _utcnow().date()
        inicio_hoy = datetime(hoy.year, hoy.month, hoy.day)
        fin_hoy = inicio_hoy + timedelta(days=1)
        inicio_ayer = inicio_hoy - timedelta(days=1)
        with self._tenant_session() as session:
            pedidos_hoy = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                    Pedido.cerrado_en >= inicio_hoy,
                    Pedido.cerrado_en < fin_hoy,
                )
            ).all()
            pedidos_ayer = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                    Pedido.cerrado_en >= inicio_ayer,
                    Pedido.cerrado_en < inicio_hoy,
                )
            ).all()
            ventas = sum(
                _to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0))
                for p in pedidos_hoy
            )
            propina_total = sum(_to_decimal(getattr(p, "propina", 0)) for p in pedidos_hoy)
            ventas_ayer = sum(
                _to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0))
                for p in pedidos_ayer
            )
            propina_ayer = sum(_to_decimal(getattr(p, "propina", 0)) for p in pedidos_ayer)
            mesas_no_libres = session.exec(
                select(Mesa).where(
                    Mesa.company_id == self._company_id(),
                    Mesa.estado != EstadoMesa.LIBRE.value,
                    Mesa.activa.is_(True),
                )
            ).all()
            pedido_ids_hoy = {p.id for p in pedidos_hoy}
            top_data: dict[int, dict] = {}
            if pedido_ids_hoy:
                detalles = session.exec(
                    select(DetallePedido).where(DetallePedido.pedido_id.in_(pedido_ids_hoy))
                ).all()
                productos = {
                    p.id: p
                    for p in session.exec(select(Producto).where(Producto.company_id == self._company_id())).all()
                }
                for d in detalles:
                    pid = d.producto_id
                    if pid not in top_data:
                        prod = productos.get(pid)
                        top_data[pid] = {
                            "nombre": prod.nombre if prod else f"Producto {pid}",
                            "cantidad": 0,
                            "total": Decimal("0.00"),
                        }
                    top_data[pid]["cantidad"] += d.cantidad
                    top_data[pid]["total"] += _to_decimal(d.subtotal)
            top_sorted = sorted(top_data.values(), key=lambda x: x["cantidad"], reverse=True)[:5]

        def _pct_change(hoy_val: Decimal, ayer_val: Decimal) -> int:
            if ayer_val <= 0:
                return 100 if hoy_val > 0 else 0
            return int(round((hoy_val - ayer_val) / ayer_val * 100))

        pedidos_count_hoy = len(pedidos_hoy)
        pedidos_count_ayer = len(pedidos_ayer)
        ticket_promedio = ventas / pedidos_count_hoy if pedidos_count_hoy else Decimal("0.00")
        ticket_promedio_ayer = (
            ventas_ayer / pedidos_count_ayer if pedidos_count_ayer else Decimal("0.00")
        )
        self.dashboard_ventas_hoy_texto = _money_text(ventas)
        self.dashboard_pedidos_hoy = pedidos_count_hoy
        self.dashboard_mesas_ocupadas = len(mesas_no_libres)
        self.dashboard_propina_hoy_texto = _money_text(propina_total)
        self.dashboard_ticket_promedio_texto = _money_text(ticket_promedio)
        self.dashboard_ventas_trend_pct = _pct_change(ventas, ventas_ayer)
        self.dashboard_pedidos_trend = pedidos_count_hoy - pedidos_count_ayer
        self.dashboard_ticket_trend_pct = _pct_change(ticket_promedio, ticket_promedio_ayer)
        self.dashboard_propina_trend_pct = _pct_change(propina_total, propina_ayer)
        self.dashboard_top_platos = [
            TopPlatoView(
                nombre=p["nombre"],
                cantidad=p["cantidad"],
                total_generado=float(p["total"]),
                total_texto=_money_text(p["total"]),
            )
            for p in top_sorted
        ]

    # ─── Historial de ventas ──────────────────────────────────────────────────

    def cargar_historial_ventas(self) -> None:
        with self._tenant_session() as session:
            query = (
                select(Pedido)
                .where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                        # Ventas anuladas: visibles con badge, no desaparecen
                        and_(
                            Pedido.estado == EstadoPedido.CANCELADO.value,
                            Pedido.cerrado_en.isnot(None),
                        ),
                    ),
                )
            )
            if self.historial_filtro_fecha_desde:
                try:
                    desde = datetime.strptime(self.historial_filtro_fecha_desde, "%Y-%m-%d")
                    query = query.where(Pedido.cerrado_en >= desde)
                except ValueError:
                    pass
            if self.historial_filtro_fecha_hasta:
                try:
                    hasta = datetime.strptime(self.historial_filtro_fecha_hasta, "%Y-%m-%d")
                    hasta = hasta.replace(hour=23, minute=59, second=59)
                    query = query.where(Pedido.cerrado_en <= hasta)
                except ValueError:
                    pass
            if self.historial_filtro_metodo:
                from app.models.food import PagoPedido as _PP
                # Captura ventas con método exacto en Pedido + ventas mixtas que
                # incluyen ese método en alguno de sus PagoPedido.
                query = query.where(
                    or_(
                        Pedido.metodo_pago == self.historial_filtro_metodo,
                        Pedido.id.in_(
                            select(_PP.pedido_id).where(
                                _PP.company_id == self._company_id(),
                                _PP.metodo == self.historial_filtro_metodo,
                            )
                        ),
                    )
                )
            query = query.order_by(Pedido.cerrado_en.desc(), Pedido.id.desc())
            total_count = len(session.exec(query).all())
            self.historial_total = total_count
            offset = self.historial_pagina * self._HISTORIAL_PAGE_SIZE
            pedidos = session.exec(query.offset(offset).limit(self._HISTORIAL_PAGE_SIZE)).all()
            mesas = {m.id: m for m in session.exec(select(Mesa).where(Mesa.company_id == self._company_id())).all()}
            usuarios = {u.id: u for u in session.exec(select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())).all()}
            historial: list[VentaHistorialView] = []
            for p in pedidos:
                total_base = _to_decimal(p.total)
                propina = _to_decimal(getattr(p, "propina", Decimal("0.00")))
                total_con_propina = total_base + propina
                anulada = p.estado == EstadoPedido.CANCELADO.value
                anulacion_texto = ""
                if anulada:
                    quien = usuarios.get(p.cancelado_por_id)
                    anulacion_texto = (p.motivo_cancelacion or "Sin motivo") + (
                        f" — {quien.nombre}" if quien else ""
                    )
                historial.append(VentaHistorialView(
                    pedido_id=p.id or 0,
                    mesa_label=_pedido_sales_label(p, mesas),
                    total=float(total_base),
                    total_texto=_money_text(total_base),
                    propina=float(propina),
                    propina_texto=_money_text(propina) if propina > 0 else "",
                    total_con_propina=float(total_con_propina),
                    total_con_propina_texto=_money_text(total_con_propina),
                    metodo_pago=getattr(p, "metodo_pago", None) or "—",
                    mozo_nombre=_actor_name(usuarios[p.mozo_id].nombre if p.mozo_id in usuarios else "Sin asignar"),
                    cajero_nombre=_actor_name(usuarios[p.cajero_id].nombre if p.cajero_id in usuarios else "Sin asignar"),
                    anulada=anulada,
                    anulacion_texto=anulacion_texto,
                ))
            self.historial_ventas = historial

    def set_historial_filtro_fecha_desde(self, v: str) -> None:
        self.historial_filtro_fecha_desde = v

    def set_historial_filtro_fecha_hasta(self, v: str) -> None:
        self.historial_filtro_fecha_hasta = v

    def set_historial_filtro_metodo(self, v: str) -> None:
        self.historial_filtro_metodo = v

    def aplicar_filtros_historial(self) -> None:
        self.historial_pagina = 0
        self.cargar_historial_ventas()
        self.cargar_analitica()
        self.cargar_descuentos_anulaciones()
        self.cargar_mermas()

    def _rango_filtros_historial(self) -> tuple[datetime | None, datetime | None]:
        desde = hasta = None
        if self.historial_filtro_fecha_desde:
            try:
                desde = datetime.strptime(self.historial_filtro_fecha_desde, "%Y-%m-%d")
            except ValueError:
                pass
        if self.historial_filtro_fecha_hasta:
            try:
                hasta = datetime.strptime(self.historial_filtro_fecha_hasta, "%Y-%m-%d")
                hasta = hasta.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass
        return desde, hasta

    def cargar_analitica(self) -> None:
        """Ranking por mozo, ventas por hora, margen por plato y desglose por método."""
        desde, hasta = self._rango_filtros_historial()
        pagos_por_metodo: dict[str, dict[str, object]] = {}
        with self._tenant_session() as session:
            mozos = ventas_por_mozo(session, self._company_id(), desde, hasta)
            horas = ventas_por_hora(session, self._company_id(), desde, hasta)
            margenes = margen_por_plato(session, self._company_id())
            q_pagos = select(PagoPedido).where(PagoPedido.company_id == self._company_id())
            if desde is not None:
                q_pagos = q_pagos.where(PagoPedido.created_at >= desde)
            if hasta is not None:
                q_pagos = q_pagos.where(PagoPedido.created_at <= hasta)
            for pago in session.exec(q_pagos).all():
                m = pago.metodo or "otro"
                if m not in pagos_por_metodo:
                    pagos_por_metodo[m] = {"total": 0.0, "count": 0}
                pagos_por_metodo[m]["total"] = round(float(pagos_por_metodo[m]["total"]) + float(pago.monto), 2)
                pagos_por_metodo[m]["count"] = int(pagos_por_metodo[m]["count"]) + 1
        self.reporte_mozos = [
            MozoRankView(
                nombre=f["nombre"],
                pedidos=f["pedidos"],
                total=float(f["total"]),
                total_texto=_money_text(f["total"]),
                propinas=float(f["propinas"]),
                propinas_texto=_money_text(f["propinas"]) if f["propinas"] > 0 else "",
            )
            for f in mozos[:10]
        ]
        max_total = max((float(f["total"]) for f in horas), default=0.0)
        self.reporte_horas = [
            FranjaHoraView(
                hora_label=f"{f['hora']:02d}:00",
                pedidos=f["pedidos"],
                total=float(f["total"]),
                total_texto=_money_text(f["total"]),
                barra_pct=int(float(f["total"]) / max_total * 100) if max_total > 0 else 0,
            )
            for f in horas
        ]
        vistas_margen: list[MargenPlatoView] = []
        for f in margenes:
            pct = f["margen_pct"]
            if not f["costo_completo"]:
                color = "#94A3B8"
            elif pct < 30:
                color = "#DC2626"
            elif pct < 60:
                color = "#D97706"
            else:
                color = "#16A34A"
            vistas_margen.append(MargenPlatoView(
                nombre=f["nombre"],
                precio_texto=_money_text(f["precio"]),
                costo_texto=_money_text(f["costo"]),
                margen_texto=_money_text(f["margen"]),
                margen_pct_texto=f"{pct:.1f}%",
                color=color,
                costo_completo=f["costo_completo"],
            ))
        self.reporte_margen = vistas_margen
        labels = {"efectivo": "Efectivo", "tarjeta": "Tarjeta", "qr": "QR / Yape", "fiado": "Fiado"}
        self.reporte_metodos = [
            {"metodo": labels.get(m, m.title()), "total": v["total"], "count": v["count"]}
            for m, v in sorted(pagos_por_metodo.items(), key=lambda x: -float(x[1]["total"]))
        ]
        self._cargar_comparativa()

    def _cargar_comparativa(self) -> None:
        desde, hasta = self._rango_filtros_historial()
        if desde is None:
            desde = datetime.combine(_utcnow().date(), datetime.min.time())
        if hasta is None:
            hasta = datetime.combine(_utcnow().date(), datetime.max.time())
        delta = hasta - desde
        anterior_hasta = desde - timedelta(seconds=1)
        anterior_desde = anterior_hasta - delta

        dias_actual = max(delta.days, 1)
        if dias_actual <= 1:
            self.comp_label_actual = "Hoy"
            self.comp_label_anterior = "Ayer"
        elif dias_actual <= 7:
            self.comp_label_actual = "Esta semana"
            self.comp_label_anterior = "Semana anterior"
        else:
            self.comp_label_actual = f"Últimos {dias_actual} días"
            self.comp_label_anterior = f"{dias_actual} días previos"

        cobrado_filter = or_(
            Pedido.pagado.is_(True),
            Pedido.estado == EstadoPedido.COBRADO.value,
        )
        with self._tenant_session() as session:
            pedidos_act = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    cobrado_filter,
                    Pedido.cerrado_en >= desde,
                    Pedido.cerrado_en <= hasta,
                )
            ).all()
            pedidos_ant = session.exec(
                select(Pedido).where(
                    Pedido.company_id == self._company_id(),
                    cobrado_filter,
                    Pedido.cerrado_en >= anterior_desde,
                    Pedido.cerrado_en <= anterior_hasta,
                )
            ).all()

        ventas_act = sum(_to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0)) for p in pedidos_act)
        ventas_ant = sum(_to_decimal(p.total) + _to_decimal(getattr(p, "propina", 0)) for p in pedidos_ant)
        count_act = len(pedidos_act)
        count_ant = len(pedidos_ant)
        ticket_act = ventas_act / count_act if count_act else Decimal("0.00")
        ticket_ant = ventas_ant / count_ant if count_ant else Decimal("0.00")

        def _pct(a: Decimal, b: Decimal) -> int:
            if b <= 0:
                return 100 if a > 0 else 0
            return int(round((a - b) / b * 100))

        self.comp_ventas_actual = _money_text(ventas_act)
        self.comp_ventas_anterior = _money_text(ventas_ant)
        self.comp_ventas_pct = _pct(ventas_act, ventas_ant)
        self.comp_pedidos_actual = count_act
        self.comp_pedidos_anterior = count_ant
        self.comp_pedidos_diff = count_act - count_ant
        self.comp_ticket_actual = _money_text(ticket_act)
        self.comp_ticket_anterior = _money_text(ticket_ant)
        self.comp_ticket_pct = _pct(ticket_act, ticket_ant)

    def buscar_historial_manual(self) -> None:
        self.historial_filtro_rapido = "personalizado"
        self.aplicar_filtros_historial()

    def limpiar_filtros_historial(self) -> None:
        self.historial_filtro_fecha_desde = ""
        self.historial_filtro_fecha_hasta = ""
        self.historial_filtro_metodo = ""
        self.historial_filtro_rapido = ""
        self.historial_pagina = 0
        self.cargar_historial_ventas()
        self.cargar_analitica()

    def filtro_rapido_hoy(self) -> None:
        hoy = _utcnow().date().isoformat()
        self.historial_filtro_fecha_desde = hoy
        self.historial_filtro_fecha_hasta = hoy
        self.historial_filtro_rapido = "hoy"
        self.aplicar_filtros_historial()

    def filtro_rapido_semana(self) -> None:
        hoy = _utcnow().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        self.historial_filtro_fecha_desde = inicio_semana.isoformat()
        self.historial_filtro_fecha_hasta = hoy.isoformat()
        self.historial_filtro_rapido = "semana"
        self.aplicar_filtros_historial()

    def filtro_rapido_mes(self) -> None:
        hoy = _utcnow().date()
        inicio_mes = hoy.replace(day=1)
        self.historial_filtro_fecha_desde = inicio_mes.isoformat()
        self.historial_filtro_fecha_hasta = hoy.isoformat()
        self.historial_filtro_rapido = "mes"
        self.aplicar_filtros_historial()

    def abrir_detalle_venta(self, pedido_id: int) -> None:
        venta = next((v for v in self.historial_ventas if v.pedido_id == pedido_id), None)
        if venta is None:
            return
        with self._tenant_session() as session:
            detalles = session.exec(
                select(DetallePedido).where(
                    DetallePedido.company_id == self._company_id(),
                    DetallePedido.pedido_id == pedido_id,
                )
            ).all()
            productos = {
                p.id: p for p in session.exec(
                    select(Producto).where(Producto.company_id == self._company_id())
                ).all()
            }
        self.venta_detalle_items = [
            VentaDetalleItemView(
                nombre=productos[d.producto_id].nombre if d.producto_id in productos else f"Producto {d.producto_id}",
                cantidad=d.cantidad,
                precio_unitario_texto=_money_text(d.precio_unitario),
                subtotal_texto=_money_text(d.subtotal),
                notas=d.notas or "",
            )
            for d in detalles
        ]
        self.venta_detalle_pedido_id = pedido_id
        self.venta_detalle_mesa_label = venta.mesa_label
        self.venta_detalle_metodo = venta.metodo_pago
        self.venta_detalle_mozo = venta.mozo_nombre
        self.venta_detalle_cajero = venta.cajero_nombre
        self.venta_detalle_total_texto = venta.total_con_propina_texto
        self.venta_detalle_propina_texto = venta.propina_texto
        self.venta_detalle_visible = True

    def set_venta_detalle_visible(self, v: bool) -> None:
        self.venta_detalle_visible = v

    def historial_pagina_anterior(self) -> None:
        if self.historial_pagina > 0:
            self.historial_pagina -= 1
            self.cargar_historial_ventas()

    def historial_pagina_siguiente(self) -> None:
        if self.historial_tiene_siguiente:
            self.historial_pagina += 1
            self.cargar_historial_ventas()

    def exportar_ventas_excel(self):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        with self._tenant_session() as session:
            query = (
                select(Pedido)
                .where(
                    Pedido.company_id == self._company_id(),
                    or_(
                        Pedido.pagado.is_(True),
                        Pedido.estado == EstadoPedido.COBRADO.value,
                    ),
                )
            )
            if self.historial_filtro_fecha_desde:
                try:
                    desde = datetime.strptime(self.historial_filtro_fecha_desde, "%Y-%m-%d")
                    query = query.where(Pedido.cerrado_en >= desde)
                except ValueError:
                    pass
            if self.historial_filtro_fecha_hasta:
                try:
                    hasta = datetime.strptime(self.historial_filtro_fecha_hasta, "%Y-%m-%d")
                    hasta = hasta.replace(hour=23, minute=59, second=59)
                    query = query.where(Pedido.cerrado_en <= hasta)
                except ValueError:
                    pass
            if self.historial_filtro_metodo:
                from app.models.food import PagoPedido as _PP
                query = query.where(
                    or_(
                        Pedido.metodo_pago == self.historial_filtro_metodo,
                        Pedido.id.in_(
                            select(_PP.pedido_id).where(
                                _PP.company_id == self._company_id(),
                                _PP.metodo == self.historial_filtro_metodo,
                            )
                        ),
                    )
                )
            query = query.order_by(Pedido.cerrado_en.desc(), Pedido.id.desc())
            pedidos = session.exec(query).all()
            mesas = {m.id: m for m in session.exec(
                select(Mesa).where(Mesa.company_id == self._company_id())
            ).all()}
            usuarios = {u.id: u for u in session.exec(
                select(UsuarioFood).where(UsuarioFood.company_id == self._company_id())
            ).all()}
            pedido_ids = [p.id for p in pedidos if p.id is not None]
            detalles_por_pedido: dict[int, list] = {}
            productos_map: dict[int, Producto] = {}
            if pedido_ids:
                detalles = session.exec(
                    select(DetallePedido).where(DetallePedido.pedido_id.in_(pedido_ids))
                ).all()
                for d in detalles:
                    detalles_por_pedido.setdefault(d.pedido_id, []).append(d)
                productos_map = {
                    pr.id: pr for pr in session.exec(
                        select(Producto).where(Producto.company_id == self._company_id())
                    ).all()
                }

        if not pedidos:
            self.mensaje = "No hay ventas para exportar con estos filtros."
            return None

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Ventas"
        ws1.append([
            "Fecha", "Hora", "Pedido #", "Mesa", "Método de pago",
            "Mozo", "Cajero", "Subtotal", "Propina", "Total",
        ])
        for p in pedidos:
            fecha = p.cerrado_en
            subtotal = float(_to_decimal(p.total))
            propina = float(_to_decimal(getattr(p, "propina", 0)))
            mozo = usuarios.get(p.mozo_id)
            cajero = usuarios.get(p.cajero_id)
            ws1.append([
                fecha.strftime("%Y-%m-%d") if fecha else "",
                fecha.strftime("%H:%M") if fecha else "",
                p.id,
                _pedido_sales_label(p, mesas),
                getattr(p, "metodo_pago", None) or "",
                _actor_name(mozo.nombre) if mozo else "Sin asignar",
                _actor_name(cajero.nombre) if cajero else "Sin asignar",
                subtotal, propina, subtotal + propina,
            ])

        ws2 = wb.create_sheet("Detalle de items")
        ws2.append(["Fecha", "Pedido #", "Mesa", "Producto", "Cantidad",
                     "Precio unitario", "Subtotal"])
        for p in pedidos:
            fecha = p.cerrado_en
            mesa_label = _pedido_sales_label(p, mesas)
            for d in detalles_por_pedido.get(p.id or 0, []):
                prod = productos_map.get(d.producto_id)
                ws2.append([
                    fecha.strftime("%Y-%m-%d") if fecha else "",
                    p.id,
                    mesa_label,
                    prod.nombre if prod else f"Producto {d.producto_id}",
                    d.cantidad,
                    float(_to_decimal(d.precio_unitario)),
                    float(_to_decimal(d.subtotal)),
                ])

        for ws in (ws1, ws2):
            for i, col in enumerate(ws.columns, start=1):
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"ventas_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return rx.download(data=buf.getvalue(), filename=filename)

    # ─── ADM-01: P&L mensual ────────────────────────────────────────────────

    def cargar_pyl(self) -> None:
        now = _utcnow()
        anio = self.pyl_anio or now.year
        mes = self.pyl_mes or now.month
        self.pyl_anio = anio
        self.pyl_mes = mes

        with self._tenant_session() as session:
            data = pyl_mensual(session, self._company_id(), anio, mes)

        meses_nombre = [
            "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]
        self.pyl_lineas = [
            PylLineView(
                concepto="Ventas brutas",
                valor_texto=_money_text(data["ventas_brutas"]),
            ),
            PylLineView(
                concepto="Descuentos otorgados",
                valor_texto=f"- {_money_text(data['descuentos'])}",
                es_negativo=True,
            ),
            PylLineView(
                concepto="Ventas netas",
                valor_texto=_money_text(data["ventas_netas"]),
                es_total=True,
            ),
            PylLineView(
                concepto="Costo de insumos consumidos",
                valor_texto=f"- {_money_text(data['costo_consumo'])}",
                es_negativo=True,
            ),
            PylLineView(
                concepto="Egresos de caja",
                valor_texto=f"- {_money_text(data['egresos_caja'])}",
                es_negativo=True,
            ),
            PylLineView(
                concepto="Utilidad bruta operativa",
                valor_texto=_money_text(data["utilidad"]),
                es_total=True,
                es_negativo=data["utilidad"] < 0,
                margen_pct_texto=f"{data['margen_pct']}%",
            ),
        ]

    def set_pyl_anio(self, v: str) -> None:
        try:
            self.pyl_anio = int(v)
        except (ValueError, TypeError):
            pass

    def set_pyl_mes(self, v: str) -> None:
        try:
            self.pyl_mes = int(v)
        except (ValueError, TypeError):
            pass

    def actualizar_pyl(self) -> None:
        self.cargar_pyl()

    # ─── ADM-02: Descuentos y anulaciones ────────────────────────────────────

    def cargar_descuentos_anulaciones(self) -> None:
        desde, hasta = self._rango_filtros_historial()
        if hasta is not None:
            hasta = hasta + timedelta(seconds=1)

        with self._tenant_session() as session:
            desc_data = reporte_descuentos(session, self._company_id(), desde, hasta)
            anul_data = reporte_anulaciones(session, self._company_id(), desde, hasta)

        self.descuentos_rank = [
            DescuentoRankView(
                cajero=d["cajero"],
                pedidos=d["pedidos"],
                total_descuento_texto=_money_text(d["total_descuento"]),
                total_ventas_texto=_money_text(d["total_ventas"]),
                pct_descuento_texto=f"{d['pct_descuento']}%",
            )
            for d in desc_data
        ]
        total_desc = sum(d["total_descuento"] for d in desc_data)
        self.descuentos_total_texto = _money_text(total_desc)

        from tuwayki_core.utils.timezone import format_local_datetime
        self.anulaciones_lista = [
            AnulacionView(
                pedido_id=a["pedido_id"],
                total_texto=_money_text(a["total"]),
                motivo=a["motivo"],
                cancelado_por=a["cancelado_por"],
                cancelado_en_texto=format_local_datetime(a["cancelado_en"], "%d/%m %H:%M", "PE") if a["cancelado_en"] else "",
                cajero_original=a["cajero_original"],
            )
            for a in anul_data
        ]
        total_anul = sum(a["total"] for a in anul_data)
        self.anulaciones_total_texto = _money_text(total_anul)

    # ─── ADM-03: Mermas valorizado ───────────────────────────────────────────

    def cargar_mermas(self) -> None:
        desde, hasta = self._rango_filtros_historial()
        if hasta is not None:
            hasta = hasta + timedelta(seconds=1)

        with self._tenant_session() as session:
            data = reporte_mermas(session, self._company_id(), desde, hasta)

        self.mermas_por_categoria = [
            MermaCategoriaView(
                categoria=c["categoria"],
                registros=c["registros"],
                valor_texto=_money_text(c["valor"]),
            )
            for c in data["por_categoria"]
        ]
        self.mermas_por_insumo = [
            MermaInsumoView(
                nombre=i["nombre"],
                unidad=i["unidad"],
                cantidad_texto=f"{i['cantidad_total']:.3f}",
                valor_texto=_money_text(i["valor"]),
                registros=i["registros"],
            )
            for i in data["por_insumo"][:20]
        ]
        self.mermas_total_texto = _money_text(data["total"])
